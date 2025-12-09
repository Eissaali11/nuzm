"""
Ù„ÙˆØ­Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Power BI Ø§Ø­ØªØ±Ø§ÙÙŠØ© - Ù†ÙØ¸Ù…
ØªØ­Ù„ÙŠÙ„Ø§Øª Ù…ØªÙ‚Ø¯Ù…Ø© Ù„Ù„Ø­Ø¶ÙˆØ± ÙˆØ§Ù„ÙˆØ«Ø§Ø¦Ù‚ ÙˆØ§Ù„Ø³ÙŠØ§Ø±Ø§Øª
ØªØµØ¯ÙŠØ± Excel Ø¨ØªØµÙ…ÙŠÙ… Ø§Ø­ØªØ±Ø§ÙÙŠ
"""
from flask import Blueprint, render_template, request, jsonify, Response, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from app import db
from models import Employee, Attendance, Document, Vehicle, Department
from sqlalchemy import func, or_, and_, case
from utils.user_helpers import require_module_access
from models import Module, Permission
import csv
from io import StringIO, BytesIO
import json

powerbi_bp = Blueprint('powerbi', __name__, url_prefix='/powerbi')

@powerbi_bp.route('/')
@login_required
def dashboard():
    """Ø§Ù„ØµÙØ­Ø© Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠØ© Ù„Ù„ÙˆØ­Ø© Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Power BI Ø§Ù„Ø§Ø­ØªØ±Ø§ÙÙŠØ© - Ø¨ÙŠØ§Ù†Ø§Øª Ø­Ù‚ÙŠÙ‚ÙŠØ©"""
    from datetime import datetime, timedelta
    
    # Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ©
    departments = Department.query.all()
    total_vehicles = Vehicle.query.count()
    total_documents = Document.query.count()
    
    # ÙÙ„Ø§ØªØ± Ø§Ù„ØªØ§Ø±ÙŠØ®
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except:
            date_from = datetime.now().date() - timedelta(days=30)
    else:
        date_from = datetime.now().date() - timedelta(days=30)
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except:
            date_to = datetime.now().date()
    else:
        date_to = datetime.now().date()
    
    # Ø¬Ù„Ø¨ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø· Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ± ÙÙŠ Ø§Ù„ÙØªØ±Ø© Ø§Ù„Ù…Ø­Ø¯Ø¯Ø©
    active_employee_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
        Attendance.date >= date_from,
        Attendance.date <= date_to
    ).distinct().all()
    active_employee_ids_with_attendance = [e[0] for e in active_employee_ids_with_attendance]
    
    # Ø¹Ø¯Ø¯ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ±
    active_employees_count = Employee.query.filter(
        Employee.status == 'active',
        Employee.id.in_(active_employee_ids_with_attendance)
    ).count()
    
    total_employees = active_employees_count
    
    # Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ø­Ø¶ÙˆØ± Ù„Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø·
    active_emp_ids = [e.id for e in Employee.query.filter(Employee.status == 'active').all()]
    
    attendance_records = Attendance.query.filter(
        Attendance.date >= date_from,
        Attendance.date <= date_to,
        Attendance.employee_id.in_(active_emp_ids)
    ).all()
    
    attendance_stats = {
        'present': sum(1 for a in attendance_records if a.status == 'present'),
        'absent': sum(1 for a in attendance_records if a.status in ['absent', 'ØºØ§Ø¦Ø¨']),
        'leave': sum(1 for a in attendance_records if a.status == 'leave'),
        'sick': sum(1 for a in attendance_records if a.status == 'sick'),
        'total': len(attendance_records)
    }
    attendance_stats['rate'] = round((attendance_stats['present'] / attendance_stats['total']) * 100, 1) if attendance_stats['total'] > 0 else 0
    
    # Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª
    vehicles = Vehicle.query.all()
    vehicle_stats = {}
    for v in vehicles:
        status = v.status or 'unknown'
        vehicle_stats[status] = vehicle_stats.get(status, 0) + 1
    
    # Ø§Ù„Ø­Ø¶ÙˆØ± Ø­Ø³Ø¨ Ø§Ù„Ù‚Ø³Ù… - Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø· Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ±
    dept_attendance = []
    for dept in departments:
        # Ø¬Ù„Ø¨ Ù…ÙˆØ¸ÙÙŠ Ø§Ù„Ù‚Ø³Ù… Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø·
        emp_ids = db.session.query(Employee.id).join(
            Employee.departments
        ).filter(
            Department.id == dept.id,
            Employee.status == 'active'
        ).all()
        emp_ids = [e[0] for e in emp_ids]
        
        if not emp_ids:
            continue
        
        # Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ± ÙØ¹Ù„ÙŠ ÙÙŠ Ø§Ù„ÙØªØ±Ø©
        emp_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to,
            Attendance.employee_id.in_(emp_ids)
        ).distinct().all()
        emp_ids_with_attendance = [e[0] for e in emp_ids_with_attendance]
        
        if not emp_ids_with_attendance:
            continue
            
        # Ø­Ø³Ø§Ø¨ Ø§Ù„Ø­Ø¶ÙˆØ±
        dept_records = Attendance.query.filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to,
            Attendance.employee_id.in_(emp_ids_with_attendance)
        ).all()
        
        present = sum(1 for a in dept_records if a.status == 'present')
        total = len(dept_records)
        rate = round((present / total) * 100, 1) if total > 0 else 0
        
        dept_attendance.append({
            'name': dept.name,
            'employee_count': len(emp_ids_with_attendance),
            'present': present,
            'total': total,
            'rate': rate
        })
    
    # ØªØ±ØªÙŠØ¨ Ø­Ø³Ø¨ Ù†Ø³Ø¨Ø© Ø§Ù„Ø­Ø¶ÙˆØ±
    dept_attendance.sort(key=lambda x: x['rate'], reverse=True)
    
    # Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚
    today = datetime.now().date()
    thirty_days = today + timedelta(days=30)
    
    docs = Document.query.all()
    doc_stats = {
        'valid': 0,
        'expiring': 0,
        'expired': 0,
        'total': len(docs)
    }
    
    for doc in docs:
        if hasattr(doc, 'expiry_date') and doc.expiry_date:
            if doc.expiry_date < today:
                doc_stats['expired'] += 1
            elif doc.expiry_date <= thirty_days:
                doc_stats['expiring'] += 1
            else:
                doc_stats['valid'] += 1
        else:
            doc_stats['valid'] += 1
    
    return render_template('powerbi/dashboard.html',
        departments=departments,
        total_employees=total_employees,
        total_vehicles=total_vehicles,
        total_documents=total_documents,
        attendance_stats=attendance_stats,
        vehicle_stats=vehicle_stats,
        dept_attendance=dept_attendance,
        doc_stats=doc_stats,
        date_from=date_from,
        date_to=date_to
    )

@powerbi_bp.route('/api/attendance-summary')
@login_required
@require_module_access(Module.ATTENDANCE, Permission.VIEW)
def attendance_summary():
    """Ù…Ù„Ø®Øµ Ø§Ù„Ø­Ø¶ÙˆØ± Ù…Ø¹ ØªØ­Ù„ÙŠÙ„Ø§Øª Ù…ØªÙ‚Ø¯Ù…Ø©"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date()
        
        query = Attendance.query.filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to
        )
        
        if department_id:
            employee_ids = [e.id for e in Employee.query.filter_by(department_id=department_id).all()]
            query = query.filter(Attendance.employee_id.in_(employee_ids))
        
        attendance_records = query.all()
        
        present = sum(1 for a in attendance_records if a.status == 'present')
        absent = sum(1 for a in attendance_records if a.status in ['absent', 'ØºØ§Ø¦Ø¨'])
        leave = sum(1 for a in attendance_records if a.status == 'leave')
        sick = sum(1 for a in attendance_records if a.status == 'sick')
        total = len(attendance_records)
        
        attendance_rate = round((present / total) * 100, 1) if total > 0 else 0
        absence_rate = round((absent / total) * 100, 1) if total > 0 else 0
        
        prev_date_from = date_from - timedelta(days=30)
        prev_query = Attendance.query.filter(
            Attendance.date >= prev_date_from,
            Attendance.date < date_from
        )
        if department_id:
            prev_query = prev_query.filter(Attendance.employee_id.in_(employee_ids))
        
        prev_records = prev_query.all()
        prev_present = sum(1 for a in prev_records if a.status == 'present')
        prev_total = len(prev_records)
        prev_rate = round((prev_present / prev_total) * 100, 1) if prev_total > 0 else 0
        
        trend = round(attendance_rate - prev_rate, 1)
        trend_direction = 'up' if trend > 0 else 'down' if trend < 0 else 'stable'
        
        return jsonify({
            'success': True,
            'data': {
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'total': total,
                'attendance_rate': attendance_rate,
                'absence_rate': absence_rate,
                'trend': {
                    'value': abs(trend),
                    'direction': trend_direction,
                    'previous_rate': prev_rate
                },
                'date_range': {
                    'from': date_from.strftime('%Y-%m-%d'),
                    'to': date_to.strftime('%Y-%m-%d')
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/attendance-by-department')
@login_required
@require_module_access(Module.ATTENDANCE, Permission.VIEW)
def attendance_by_department():
    """Ø§Ù„Ø­Ø¶ÙˆØ± Ø­Ø³Ø¨ Ø§Ù„Ù‚Ø³Ù… Ù…Ø¹ ØªØ­Ù„ÙŠÙ„ Ù…ÙØµÙ„"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date()
        
        departments_data = []
        departments = Department.query.all()
        
        for dept in departments:
            employees = Employee.query.filter_by(department_id=dept.id).all()
            employee_ids = [e.id for e in employees]
            
            if not employee_ids:
                continue
            
            attendance_records = Attendance.query.filter(
                Attendance.date >= date_from,
                Attendance.date <= date_to,
                Attendance.employee_id.in_(employee_ids)
            ).all()
            
            present = sum(1 for a in attendance_records if a.status == 'present')
            absent = sum(1 for a in attendance_records if a.status in ['absent', 'ØºØ§Ø¦Ø¨'])
            leave = sum(1 for a in attendance_records if a.status == 'leave')
            sick = sum(1 for a in attendance_records if a.status == 'sick')
            total = len(attendance_records)
            
            attendance_rate = round((present / total) * 100, 1) if total > 0 else 0
            
            performance = 'excellent' if attendance_rate >= 90 else 'good' if attendance_rate >= 75 else 'average' if attendance_rate >= 60 else 'poor'
            
            departments_data.append({
                'department': dept.name,
                'department_id': dept.id,
                'employee_count': len(employees),
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'total': total,
                'attendance_rate': attendance_rate,
                'performance': performance
            })
        
        departments_data.sort(key=lambda x: x['attendance_rate'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': departments_data,
            'summary': {
                'total_departments': len(departments_data),
                'avg_attendance_rate': round(sum(d['attendance_rate'] for d in departments_data) / len(departments_data), 1) if departments_data else 0,
                'best_department': departments_data[0]['department'] if departments_data else None,
                'worst_department': departments_data[-1]['department'] if departments_data else None
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/documents-status')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def documents_status():
    """Ø­Ø§Ù„Ø© Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø© Ù…Ø¹ ØªØ­Ù„ÙŠÙ„ Ø´Ø§Ù…Ù„"""
    try:
        required_docs = [
            {'type': 'Ø§Ù„Ù‡ÙˆÙŠØ© Ø§Ù„ÙˆØ·Ù†ÙŠØ©', 'priority': 'high'},
            {'type': 'Ø¬ÙˆØ§Ø² Ø§Ù„Ø³ÙØ±', 'priority': 'high'},
            {'type': 'Ø±Ø®ØµØ© Ø§Ù„Ù‚ÙŠØ§Ø¯Ø©', 'priority': 'medium'},
            {'type': 'ÙˆØ«ÙŠÙ‚Ø© Ø§Ù„ØªØ£Ù…ÙŠÙ†', 'priority': 'medium'}
        ]
        
        total_employees = Employee.query.count()
        documents_summary = []
        
        today = datetime.now().date()
        thirty_days_later = today + timedelta(days=30)
        
        for doc_info in required_docs:
            doc_type = doc_info['type']
            
            docs = Document.query.filter(
                Document.document_type.ilike(f'%{doc_type}%')
            ).all()
            
            available = len(docs)
            missing = total_employees - available
            
            expiring_soon = 0
            expired = 0
            valid = 0
            
            for doc in docs:
                if hasattr(doc, 'expiry_date') and doc.expiry_date:
                    if doc.expiry_date < today:
                        expired += 1
                    elif doc.expiry_date <= thirty_days_later:
                        expiring_soon += 1
                    else:
                        valid += 1
                else:
                    valid += 1
            
            documents_summary.append({
                'type': doc_type,
                'priority': doc_info['priority'],
                'available': available,
                'missing': missing,
                'valid': valid,
                'expiring_soon': expiring_soon,
                'expired': expired,
                'completion_rate': round((available / total_employees * 100), 1) if total_employees > 0 else 0,
                'health_score': round(((valid) / available * 100), 1) if available > 0 else 0
            })
        
        total_available = sum(d['available'] for d in documents_summary)
        total_required = total_employees * len(required_docs)
        overall_completion = round((total_available / total_required * 100), 1) if total_required > 0 else 0
        
        return jsonify({
            'success': True,
            'data': documents_summary,
            'total_employees': total_employees,
            'overall_completion': overall_completion,
            'total_expiring_soon': sum(d['expiring_soon'] for d in documents_summary),
            'total_expired': sum(d['expired'] for d in documents_summary),
            'total_missing': sum(d['missing'] for d in documents_summary)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/employee-documents')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def employee_documents():
    """Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† ÙˆØ§Ù„ÙˆØ«Ø§Ø¦Ù‚ Ø§Ù„Ù†Ø§Ù‚ØµØ© Ù…Ø¹ ØªÙØ§ØµÙŠÙ„"""
    department_id = request.args.get('department_id')
    limit = request.args.get('limit', 50, type=int)
    
    try:
        query = Employee.query
        
        if department_id:
            query = query.filter_by(department_id=department_id)
        
        employees = query.limit(limit).all()
        
        required_docs = ['Ø§Ù„Ù‡ÙˆÙŠØ© Ø§Ù„ÙˆØ·Ù†ÙŠØ©', 'Ø¬ÙˆØ§Ø² Ø§Ù„Ø³ÙØ±', 'Ø±Ø®ØµØ© Ø§Ù„Ù‚ÙŠØ§Ø¯Ø©', 'ÙˆØ«ÙŠÙ‚Ø© Ø§Ù„ØªØ£Ù…ÙŠÙ†']
        employees_data = []
        
        for emp in employees:
            docs = Document.query.filter_by(employee_id=emp.id).all()
            doc_types = [d.document_type for d in docs]
            
            missing_docs = []
            for req_doc in required_docs:
                if not any(req_doc in dt for dt in doc_types):
                    missing_docs.append(req_doc)
            
            completion_rate = round(((len(required_docs) - len(missing_docs)) / len(required_docs)) * 100, 0)
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id or '-',
                'department': emp.department.name if emp.department else 'Ø¨Ø¯ÙˆÙ† Ù‚Ø³Ù…',
                'total_docs': len(docs),
                'missing_docs': missing_docs,
                'missing_count': len(missing_docs),
                'documents_complete': len(missing_docs) == 0,
                'completion_rate': completion_rate
            })
        
        employees_data.sort(key=lambda x: x['missing_count'], reverse=True)
        
        complete_count = sum(1 for e in employees_data if e['documents_complete'])
        incomplete_count = len(employees_data) - complete_count
        
        return jsonify({
            'success': True,
            'data': employees_data,
            'summary': {
                'total': len(employees_data),
                'complete': complete_count,
                'incomplete': incomplete_count,
                'completion_rate': round((complete_count / len(employees_data)) * 100, 1) if employees_data else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/vehicles-summary')
@login_required
@require_module_access(Module.VEHICLES, Permission.VIEW)
def vehicles_summary():
    """Ù…Ù„Ø®Øµ Ø­Ø§Ù„Ø© Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª Ù…Ø¹ ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø£Ø³Ø·ÙˆÙ„"""
    try:
        vehicles = Vehicle.query.all()
        
        statuses = {}
        brands = {}
        years = {}
        
        for vehicle in vehicles:
            status = vehicle.status or 'unknown'
            statuses[status] = statuses.get(status, 0) + 1
            
            if hasattr(vehicle, 'make') and vehicle.make:
                brands[vehicle.make] = brands.get(vehicle.make, 0) + 1
            
            if hasattr(vehicle, 'year') and vehicle.year:
                years[str(vehicle.year)] = years.get(str(vehicle.year), 0) + 1
        
        vehicles_by_status = []
        status_labels = {
            'in_project': 'ÙÙŠ Ø§Ù„Ù…Ø´Ø±ÙˆØ¹',
            'in_workshop': 'ÙÙŠ Ø§Ù„ÙˆØ±Ø´Ø©',
            'out_of_service': 'Ø®Ø§Ø±Ø¬ Ø§Ù„Ø®Ø¯Ù…Ø©',
            'accident': 'Ø­Ø§Ø¯Ø«',
            'unknown': 'ØºÙŠØ± Ù…Ø­Ø¯Ø¯'
        }
        
        for status, count in statuses.items():
            vehicles_by_status.append({
                'status': status_labels.get(status, status),
                'status_key': status,
                'count': count,
                'percentage': round((count / len(vehicles) * 100), 1) if vehicles else 0
            })
        
        vehicles_by_brand = [{'brand': b, 'count': c} for b, c in sorted(brands.items(), key=lambda x: x[1], reverse=True)]
        
        total = len(vehicles)
        in_project = statuses.get('in_project', 0)
        in_workshop = statuses.get('in_workshop', 0)
        out_of_service = statuses.get('out_of_service', 0)
        accident = statuses.get('accident', 0)
        
        fleet_health = 'excellent' if in_project >= total * 0.8 else 'good' if in_project >= total * 0.6 else 'average' if in_project >= total * 0.4 else 'poor'
        
        return jsonify({
            'success': True,
            'data': {
                'by_status': vehicles_by_status,
                'by_brand': vehicles_by_brand[:5],
                'total_vehicles': total,
                'in_project': in_project,
                'in_workshop': in_workshop,
                'out_of_service': out_of_service,
                'accident': accident,
                'utilization_rate': round((in_project / total) * 100, 1) if total > 0 else 0,
                'fleet_health': fleet_health
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/vehicle-operations-summary')
@login_required
@require_module_access(Module.VEHICLES, Permission.VIEW)
def vehicle_operations_summary():
    """Ù…Ù„Ø®Øµ Ø¹Ù…Ù„ÙŠØ§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª"""
    try:
        vehicles = Vehicle.query.all()
        
        handovers = sum(1 for v in vehicles if v.handover_records)
        in_workshop = sum(1 for v in vehicles if v.status == 'in_workshop')
        in_project = sum(1 for v in vehicles if v.status == 'in_project')
        out_of_service = sum(1 for v in vehicles if v.status == 'out_of_service')
        accident = sum(1 for v in vehicles if v.status == 'accident')
        
        operations_data = [
            {'type': 'ÙÙŠ Ø§Ù„Ù…Ø´Ø±ÙˆØ¹', 'count': in_project, 'color': '#38ef7d'},
            {'type': 'Ù…Ø³ØªÙ„Ù…Ø©', 'count': handovers, 'color': '#667eea'},
            {'type': 'ÙÙŠ Ø§Ù„ÙˆØ±Ø´Ø©', 'count': in_workshop, 'color': '#fbbf24'},
            {'type': 'Ø®Ø§Ø±Ø¬ Ø§Ù„Ø®Ø¯Ù…Ø©', 'count': out_of_service, 'color': '#ef4444'},
            {'type': 'Ø­Ø§Ø¯Ø«', 'count': accident, 'color': '#ff6b6b'}
        ]
        
        return jsonify({
            'success': True,
            'data': {
                'by_type': operations_data,
                'total_vehicles': len(vehicles),
                'summary': {
                    'active_percentage': round((in_project / len(vehicles)) * 100, 1) if vehicles else 0,
                    'handover_percentage': round((handovers / len(vehicles)) * 100, 1) if vehicles else 0
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/export-data')
@login_required
def export_data():
    """ØªØµØ¯ÙŠØ± Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø¨ØµÙŠØºØ© Excel Ø§Ø­ØªØ±Ø§ÙÙŠØ© Ø¨ØªØµÙ…ÙŠÙ… Power BI - Ù†ÙØ³ ØªØµÙ…ÙŠÙ… Ø§Ù„ØµÙØ­Ø©"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
    from openpyxl.chart.label import DataLabelList
    from sqlalchemy import func
    
    data_type = request.args.get('type', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    try:
        wb = Workbook()
        
        # Ø£Ù„ÙˆØ§Ù† Ø§Ù„ØªØµÙ…ÙŠÙ… Ø§Ù„Ø¬Ø¯ÙŠØ¯ - cyan/teal theme
        dark_bg = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        card_fill = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        
        # Ø£Ù„ÙˆØ§Ù† cyan/teal
        cyan_fill = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
        teal_fill = PatternFill(start_color="0ABAB5", end_color="0ABAB5", fill_type="solid")
        
        # Ø£Ù„ÙˆØ§Ù† Ø§Ù„Ø­Ø§Ù„Ø§Øª
        green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        orange_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        blue_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        purple_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        
        # Ø®Ø·ÙˆØ·
        cyan_font = Font(bold=True, color="00D4AA", size=12)
        white_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, color="00D4AA", size=24)
        subtitle_font = Font(color="9CA3AF", size=10)
        kpi_value_font = Font(bold=True, color="00D4AA", size=28)
        kpi_label_font = Font(color="9CA3AF", size=10)
        
        # Ø­Ø¯ÙˆØ¯
        cyan_border = Border(
            left=Side(style='medium', color='00D4AA'),
            right=Side(style='medium', color='00D4AA'),
            top=Side(style='medium', color='00D4AA'),
            bottom=Side(style='medium', color='00D4AA')
        )
        thin_border = Border(
            left=Side(style='thin', color='374151'),
            right=Side(style='thin', color='374151'),
            top=Side(style='thin', color='374151'),
            bottom=Side(style='thin', color='374151')
        )
        
        # Ù…Ø¹Ø§Ù„Ø¬Ø© Ø§Ù„ØªÙˆØ§Ø±ÙŠØ®
        d_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else (datetime.now().date() - timedelta(days=30))
        d_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else datetime.now().date()
        
        # Ø¬Ù„Ø¨ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ± ÙÙŠ Ø§Ù„ÙØªØ±Ø©
        active_employee_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
            Attendance.date >= d_from,
            Attendance.date <= d_to
        ).distinct().all()
        active_employee_ids_with_attendance = [e[0] for e in active_employee_ids_with_attendance]
        
        # Ø¹Ø¯Ø¯ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ±
        total_employees = Employee.query.filter(
            Employee.status == 'active',
            Employee.id.in_(active_employee_ids_with_attendance)
        ).count()
        
        # Ø¨Ø§Ù‚ÙŠ Ø§Ù„Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª
        total_vehicles = Vehicle.query.count()
        in_project_vehicles = Vehicle.query.filter_by(status='in_project').count()
        in_workshop_vehicles = Vehicle.query.filter_by(status='in_workshop').count()
        out_of_service_vehicles = Vehicle.query.filter_by(status='out_of_service').count()
        accident_vehicles = Vehicle.query.filter_by(status='accident').count()
        total_documents = Document.query.count()
        total_departments = Department.query.count()
        
        # Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ø­Ø¶ÙˆØ± Ù„Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø·
        active_emp_ids = [e.id for e in Employee.query.filter(Employee.status == 'active').all()]
        
        attendance_records = Attendance.query.filter(
            Attendance.date >= d_from,
            Attendance.date <= d_to,
            Attendance.employee_id.in_(active_emp_ids)
        ).all()
        
        att_data = {
            'present': sum(1 for a in attendance_records if a.status == 'present'),
            'absent': sum(1 for a in attendance_records if a.status in ['absent', 'ØºØ§Ø¦Ø¨']),
            'leave': sum(1 for a in attendance_records if a.status == 'leave'),
            'sick': sum(1 for a in attendance_records if a.status == 'sick')
        }
        total_attendance = sum(att_data.values())
        
        ws = wb.active
        ws.title = "Power BI Dashboard"
        ws.sheet_view.rightToLeft = True
        
        for row in range(1, 60):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = dark_bg
        
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 14
        
        ws.merge_cells('A1:S1')
        ws['A1'] = "Ù†ÙÙ€Ø¸Ù€Ù… | Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­Ù„ÙŠÙ„Ø§Øª Ø§Ù„Ø§Ø­ØªØ±Ø§ÙÙŠØ©"
        ws['A1'].font = title_font
        ws['A1'].fill = dark_bg
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 50
        
        ws.merge_cells('A2:S2')
        ws['A2'] = f"ØªØ§Ø±ÙŠØ® Ø§Ù„ØªÙ‚Ø±ÙŠØ±: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Ø§Ù„ÙØªØ±Ø©: {d_from} Ø¥Ù„Ù‰ {d_to}"
        ws['A2'].font = subtitle_font
        ws['A2'].fill = dark_bg
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A3:S3')
        ws['A3'] = f"Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙŠ Ø§Ù„ÙØªØ±Ø© Ø§Ù„Ù…Ø­Ø¯Ø¯Ø© | Active Employees in Selected Period"
        ws['A3'].font = Font(color="9CA3AF", size=9)
        ws['A3'].fill = dark_bg
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        kpi_data = [
            ("Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†", total_employees, "ğŸ‘¥", "A"),
            ("Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª", total_vehicles, "ğŸš—", "E"),
            ("ÙÙŠ Ø§Ù„Ù…Ø´Ø§Ø±ÙŠØ¹", in_project_vehicles, "âœ…", "I"),
            ("Ø§Ù„Ø£Ù‚Ø³Ø§Ù…", total_departments, "ğŸ¢", "M"),
        ]
        
        for label, value, icon, start_col in kpi_data:
            col_idx = ord(start_col) - ord('A') + 1
            end_col = get_column_letter(col_idx + 2)
            
            ws.merge_cells(f'{start_col}5:{end_col}5')
            ws[f'{start_col}5'] = f"{icon} {label}"
            ws[f'{start_col}5'].font = kpi_label_font
            ws[f'{start_col}5'].fill = card_fill
            ws[f'{start_col}5'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{start_col}5'].border = cyan_border
            
            ws.merge_cells(f'{start_col}6:{end_col}6')
            ws[f'{start_col}6'] = value
            ws[f'{start_col}6'].font = kpi_value_font
            ws[f'{start_col}6'].fill = card_fill
            ws[f'{start_col}6'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{start_col}6'].border = cyan_border
            
            ws.row_dimensions[5].height = 25
            ws.row_dimensions[6].height = 40
        
        ws.merge_cells('A8:H8')
        ws['A8'] = "ğŸ“Š ØªÙˆØ²ÙŠØ¹ Ø­Ø§Ù„Ø§Øª Ø§Ù„Ø­Ø¶ÙˆØ± - Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†"
        ws['A8'].font = cyan_font
        ws['A8'].fill = header_fill
        ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[8].height = 30
        
        att_headers = ['Ø§Ù„Ø­Ø§Ù„Ø©', 'Ø§Ù„Ø¹Ø¯Ø¯', 'Ø§Ù„Ù†Ø³Ø¨Ø©', 'Ø§Ù„Ø±Ø³Ù…']
        for i, h in enumerate(att_headers):
            cell = ws.cell(row=9, column=i+1)
            cell.value = h
            cell.font = cyan_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        att_rows = [
            ('Ø­Ø§Ø¶Ø± âœ…', att_data['present'], green_fill),
            ('ØºØ§Ø¦Ø¨ âŒ', att_data['absent'], red_fill),
            ('Ø¥Ø¬Ø§Ø²Ø© ğŸ“‹', att_data['leave'], blue_fill),
            ('Ù…Ø±ÙŠØ¶ ğŸ¥', att_data['sick'], orange_fill)
        ]
        
        for idx, (label, count, fill) in enumerate(att_rows, start=10):
            pct = round((count / total_attendance * 100), 1) if total_attendance > 0 else 0
            
            ws.cell(row=idx, column=1).value = label
            ws.cell(row=idx, column=1).font = white_font
            ws.cell(row=idx, column=1).fill = card_fill
            ws.cell(row=idx, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=1).border = thin_border
            
            ws.cell(row=idx, column=2).value = count
            ws.cell(row=idx, column=2).font = white_font
            ws.cell(row=idx, column=2).fill = card_fill
            ws.cell(row=idx, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=2).border = thin_border
            
            ws.cell(row=idx, column=3).value = f"{pct}%"
            ws.cell(row=idx, column=3).font = white_font
            ws.cell(row=idx, column=3).fill = card_fill
            ws.cell(row=idx, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=3).border = thin_border
            
            bar_width = int(pct / 5) if pct > 0 else 1
            ws.cell(row=idx, column=4).value = "â–ˆ" * bar_width
            ws.cell(row=idx, column=4).font = Font(color=fill.fgColor.rgb[2:] if fill.fgColor else "FFFFFF", size=10)
            ws.cell(row=idx, column=4).fill = card_fill
            ws.cell(row=idx, column=4).border = thin_border
        
        pie1 = PieChart()
        pie1.title = "ØªÙˆØ²ÙŠØ¹ Ø§Ù„Ø­Ø¶ÙˆØ±"
        labels1 = Reference(ws, min_col=1, min_row=10, max_row=13)
        data1 = Reference(ws, min_col=2, min_row=9, max_row=13)
        pie1.add_data(data1, titles_from_data=True)
        pie1.set_categories(labels1)
        pie1.width = 10
        pie1.height = 8
        pie1.dataLabels = DataLabelList()
        pie1.dataLabels.showPercent = True
        pie1.dataLabels.showCatName = True
        ws.add_chart(pie1, "E9")
        
        ws.merge_cells('A16:H16')
        ws['A16'] = "ğŸš— Ø­Ø§Ù„Ø© Ø£Ø³Ø·ÙˆÙ„ Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª"
        ws['A16'].font = cyan_font
        ws['A16'].fill = header_fill
        ws['A16'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[16].height = 30
        
        veh_headers = ['Ø§Ù„Ø­Ø§Ù„Ø©', 'Ø§Ù„Ø¹Ø¯Ø¯', 'Ø§Ù„Ù†Ø³Ø¨Ø©', 'Ø§Ù„Ø±Ø³Ù…']
        for i, h in enumerate(veh_headers):
            cell = ws.cell(row=17, column=i+1)
            cell.value = h
            cell.font = cyan_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        veh_rows = [
            ('ÙÙŠ Ø§Ù„Ù…Ø´Ø±ÙˆØ¹ ğŸŸ¢', in_project_vehicles, green_fill),
            ('ÙÙŠ Ø§Ù„ÙˆØ±Ø´Ø© ğŸŸ¡', in_workshop_vehicles, orange_fill),
            ('Ø®Ø§Ø±Ø¬ Ø§Ù„Ø®Ø¯Ù…Ø© ğŸ”´', out_of_service_vehicles, red_fill),
            ('Ø­Ø§Ø¯Ø« âš ï¸', accident_vehicles, purple_fill)
        ]
        
        for idx, (label, count, fill) in enumerate(veh_rows, start=18):
            pct = round((count / total_vehicles * 100), 1) if total_vehicles > 0 else 0
            
            ws.cell(row=idx, column=1).value = label
            ws.cell(row=idx, column=1).font = white_font
            ws.cell(row=idx, column=1).fill = card_fill
            ws.cell(row=idx, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=1).border = thin_border
            
            ws.cell(row=idx, column=2).value = count
            ws.cell(row=idx, column=2).font = white_font
            ws.cell(row=idx, column=2).fill = card_fill
            ws.cell(row=idx, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=2).border = thin_border
            
            ws.cell(row=idx, column=3).value = f"{pct}%"
            ws.cell(row=idx, column=3).font = white_font
            ws.cell(row=idx, column=3).fill = card_fill
            ws.cell(row=idx, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=3).border = thin_border
            
            bar_width = int(pct / 5) if pct > 0 else 1
            ws.cell(row=idx, column=4).value = "â–ˆ" * bar_width
            ws.cell(row=idx, column=4).font = Font(color=fill.fgColor.rgb[2:] if fill.fgColor else "FFFFFF", size=10)
            ws.cell(row=idx, column=4).fill = card_fill
            ws.cell(row=idx, column=4).border = thin_border
        
        doughnut1 = DoughnutChart()
        doughnut1.title = "Ø­Ø§Ù„Ø© Ø§Ù„Ø£Ø³Ø·ÙˆÙ„"
        labels2 = Reference(ws, min_col=1, min_row=18, max_row=21)
        data2 = Reference(ws, min_col=2, min_row=17, max_row=21)
        doughnut1.add_data(data2, titles_from_data=True)
        doughnut1.set_categories(labels2)
        doughnut1.width = 10
        doughnut1.height = 8
        doughnut1.dataLabels = DataLabelList()
        doughnut1.dataLabels.showPercent = True
        ws.add_chart(doughnut1, "E16")
        
        # Ø¬Ù„Ø¨ Ø§Ù„Ø£Ù‚Ø³Ø§Ù… Ù…Ø¹ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø· Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ±
        departments = Department.query.all()
        
        ws.merge_cells('A24:H24')
        ws['A24'] = "ğŸ¢ Ù†Ø³Ø¨Ø© Ø§Ù„Ø­Ø¶ÙˆØ± Ø­Ø³Ø¨ Ø§Ù„Ù‚Ø³Ù… - Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†"
        ws['A24'].font = cyan_font
        ws['A24'].fill = header_fill
        ws['A24'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[24].height = 30
        
        dept_headers = ['Ø§Ù„Ù‚Ø³Ù…', 'Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†', 'Ø§Ù„Ø­Ø¶ÙˆØ±', 'Ø§Ù„Ù†Ø³Ø¨Ø©', 'Ø§Ù„ØªÙ‚ÙŠÙŠÙ…']
        for i, h in enumerate(dept_headers):
            cell = ws.cell(row=25, column=i+1)
            cell.value = h
            cell.font = cyan_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        dept_row = 26
        for dept in departments[:10]:
            # Ø¬Ù„Ø¨ Ù…ÙˆØ¸ÙÙŠ Ø§Ù„Ù‚Ø³Ù… Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø·
            emp_ids = db.session.query(Employee.id).join(
                Employee.departments
            ).filter(
                Department.id == dept.id,
                Employee.status == 'active'
            ).all()
            emp_ids = [e[0] for e in emp_ids]
            
            if not emp_ids:
                continue
            
            # Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ø°ÙŠÙ† Ù„Ù‡Ù… Ø­Ø¶ÙˆØ± ÙØ¹Ù„ÙŠ ÙÙŠ Ø§Ù„ÙØªØ±Ø©
            emp_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
                Attendance.date >= d_from,
                Attendance.date <= d_to,
                Attendance.employee_id.in_(emp_ids)
            ).distinct().all()
            emp_ids_with_attendance = [e[0] for e in emp_ids_with_attendance]
            
            if not emp_ids_with_attendance:
                continue
            
            dept_attendance = Attendance.query.filter(
                Attendance.date >= d_from,
                Attendance.date <= d_to,
                Attendance.employee_id.in_(emp_ids_with_attendance)
            ).all()
            
            present = sum(1 for a in dept_attendance if a.status == 'present')
            total = len(dept_attendance)
            rate = round((present / total) * 100) if total > 0 else 0
            
            if rate >= 90:
                rating = "Ù…Ù…ØªØ§Ø² â­"
                rating_fill = green_fill
            elif rate >= 75:
                rating = "Ø¬ÙŠØ¯ ğŸ‘"
                rating_fill = teal_fill
            elif rate >= 60:
                rating = "Ù…ØªÙˆØ³Ø· âš¡"
                rating_fill = orange_fill
            else:
                rating = "ÙŠØ­ØªØ§Ø¬ ØªØ­Ø³ÙŠÙ† âš ï¸"
                rating_fill = red_fill
            
            ws.cell(row=dept_row, column=1).value = dept.name
            ws.cell(row=dept_row, column=1).font = white_font
            ws.cell(row=dept_row, column=1).fill = card_fill
            ws.cell(row=dept_row, column=1).border = thin_border
            
            ws.cell(row=dept_row, column=2).value = len(emp_ids_with_attendance)
            ws.cell(row=dept_row, column=2).font = white_font
            ws.cell(row=dept_row, column=2).fill = card_fill
            ws.cell(row=dept_row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=dept_row, column=2).border = thin_border
            
            ws.cell(row=dept_row, column=3).value = present
            ws.cell(row=dept_row, column=3).font = white_font
            ws.cell(row=dept_row, column=3).fill = card_fill
            ws.cell(row=dept_row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=dept_row, column=3).border = thin_border
            
            ws.cell(row=dept_row, column=4).value = f"{rate}%"
            ws.cell(row=dept_row, column=4).font = white_font
            ws.cell(row=dept_row, column=4).fill = card_fill
            ws.cell(row=dept_row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=dept_row, column=4).border = thin_border
            
            ws.cell(row=dept_row, column=5).value = rating
            ws.cell(row=dept_row, column=5).font = Font(bold=True, color="FFFFFF", size=10)
            ws.cell(row=dept_row, column=5).fill = rating_fill
            ws.cell(row=dept_row, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=dept_row, column=5).border = thin_border
            
            dept_row += 1
        
        if dept_row > 26:
            bar1 = BarChart()
            bar1.type = "col"
            bar1.style = 12
            bar1.title = "Ù†Ø³Ø¨Ø© Ø§Ù„Ø­Ø¶ÙˆØ± Ø¨Ø§Ù„Ø£Ù‚Ø³Ø§Ù… - Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†"
            bar1.y_axis.title = "Ø§Ù„Ù†Ø³Ø¨Ø© %"
            
            data_bar = Reference(ws, min_col=4, min_row=25, max_row=dept_row-1)
            cats_bar = Reference(ws, min_col=1, min_row=26, max_row=dept_row-1)
            bar1.add_data(data_bar, titles_from_data=True)
            bar1.set_categories(cats_bar)
            bar1.width = 14
            bar1.height = 10
            bar1.dataLabels = DataLabelList()
            bar1.dataLabels.showVal = True
            ws.add_chart(bar1, "G24")
        
        doc_counts = db.session.query(
            Document.employee_id,
            func.count(Document.id)
        ).group_by(Document.employee_id).all()
        
        complete_docs = sum(1 for _, cnt in doc_counts if cnt >= 4)
        incomplete_docs = max(0, total_employees - complete_docs)
        
        doc_start = dept_row + 2
        ws.merge_cells(f'A{doc_start}:H{doc_start}')
        ws[f'A{doc_start}'] = "ğŸ“„ Ø­Ø§Ù„Ø© Ø§ÙƒØªÙ…Ø§Ù„ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚ - Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†"
        ws[f'A{doc_start}'].font = cyan_font
        ws[f'A{doc_start}'].fill = header_fill
        ws[f'A{doc_start}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=doc_start+1, column=1).value = "Ø§Ù„Ø­Ø§Ù„Ø©"
        ws.cell(row=doc_start+1, column=2).value = "Ø§Ù„Ø¹Ø¯Ø¯"
        ws.cell(row=doc_start+1, column=3).value = "Ø§Ù„Ù†Ø³Ø¨Ø©"
        for i in range(1, 4):
            ws.cell(row=doc_start+1, column=i).font = cyan_font
            ws.cell(row=doc_start+1, column=i).fill = header_fill
            ws.cell(row=doc_start+1, column=i).border = thin_border
        
        complete_pct = round((complete_docs / total_employees * 100), 1) if total_employees > 0 else 0
        incomplete_pct = round((incomplete_docs / total_employees * 100), 1) if total_employees > 0 else 0
        
        ws.cell(row=doc_start+2, column=1).value = "Ù…ÙƒØªÙ…Ù„ âœ…"
        ws.cell(row=doc_start+2, column=2).value = complete_docs
        ws.cell(row=doc_start+2, column=3).value = f"{complete_pct}%"
        for i in range(1, 4):
            ws.cell(row=doc_start+2, column=i).font = white_font
            ws.cell(row=doc_start+2, column=i).fill = card_fill
            ws.cell(row=doc_start+2, column=i).border = thin_border
        
        ws.cell(row=doc_start+3, column=1).value = "Ù†Ø§Ù‚Øµ âš ï¸"
        ws.cell(row=doc_start+3, column=2).value = incomplete_docs
        ws.cell(row=doc_start+3, column=3).value = f"{incomplete_pct}%"
        for i in range(1, 4):
            ws.cell(row=doc_start+3, column=i).font = white_font
            ws.cell(row=doc_start+3, column=i).fill = card_fill
            ws.cell(row=doc_start+3, column=i).border = thin_border
        
        pie3 = PieChart()
        pie3.title = "Ø§ÙƒØªÙ…Ø§Ù„ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚"
        labels3 = Reference(ws, min_col=1, min_row=doc_start+2, max_row=doc_start+3)
        data3 = Reference(ws, min_col=2, min_row=doc_start+1, max_row=doc_start+3)
        pie3.add_data(data3, titles_from_data=True)
        pie3.set_categories(labels3)
        pie3.width = 10
        pie3.height = 8
        pie3.dataLabels = DataLabelList()
        pie3.dataLabels.showPercent = True
        ws.add_chart(pie3, f"E{doc_start}")
        
        # Ø£Ù„ÙˆØ§Ù† Ù„Ù„ØªÙ‚Ø§Ø±ÙŠØ± Ø§Ù„ØªÙØµÙŠÙ„ÙŠØ©
        detail_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        detail_header_font = Font(bold=True, color="00D4AA", size=12)
        success_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        danger_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        info_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        if data_type in ['attendance', 'all']:
            ws_att = wb.create_sheet("ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø­Ø¶ÙˆØ±")
            ws_att.sheet_view.rightToLeft = True
            
            ws_att.merge_cells('A1:G1')
            ws_att['A1'] = "ğŸ“‹ ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø­Ø¶ÙˆØ± Ø§Ù„ØªÙØµÙŠÙ„ÙŠ - Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ†"
            ws_att['A1'].font = title_font
            ws_att['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_att.row_dimensions[1].height = 30
            
            ws_att.merge_cells('A2:G2')
            ws_att['A2'] = f"Ø§Ù„ÙØªØ±Ø©: {d_from} Ø¥Ù„Ù‰ {d_to}"
            ws_att['A2'].font = subtitle_font
            ws_att['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            headers = ['Ø§Ù„ØªØ§Ø±ÙŠØ®', 'Ø§Ø³Ù… Ø§Ù„Ù…ÙˆØ¸Ù', 'Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ¸ÙŠÙÙŠ', 'Ø§Ù„Ù‚Ø³Ù…', 'Ø§Ù„Ø­Ø§Ù„Ø©', 'ÙˆÙ‚Øª Ø§Ù„Ø­Ø¶ÙˆØ±', 'ÙˆÙ‚Øª Ø§Ù„Ø§Ù†ØµØ±Ø§Ù']
            for col, header in enumerate(headers, start=1):
                cell = ws_att.cell(row=4, column=col)
                cell.value = header
                cell.font = detail_header_font
                cell.fill = detail_header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_att.row_dimensions[4].height = 25
            
            # Ø¬Ù„Ø¨ Ø§Ù„Ø­Ø¶ÙˆØ± Ù„Ù„Ù…ÙˆØ¸ÙÙŠÙ† Ø§Ù„Ù†Ø´Ø·ÙŠÙ† ÙÙ‚Ø·
            query = Attendance.query.filter(
                Attendance.employee_id.in_(active_emp_ids)
            ).order_by(Attendance.date.desc())
            query = query.filter(Attendance.date >= d_from, Attendance.date <= d_to)
            
            status_translations = {
                'present': 'Ø­Ø§Ø¶Ø±',
                'absent': 'ØºØ§Ø¦Ø¨',
                'late': 'Ù…ØªØ£Ø®Ø±',
                'excused': 'Ù…Ø¹Ø°ÙˆØ±',
                'leave': 'Ø¥Ø¬Ø§Ø²Ø©',
                'sick': 'Ù…Ø±ÙŠØ¶'
            }
            
            att_records = query.limit(500).all()
            for row_idx, record in enumerate(att_records, start=5):
                emp = Employee.query.get(record.employee_id) if record.employee_id else None
                status = record.status or 'unknown'
                
                data_row = [
                    record.date.strftime('%Y-%m-%d') if record.date else '',
                    emp.name if emp else 'ØºÙŠØ± Ù…Ø¹Ø±ÙˆÙ',
                    emp.employee_id if emp else '',
                    emp.department.name if emp and emp.department else '',
                    status_translations.get(status, status),
                    record.time.strftime('%H:%M') if hasattr(record, 'time') and record.time else '',
                    record.checkout_time.strftime('%H:%M') if hasattr(record, 'checkout_time') and record.checkout_time else ''
                ]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_att.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 5:
                        if status == 'present':
                            cell.fill = success_fill
                        elif status == 'absent':
                            cell.fill = danger_fill
                        elif status == 'late':
                            cell.fill = warning_fill
                        elif status == 'excused':
                            cell.fill = info_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 8):
                ws_att.column_dimensions[get_column_letter(col)].width = 18
        
        if data_type in ['employees', 'all']:
            ws_emp = wb.create_sheet("ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†")
            ws_emp.sheet_view.rightToLeft = True
            
            ws_emp.merge_cells('A1:F1')
            ws_emp['A1'] = "ğŸ‘¥ ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ† ÙˆØ§Ù„ÙˆØ«Ø§Ø¦Ù‚"
            ws_emp['A1'].font = title_font
            ws_emp['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_emp.row_dimensions[1].height = 30
            
            headers = ['Ø§Ù„Ø±Ù‚Ù…', 'Ø§Ø³Ù… Ø§Ù„Ù…ÙˆØ¸Ù', 'Ø§Ù„Ø±Ù‚Ù… Ø§Ù„ÙˆØ¸ÙŠÙÙŠ', 'Ø§Ù„Ù‚Ø³Ù…', 'Ø¹Ø¯Ø¯ Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚', 'Ø­Ø§Ù„Ø© Ø§Ù„ÙˆØ«Ø§Ø¦Ù‚']
            for col, header in enumerate(headers, start=1):
                cell = ws_emp.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_emp.row_dimensions[3].height = 25
            
            employees = Employee.query.all()
            
            doc_counts = db.session.query(
                Document.employee_id,
                func.count(Document.id).label('count')
            ).group_by(Document.employee_id).all()
            doc_count_map = {e_id: cnt for e_id, cnt in doc_counts}
            
            for row_idx, emp in enumerate(employees, start=4):
                docs_count = doc_count_map.get(emp.id, 0)
                status = 'Ù…ÙƒØªÙ…Ù„ âœ…' if docs_count >= 4 else 'Ù†Ø§Ù‚Øµ âš ï¸'
                
                data_row = [
                    emp.id,
                    emp.name,
                    emp.employee_id or '-',
                    emp.department.name if emp.department else 'Ø¨Ø¯ÙˆÙ† Ù‚Ø³Ù…',
                    docs_count,
                    status
                ]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_emp.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 6:
                        cell.fill = success_fill if docs_count >= 4 else warning_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 7):
                ws_emp.column_dimensions[get_column_letter(col)].width = 18
        
        if data_type in ['vehicles', 'all']:
            ws_veh = wb.create_sheet("ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª")
            ws_veh.sheet_view.rightToLeft = True
            
            ws_veh.merge_cells('A1:F1')
            ws_veh['A1'] = "ğŸš— ØªÙ‚Ø±ÙŠØ± Ø£Ø³Ø·ÙˆÙ„ Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª"
            ws_veh['A1'].font = title_font
            ws_veh['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_veh.row_dimensions[1].height = 30
            
            headers = ['Ø±Ù‚Ù… Ø§Ù„Ù„ÙˆØ­Ø©', 'Ø§Ù„Ù…Ø§Ø±ÙƒØ©', 'Ø§Ù„Ù…ÙˆØ¯ÙŠÙ„', 'Ø§Ù„Ø³Ù†Ø©', 'Ø§Ù„Ø­Ø§Ù„Ø©', 'Ø­Ø§Ù„Ø© Ø§Ù„ØªØ³Ù„ÙŠÙ…']
            for col, header in enumerate(headers, start=1):
                cell = ws_veh.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_veh.row_dimensions[3].height = 25
            
            status_translations = {
                'working': 'Ù†Ø´Ø·',
                'maintenance': 'ØµÙŠØ§Ù†Ø©',
                'inactive': 'ØºÙŠØ± Ù†Ø´Ø·'
            }
            
            vehicles = Vehicle.query.all()
            for row_idx, v in enumerate(vehicles, start=4):
                status = v.status or 'unknown'
                
                data_row = [
                    v.plate_number if hasattr(v, 'plate_number') else '',
                    v.make if hasattr(v, 'make') else '',
                    v.model if hasattr(v, 'model') else '',
                    v.year if hasattr(v, 'year') else '',
                    status_translations.get(status, status),
                    'Ù…Ø³ØªÙ„Ù…Ø© âœ…' if v.handover_records else 'ØºÙŠØ± Ù…Ø³ØªÙ„Ù…Ø©'
                ]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_veh.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 5:
                        if status == 'working':
                            cell.fill = success_fill
                        elif status == 'maintenance':
                            cell.fill = warning_fill
                        elif status == 'inactive':
                            cell.fill = danger_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 7):
                ws_veh.column_dimensions[get_column_letter(col)].width = 16
        
        if data_type == 'all':
            ws_dept = wb.create_sheet("ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø£Ù‚Ø³Ø§Ù…")
            ws_dept.sheet_view.rightToLeft = True
            
            ws_dept.merge_cells('A1:G1')
            ws_dept['A1'] = "ğŸ¢ ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø­Ø¶ÙˆØ± Ø­Ø³Ø¨ Ø§Ù„Ø£Ù‚Ø³Ø§Ù…"
            ws_dept['A1'].font = title_font
            ws_dept['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_dept.row_dimensions[1].height = 30
            
            headers = ['Ø§Ù„Ù‚Ø³Ù…', 'Ø¹Ø¯Ø¯ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†', 'Ø­Ø§Ø¶Ø±', 'ØºØ§Ø¦Ø¨', 'Ù…ØªØ£Ø®Ø±', 'Ù†Ø³Ø¨Ø© Ø§Ù„Ø­Ø¶ÙˆØ±', 'Ø§Ù„ØªÙ‚ÙŠÙŠÙ…']
            for col, header in enumerate(headers, start=1):
                cell = ws_dept.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_dept.row_dimensions[3].height = 25
            
            if date_from:
                d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            else:
                d_from = datetime.now().date() - timedelta(days=30)
            
            if date_to:
                d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            else:
                d_to = datetime.now().date()
            
            departments = Department.query.all()
            
            all_employees = Employee.query.all()
            dept_employees = {}
            for e in all_employees:
                if e.department_id:
                    if e.department_id not in dept_employees:
                        dept_employees[e.department_id] = []
                    dept_employees[e.department_id].append(e.id)
            
            all_attendance = Attendance.query.filter(
                Attendance.date >= d_from,
                Attendance.date <= d_to
            ).all()
            
            dept_attendance = {}
            for a in all_attendance:
                for d_id, emp_ids in dept_employees.items():
                    if a.employee_id in emp_ids:
                        if d_id not in dept_attendance:
                            dept_attendance[d_id] = {'present': 0, 'absent': 0, 'late': 0, 'total': 0}
                        dept_attendance[d_id]['total'] += 1
                        if a.status == 'present':
                            dept_attendance[d_id]['present'] += 1
                        elif a.status == 'absent':
                            dept_attendance[d_id]['absent'] += 1
                        elif a.status == 'late':
                            dept_attendance[d_id]['late'] += 1
                        break
            
            actual_row = 4
            for dept in departments:
                emp_count = len(dept_employees.get(dept.id, []))
                if emp_count == 0:
                    continue
                
                stats = dept_attendance.get(dept.id, {'present': 0, 'absent': 0, 'late': 0, 'total': 0})
                present = stats['present']
                absent = stats['absent']
                late = stats['late']
                total = stats['total']
                rate = round((present / total) * 100, 1) if total > 0 else 0
                
                if rate >= 90:
                    performance = 'Ù…Ù…ØªØ§Ø² â­'
                    perf_fill = success_fill
                elif rate >= 75:
                    performance = 'Ø¬ÙŠØ¯ ğŸ‘'
                    perf_fill = info_fill
                elif rate >= 60:
                    performance = 'Ù…ØªÙˆØ³Ø· âš¡'
                    perf_fill = warning_fill
                else:
                    performance = 'ÙŠØ­ØªØ§Ø¬ ØªØ­Ø³ÙŠÙ† âš ï¸'
                    perf_fill = danger_fill
                
                data_row = [dept.name, len(employees), present, absent, late, f'{rate}%', performance]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_dept.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 7:
                        cell.fill = perf_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 8):
                ws_dept.column_dimensions[get_column_letter(col)].width = 16
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"powerbi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    """Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª Ø´Ø§Ù…Ù„Ø© Ù„Ù„ÙˆØ­Ø© Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª"""
    try:
        total_employees = Employee.query.count()
        total_vehicles = Vehicle.query.count()
        total_documents = Document.query.count()
        total_departments = Department.query.count()
        
        today = datetime.now().date()
        today_attendance = Attendance.query.filter(Attendance.date == today).all()
        today_present = sum(1 for a in today_attendance if a.status == 'present')
        
        working_vehicles = Vehicle.query.filter_by(status='working').count()
        
        return jsonify({
            'success': True,
            'data': {
                'employees': {
                    'total': total_employees,
                    'present_today': today_present
                },
                'vehicles': {
                    'total': total_vehicles,
                    'working': working_vehicles
                },
                'documents': {
                    'total': total_documents
                },
                'departments': {
                    'total': total_departments
                },
                'last_updated': datetime.now().isoformat()
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
