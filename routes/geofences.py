from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from models import Geofence, GeofenceEvent, GeofenceSession, Employee, Department, Attendance, EmployeeLocation, db, employee_departments
from datetime import datetime, timedelta
from utils.geofence_session_manager import SessionManager
from sqlalchemy import func, desc
import re
import requests

geofences_bp = Blueprint('geofences', __name__, url_prefix='/employees/geofences')


@geofences_bp.route('/')
@login_required
def index():
    """صفحة إدارة الدوائر الجغرافية"""
    geofences = Geofence.query.filter_by(is_active=True).all()
    departments = Department.query.all()
    
    geofences_data = []
    for geofence in geofences:
        employees_inside = geofence.get_department_employees_inside()
        
        # تحويل الموظفين إلى قواميس قابلة للتحويل إلى JSON
        employees_list = []
        for emp_data in employees_inside:
            emp = emp_data.get('employee') if isinstance(emp_data, dict) else emp_data
            employees_list.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'profile_image': emp.profile_image if hasattr(emp, 'profile_image') else None
            })
        
        geofences_data.append({
            'geofence': {
                'id': geofence.id,
                'name': geofence.name,
                'type': geofence.type,
                'center_latitude': geofence.center_latitude,
                'center_longitude': geofence.center_longitude,
                'radius_meters': geofence.radius_meters,
                'color': geofence.color,
                'description': geofence.description,
                'department_id': geofence.department_id,
                'department': {
                    'id': geofence.department.id,
                    'name': geofence.department.name
                }
            },
            'employees_count': len(employees_inside),
            'employees_inside': employees_list
        })
    
    return render_template(
        'geofences/index.html',
        geofences_data=geofences_data,
        departments=departments
    )


@geofences_bp.route('/create', methods=['POST'])
@login_required
def create():
    """إنشاء دائرة جديدة"""
    try:
        data = request.get_json()
        
        geofence = Geofence(
            name=data['name'],
            type=data.get('type', 'project'),
            description=data.get('description'),
            center_latitude=data['latitude'],
            center_longitude=data['longitude'],
            radius_meters=data['radius'],
            color=data.get('color', '#667eea'),
            department_id=data['department_id'],
            notify_on_entry=data.get('notify_on_entry', False),
            notify_on_exit=data.get('notify_on_exit', False),
            created_by=current_user.id
        )
        
        db.session.add(geofence)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'geofence_id': geofence.id,
            'message': 'تم إنشاء الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الدائرة: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>')
@login_required
def view(geofence_id):
    """عرض تفاصيل دائرة معينة"""
    from collections import defaultdict
    import json
    
    geofence = Geofence.query.get_or_404(geofence_id)
    
    employees_inside = geofence.get_department_employees_inside()
    all_employees = geofence.get_all_employees_inside()
    
    recent_events = GeofenceEvent.query.filter_by(
        geofence_id=geofence_id
    ).order_by(GeofenceEvent.recorded_at.desc()).limit(50).all()
    
    # جلب جميع موظفي القسم للإضافة
    department_employees = Employee.query.join(employee_departments).filter(
        employee_departments.c.department_id == geofence.department_id
    ).all()
    
    # الموظفون المتاحون للربط (غير مرتبطين بالدائرة حالياً)
    assigned_employee_ids = [emp.id for emp in geofence.assigned_employees]
    available_employees = [emp for emp in department_employees if emp.id not in assigned_employee_ids]
    
    # جلب الجلسات النشطة (الموظفون داخل الدائرة الآن)
    active_sessions = SessionManager.get_active_sessions(geofence_id=geofence_id)
    
    # جلب جميع الجلسات مع تفاصيل الموظفين
    all_sessions = GeofenceSession.query.filter_by(
        geofence_id=geofence_id
    ).options(
        db.joinedload(GeofenceSession.employee)
    ).order_by(GeofenceSession.entry_time.desc()).limit(100).all()
    
    # حساب إحصائيات الموظفين
    employee_stats = {}
    for emp in geofence.assigned_employees:
        total_time = SessionManager.get_employee_total_time(emp.id, geofence_id)
        visit_count = SessionManager.get_employee_visit_count(emp.id, geofence_id)
        is_inside = any(s.employee_id == emp.id for s in active_sessions)
        
        employee_stats[emp.id] = {
            'employee': emp,
            'total_time_minutes': total_time,
            'total_time_hours': round(total_time / 60, 1) if total_time else 0,
            'visit_count': visit_count,
            'avg_duration': round(total_time / visit_count, 1) if visit_count > 0 else 0,
            'is_inside': is_inside
        }
    
    # إحصائيات الحضور
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    
    total_assigned = len(geofence.assigned_employees)
    # حساب فقط الموظفين المعينين للدائرة والموجودين داخلها
    assigned_ids = {emp.id for emp in geofence.assigned_employees}
    present_count = len([s for s in active_sessions if s.employee_id in assigned_ids])
    absent_count = total_assigned - present_count
    attendance_rate = (present_count / total_assigned * 100) if total_assigned > 0 else 0
    
    # إحصائيات آخر 24 ساعة (كل ساعة)
    hourly_data = []
    for hour in range(24):
        hour_start = today_start + timedelta(hours=hour)
        hour_end = hour_start + timedelta(hours=1)
        
        events_count = GeofenceEvent.query.filter(
            GeofenceEvent.geofence_id == geofence_id,
            GeofenceEvent.recorded_at >= hour_start,
            GeofenceEvent.recorded_at < hour_end
        ).count()
        
        hourly_data.append({
            'hour': hour,
            'count': events_count
        })
    
    # إحصائيات أسبوعية
    weekly_data = []
    days_ar = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت']
    for day in range(7):
        day_start = week_start + timedelta(days=day)
        day_end = day_start + timedelta(days=1)
        
        events_count = GeofenceEvent.query.filter(
            GeofenceEvent.geofence_id == geofence_id,
            GeofenceEvent.recorded_at >= day_start,
            GeofenceEvent.recorded_at < day_end
        ).count()
        
        weekly_data.append({
            'day': days_ar[day_start.weekday()],
            'count': events_count
        })
    
    stats = {
        'total_assigned': total_assigned,
        'present_count': present_count,
        'absent_count': absent_count,
        'attendance_rate': round(attendance_rate, 1),
        'hourly_data': json.dumps(hourly_data),
        'weekly_data': json.dumps(weekly_data)
    }
    
    return render_template(
        'geofences/view.html',
        geofence=geofence,
        employees_inside=employees_inside,
        all_employees=all_employees,
        recent_events=recent_events,
        assigned_employees=geofence.assigned_employees,
        available_employees=available_employees,
        stats=stats,
        employee_stats=employee_stats,
        all_sessions=all_sessions,
        active_sessions=active_sessions
    )


@geofences_bp.route('/<int:geofence_id>/bulk-check-in', methods=['POST'])
@login_required
def bulk_check_in(geofence_id):
    """تسجيل حضور جماعي فقط لموظفي القسم المرتبط بالدائرة والموجودين داخل دوائرهم المخصصة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        employees_inside = geofence.get_department_employees_inside()
        
        if not employees_inside:
            return jsonify({
                'success': False,
                'message': f'لا يوجد موظفين من قسم "{geofence.department.name}" داخل الدائرة حالياً'
            })
        
        checked_in = []
        already_checked = []
        errors = []
        not_assigned = []
        
        for emp_data in employees_inside:
            employee = emp_data['employee']
            location = emp_data['location']
            
            # التحقق من أن الموظف مرتبط بهذه الدائرة
            if employee not in geofence.assigned_employees:
                not_assigned.append(employee.name)
                continue
            
            today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
            existing_attendance = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.check_in_time >= today_start
            ).first()
            
            if existing_attendance:
                already_checked.append(employee.name)
                continue
            
            try:
                attendance = Attendance(
                    employee_id=employee.id,
                    check_in_time=datetime.utcnow(),
                    status='present',
                    notes=f'تسجيل جماعي من دائرة: {geofence.name} (قسم: {geofence.department.name})'
                )
                db.session.add(attendance)
                db.session.flush()
                
                event = GeofenceEvent(
                    geofence_id=geofence.id,
                    employee_id=employee.id,
                    event_type='bulk_check_in',
                    location_latitude=location.latitude,
                    location_longitude=location.longitude,
                    distance_from_center=int(emp_data['distance']),
                    source='bulk',
                    attendance_id=attendance.id,
                    notes=f'تسجيل جماعي بواسطة {current_user.username} - قسم: {geofence.department.name}'
                )
                db.session.add(event)
                
                checked_in.append(employee.name)
                
            except Exception as e:
                errors.append(f'{employee.name}: {str(e)}')
        
        db.session.commit()
        
        message_parts = [f'تم تسجيل حضور {len(checked_in)} موظف']
        if not_assigned:
            message_parts.append(f'({len(not_assigned)} موظف غير مرتبط بهذه الدائرة)')
        if already_checked:
            message_parts.append(f'({len(already_checked)} تم تسجيلهم مسبقاً)')
        
        return jsonify({
            'success': True,
            'department_name': geofence.department.name,
            'checked_in_count': len(checked_in),
            'already_checked_count': len(already_checked),
            'not_assigned_count': len(not_assigned),
            'error_count': len(errors),
            'checked_in': checked_in,
            'already_checked': already_checked,
            'not_assigned': not_assigned,
            'errors': errors,
            'message': ' '.join(message_parts)
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في تسجيل الحضور: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/employees')
@login_required
def get_employees(geofence_id):
    """جلب الموظفين داخل الدائرة (API)"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        employees_inside = geofence.get_department_employees_inside()
        all_employees = geofence.get_all_employees_inside()
        
        return jsonify({
            'success': True,
            'department_employees': [{
                'id': emp['employee'].id,
                'name': emp['employee'].name,
                'employee_id': emp['employee'].employee_id,
                'distance': round(emp['distance'], 2),
                'latitude': float(emp['location'].latitude),
                'longitude': float(emp['location'].longitude),
                'profile_image': emp['employee'].profile_image
            } for emp in employees_inside],
            'other_employees': [{
                'id': emp['employee'].id,
                'name': emp['employee'].name,
                'employee_id': emp['employee'].employee_id,
                'distance': round(emp['distance'], 2),
                'latitude': float(emp['location'].latitude),
                'longitude': float(emp['location'].longitude),
                'is_eligible': emp['is_eligible'],
                'profile_image': emp['employee'].profile_image
            } for emp in all_employees if not emp['is_eligible']]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/events')
@login_required
def get_events(geofence_id):
    """جلب أحداث دائرة معينة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        events = GeofenceEvent.query.filter_by(
            geofence_id=geofence_id
        ).order_by(GeofenceEvent.recorded_at.desc()).limit(100).all()
        
        return jsonify({
            'success': True,
            'events': [{
                'id': event.id,
                'employee_name': event.employee.name,
                'event_type': event.event_type,
                'event_time': event.recorded_at.isoformat(),
                'distance': event.distance_from_center,
                'notes': event.notes
            } for event in events]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/api/list')
@login_required
def api_list():
    """API: جلب قائمة جميع الدوائر النشطة"""
    try:
        geofences = Geofence.query.filter_by(is_active=True).all()
        
        return jsonify({
            'success': True,
            'geofences': [{
                'id': g.id,
                'name': g.name,
                'type': g.type,
                'center_lat': float(g.center_latitude),
                'center_lng': float(g.center_longitude),
                'radius': g.radius_meters,
                'color': g.color,
                'department_id': g.department_id,
                'department_name': g.department.name
            } for g in geofences]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/update-settings', methods=['POST'])
@login_required
def update_settings(geofence_id):
    """تحديث إعدادات الحضور للدائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        
        if 'attendance_start_time' in data:
            geofence.attendance_start_time = data['attendance_start_time']
        if 'attendance_required_minutes' in data:
            geofence.attendance_required_minutes = data['attendance_required_minutes']
        
        geofence.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث إعدادات الحضور بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/update', methods=['PUT'])
@login_required
def update(geofence_id):
    """تحديث دائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        
        if 'name' in data:
            geofence.name = data['name']
        if 'description' in data:
            geofence.description = data['description']
        if 'radius' in data:
            geofence.radius_meters = data['radius']
        if 'color' in data:
            geofence.color = data['color']
        if 'notify_on_entry' in data:
            geofence.notify_on_entry = data['notify_on_entry']
        if 'notify_on_exit' in data:
            geofence.notify_on_exit = data['notify_on_exit']
        
        geofence.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في التحديث: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/delete', methods=['POST', 'DELETE'])
@login_required
def delete(geofence_id):
    """حذف دائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        db.session.delete(geofence)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم حذف الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في الحذف: {str(e)}'
        }), 400


@geofences_bp.route('/extract-google-maps-coords', methods=['POST'])
@login_required
def extract_google_maps_coords():
    """استخراج الإحداثيات من روابط Google Maps (بما في ذلك الروابط المختصرة)"""
    try:
        data = request.get_json()
        url = data.get('url', '').strip()
        
        if not url:
            return jsonify({
                'success': False,
                'message': 'الرجاء إدخال رابط Google Maps'
            }), 400
        
        # إذا كان الرابط مختصر (goo.gl أو maps.app.goo.gl)، نحتاج لفتحه للحصول على الرابط الكامل
        if 'goo.gl' in url or 'maps.app.goo.gl' in url:
            try:
                # إرسال طلب للحصول على الرابط الكامل بعد إعادة التوجيه
                response = requests.get(url, allow_redirects=True, timeout=10)
                url = response.url  # الرابط الكامل بعد إعادة التوجيه
            except Exception as e:
                return jsonify({
                    'success': False,
                    'message': f'فشل فتح الرابط المختصر: {str(e)}'
                }), 400
        
        # استخراج الإحداثيات من الرابط
        coords = None
        
        # نمط 1: !3d/!4d (الأكثر دقة - موقع العلامة)
        match = re.search(r'!3d(-?\d+\.?\d*)!4d(-?\d+\.?\d*)', url)
        if match:
            coords = {
                'lat': float(match.group(1)),
                'lng': float(match.group(2))
            }
        
        # نمط 2: @ (مركز الخريطة)
        if not coords:
            match = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', url)
            if match:
                coords = {
                    'lat': float(match.group(1)),
                    'lng': float(match.group(2))
                }
        
        # نمط 3: ?q= أو ll= (روابط قصيرة)
        if not coords:
            match = re.search(r'[?&](q|ll)=(-?\d+\.?\d*),(-?\d+\.?\d*)', url)
            if match:
                coords = {
                    'lat': float(match.group(2)),
                    'lng': float(match.group(3))
                }
        
        if coords:
            return jsonify({
                'success': True,
                'coords': coords,
                'full_url': url
            })
        else:
            return jsonify({
                'success': False,
                'message': 'لم يتم العثور على إحداثيات في الرابط'
            }), 400
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/assign-employees', methods=['POST'])
@login_required
def assign_employees(geofence_id):
    """ربط موظفين بدائرة جغرافية محددة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        employee_ids = data.get('employee_ids', [])
        
        if not employee_ids:
            return jsonify({
                'success': False,
                'message': 'الرجاء اختيار موظف واحد على الأقل'
            }), 400
        
        # إضافة الموظفين إلى الدائرة
        for employee_id in employee_ids:
            employee = Employee.query.get(employee_id)
            if employee and employee not in geofence.assigned_employees:
                geofence.assigned_employees.append(employee)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم ربط {len(employee_ids)} موظف بالدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في ربط الموظفين: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/unassign-employee/<int:employee_id>', methods=['POST'])
@login_required
def unassign_employee(geofence_id, employee_id):
    """إلغاء ربط موظف من دائرة جغرافية"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        employee = Employee.query.get_or_404(employee_id)
        
        if employee in geofence.assigned_employees:
            geofence.assigned_employees.remove(employee)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'تم إلغاء ربط الموظف من الدائرة'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'الموظف غير مرتبط بهذه الدائرة'
            }), 400
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/update', methods=['POST'])
@login_required
def update_geofence(geofence_id):
    """تحديث بيانات الدائرة الجغرافية"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        
        name = data.get('name', '').strip()
        radius_meters = data.get('radius_meters')
        color = data.get('color', '').strip()
        
        if not name:
            return jsonify({
                'success': False,
                'message': 'الرجاء إدخال اسم الدائرة'
            }), 400
        
        if not radius_meters or radius_meters < 10:
            return jsonify({
                'success': False,
                'message': 'نصف القطر يجب أن يكون 10 متر على الأقل'
            }), 400
        
        if not color or not color.startswith('#'):
            return jsonify({
                'success': False,
                'message': 'الرجاء اختيار لون صحيح'
            }), 400
        
        geofence.name = name
        geofence.radius_meters = radius_meters
        geofence.color = color
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في التحديث: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/export-events')
@login_required
def export_events(geofence_id):
    """تصدير بيانات الوصول والمغادرة لموظفي الدائرة"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "سجل الحضور والمغادرة"
        
        ws.right_to_left = True
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['الحالة', 'نوع الحدث', 'وقت الحضور', 'وقت الخروج', 'رقم الموظف', 'اسم الموظف', 'القسم', 'الدائرة']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        employees_inside = geofence.get_department_employees_inside()
        inside_employee_ids = {emp['employee'].id for emp in employees_inside}
        
        all_assigned_employees = geofence.assigned_employees
        
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        
        row_num = 2
        for employee in all_assigned_employees:
            is_inside = employee.id in inside_employee_ids
            
            entry_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence_id,
                employee_id=employee.id,
                event_type='entry'
            ).filter(
                GeofenceEvent.recorded_at >= today_start
            ).order_by(GeofenceEvent.recorded_at.asc()).first()
            
            bulk_entry_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence_id,
                employee_id=employee.id,
                event_type='bulk_check_in'
            ).filter(
                GeofenceEvent.recorded_at >= today_start
            ).order_by(GeofenceEvent.recorded_at.asc()).first()
            
            exit_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence_id,
                employee_id=employee.id,
                event_type='exit'
            ).filter(
                GeofenceEvent.recorded_at >= today_start
            ).order_by(GeofenceEvent.recorded_at.desc()).first()
            
            first_entry = entry_event if entry_event else bulk_entry_event
            if entry_event and bulk_entry_event:
                first_entry = entry_event if entry_event.recorded_at < bulk_entry_event.recorded_at else bulk_entry_event
            
            entry_time = first_entry.recorded_at.strftime('%H:%M:%S') if first_entry else '-'
            exit_time = exit_event.recorded_at.strftime('%H:%M:%S') if exit_event else '-'
            
            if is_inside:
                status = 'موجود داخل الدائرة'
                event_type = 'حضور'
            else:
                if first_entry:
                    status = 'خارج الدائرة'
                    event_type = 'غادر'
                else:
                    status = 'خارج الحضور'
                    event_type = 'غائب'
            
            ws.cell(row=row_num, column=1, value=status)
            ws.cell(row=row_num, column=2, value=event_type)
            ws.cell(row=row_num, column=3, value=entry_time)
            ws.cell(row=row_num, column=4, value=exit_time)
            ws.cell(row=row_num, column=5, value=employee.employee_id)
            ws.cell(row=row_num, column=6, value=employee.name)
            ws.cell(row=row_num, column=7, value=geofence.department.name if geofence.department else '-')
            ws.cell(row=row_num, column=8, value=geofence.name)
            
            for col in range(1, 9):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if col == 1:
                    if is_inside:
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = Font(color="065F46", bold=True)
                    elif status == 'خارج الحضور':
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(color="991B1B", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(color="92400E", bold=True)
            
            row_num += 1
        
        for col in range(1, 9):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"تقرير_الحضور_{geofence.name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('geofences.view', geofence_id=geofence_id))
