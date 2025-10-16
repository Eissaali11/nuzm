from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func, or_, and_
import os
import uuid
from PIL import Image
import pillow_heif
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from app import db
from models import RentalProperty, PropertyImage, PropertyPayment, PropertyFurnishing, User
from forms.property_forms import (
    RentalPropertyForm, PropertyImagesForm, PropertyPaymentForm, PropertyFurnishingForm
)
from utils.audit_logger import log_activity

properties_bp = Blueprint('properties', __name__)

# إعدادات رفع الملفات
UPLOAD_FOLDER = 'uploads/properties'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'heic', 'webp'}


def allowed_file(filename):
    """التحقق من امتداد الملف المسموح به"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def process_and_save_image(file, property_id):
    """معالجة وحفظ الصورة مع دعم HEIC"""
    try:
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        unique_filename = f"{uuid.uuid4()}.{file_ext}"
        
        # إنشاء مجلد التخزين
        property_folder = os.path.join(UPLOAD_FOLDER, str(property_id))
        os.makedirs(property_folder, exist_ok=True)
        filepath = os.path.join(property_folder, unique_filename)
        
        # معالجة صور HEIC
        if file_ext == 'heic':
            heif_file = pillow_heif.read_heif(file)
            image = Image.frombytes(
                heif_file.mode,
                heif_file.size,
                heif_file.data,
                "raw",
            )
            # حفظ كـ JPG
            unique_filename = f"{uuid.uuid4()}.jpg"
            filepath = os.path.join(property_folder, unique_filename)
            image.save(filepath, "JPEG", quality=85)
        else:
            file.save(filepath)
        
        return filepath
    except Exception as e:
        print(f"خطأ في معالجة الصورة: {e}")
        return None


@properties_bp.route('/dashboard')
@login_required
def dashboard():
    """لوحة التحكم الرئيسية للعقارات المستأجرة"""
    
    # إحصائيات العقارات
    total_properties = RentalProperty.query.filter_by(is_active=True).count()
    active_properties = RentalProperty.query.filter_by(status='active', is_active=True).count()
    
    # العقود المنتهية
    expired_properties = RentalProperty.query.filter(
        RentalProperty.contract_end_date < date.today(),
        RentalProperty.is_active == True
    ).count()
    
    # العقود القريبة من الانتهاء (60 يوم)
    expiring_soon_date = date.today() + timedelta(days=60)
    expiring_soon = RentalProperty.query.filter(
        RentalProperty.contract_end_date.between(date.today(), expiring_soon_date),
        RentalProperty.is_active == True
    ).count()
    
    # إجمالي الإيجار السنوي
    total_annual_rent = db.session.query(
        func.sum(RentalProperty.annual_rent_amount)
    ).filter_by(is_active=True, status='active').scalar() or 0
    
    # الدفعات المعلقة
    pending_payments = PropertyPayment.query.filter_by(status='pending').count()
    
    # الدفعات المتأخرة
    overdue_payments = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date < date.today()
    ).count()
    
    # إجمالي المدفوعات (المدفوعة فقط)
    total_paid = db.session.query(
        func.sum(PropertyPayment.amount)
    ).filter_by(status='paid').scalar() or 0
    
    # قائمة العقارات مع التفاصيل
    properties = RentalProperty.query.filter_by(is_active=True).order_by(
        RentalProperty.created_at.desc()
    ).all()
    
    # قائمة العقارات القريبة من الانتهاء
    expiring_properties = RentalProperty.query.filter(
        RentalProperty.contract_end_date.between(date.today(), expiring_soon_date),
        RentalProperty.is_active == True
    ).order_by(RentalProperty.contract_end_date).all()
    
    # الدفعات القادمة (خلال 30 يوم)
    upcoming_payments_date = date.today() + timedelta(days=30)
    upcoming_payments = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date.between(date.today(), upcoming_payments_date)
    ).order_by(PropertyPayment.payment_date).all()
    
    return render_template('properties/dashboard.html',
                         total_properties=total_properties,
                         active_properties=active_properties,
                         expired_properties=expired_properties,
                         expiring_soon=expiring_soon,
                         total_annual_rent=total_annual_rent,
                         pending_payments=pending_payments,
                         overdue_payments=overdue_payments,
                         total_paid=total_paid,
                         properties=properties,
                         expiring_properties=expiring_properties,
                         upcoming_payments=upcoming_payments)


@properties_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """إضافة عقار جديد"""
    form = RentalPropertyForm()
    
    if form.validate_on_submit():
        try:
            # إنشاء العقار
            property = RentalProperty(
                city=form.name.data,
                address=form.address.data,
                map_link='',
                contract_number=form.contract_number.data or None,  # استخدام None بدلاً من قيمة فارغة
                owner_name=form.landlord_name.data,
                owner_id=form.property_type.data,  # استخدام owner_id لحفظ نوع العقار مؤقتاً
                contract_start_date=form.contract_start_date.data,
                contract_end_date=form.contract_end_date.data,
                annual_rent_amount=form.monthly_rent.data * 12,
                includes_utilities=False,
                payment_method=form.payment_method.data,
                status='active',
                notes=form.notes.data,
                created_by=current_user.id
            )
            
            db.session.add(property)
            db.session.commit()
            
            # معالجة الصور إن وجدت
            if form.images.data:
                files = request.files.getlist('images')
                for file in files:
                    if file and allowed_file(file.filename):
                        filepath = process_and_save_image(file, property.id)
                        if filepath:
                            image = PropertyImage(
                                property_id=property.id,
                                image_path=filepath
                            )
                            db.session.add(image)
                db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='إضافة عقار مستأجر',
                entity_type='RentalProperty',
                entity_id=property.id,
                details=f'تم إضافة عقار جديد: {property.contract_number} - {property.city}'
            )
            
            flash('تم إضافة العقار بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة العقار: {str(e)}', 'danger')
    
    return render_template('properties/create.html', form=form)


@properties_bp.route('/<int:property_id>')
@login_required
def view(property_id):
    """عرض تفاصيل العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب الصور
    images = PropertyImage.query.filter_by(property_id=property_id).order_by(PropertyImage.uploaded_at.desc()).all()
    
    # جلب الدفعات
    payments = PropertyPayment.query.filter_by(property_id=property_id).order_by(PropertyPayment.payment_date.desc()).all()
    
    # جلب التجهيزات
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    return render_template('properties/view.html',
                         property=property,
                         images=images,
                         payments=payments,
                         furnishing=furnishing)


@properties_bp.route('/<int:property_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(property_id):
    """تعديل بيانات العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # ملء النموذج بالبيانات الحالية
    form = RentalPropertyForm()
    if request.method == 'GET':
        form.name.data = property.city
        form.property_type.data = property.owner_id
        form.address.data = property.address
        form.contract_number.data = property.contract_number
        form.landlord_name.data = property.owner_name
        form.landlord_phone.data = ''
        form.contract_start_date.data = property.contract_start_date
        form.contract_end_date.data = property.contract_end_date
        form.monthly_rent.data = property.annual_rent_amount / 12
        form.payment_method.data = property.payment_method
        form.notes.data = property.notes
    
    if form.validate_on_submit():
        try:
            property.city = form.name.data
            property.address = form.address.data
            property.map_link = ''
            property.contract_number = form.contract_number.data or None  # استخدام None بدلاً من قيمة فارغة
            property.owner_name = form.landlord_name.data
            property.owner_id = form.property_type.data  # استخدام owner_id لحفظ نوع العقار مؤقتاً
            property.contract_start_date = form.contract_start_date.data
            property.contract_end_date = form.contract_end_date.data
            property.annual_rent_amount = form.monthly_rent.data * 12
            property.includes_utilities = False
            property.payment_method = form.payment_method.data
            property.status = 'active'
            property.notes = form.notes.data
            
            db.session.commit()
            
            # معالجة الصور الجديدة إن وجدت
            if form.images.data:
                files = request.files.getlist('images')
                for file in files:
                    if file and allowed_file(file.filename):
                        filepath = process_and_save_image(file, property.id)
                        if filepath:
                            image = PropertyImage(
                                property_id=property.id,
                                image_path=filepath
                            )
                            db.session.add(image)
                db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='تعديل عقار مستأجر',
                entity_type='RentalProperty',
                entity_id=property.id,
                details=f'تم تعديل العقار: {property.contract_number}'
            )
            
            flash('تم تحديث بيانات العقار بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث البيانات: {str(e)}', 'danger')
    
    return render_template('properties/edit.html', form=form, property=property)


@properties_bp.route('/<int:property_id>/delete', methods=['POST'])
@login_required
def delete(property_id):
    """حذف العقار (حذف منطقي)"""
    property = RentalProperty.query.get_or_404(property_id)
    
    try:
        property.is_active = False
        db.session.commit()
        
        # تسجيل النشاط
        log_activity(
            action='حذف عقار مستأجر',
            entity_type='RentalProperty',
            entity_id=property.id,
            details=f'تم حذف العقار: {property.contract_number}'
        )
        
        flash('تم حذف العقار بنجاح!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف العقار: {str(e)}', 'danger')
    
    return redirect(url_for('properties.dashboard'))


@properties_bp.route('/<int:property_id>/images/upload', methods=['POST'])
@login_required
def upload_images(property_id):
    """رفع صور العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    image_type = request.form.get('image_type', 'أخرى')
    description = request.form.get('description', '')
    
    files = request.files.getlist('images')
    
    if not files:
        flash('الرجاء اختيار صور للرفع', 'warning')
        return redirect(url_for('properties.view', property_id=property_id))
    
    uploaded_count = 0
    for file in files:
        if file and allowed_file(file.filename):
            filepath = process_and_save_image(file, property_id)
            if filepath:
                image = PropertyImage(
                    property_id=property_id,
                    image_path=filepath,
                    image_type=image_type,
                    description=description
                )
                db.session.add(image)
                uploaded_count += 1
    
    try:
        db.session.commit()
        flash(f'تم رفع {uploaded_count} صورة بنجاح!', 'success')
        
        # تسجيل النشاط
        log_activity(
            action='رفع صور عقار',
            entity_type='RentalProperty',
            entity_id=property.id,
            details=f'تم رفع {uploaded_count} صورة للعقار: {property.contract_number}'
        )
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء رفع الصور: {str(e)}', 'danger')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/images/<int:image_id>/delete', methods=['POST'])
@login_required
def delete_image(image_id):
    """حذف صورة"""
    image = PropertyImage.query.get_or_404(image_id)
    property_id = image.property_id
    
    try:
        # حذف الملف الفعلي
        if os.path.exists(image.image_path):
            os.remove(image.image_path)
        
        db.session.delete(image)
        db.session.commit()
        
        flash('تم حذف الصورة بنجاح!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الصورة: {str(e)}', 'danger')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/<int:property_id>/payments/add', methods=['GET', 'POST'])
@login_required
def add_payment(property_id):
    """إضافة دفعة جديدة"""
    property = RentalProperty.query.get_or_404(property_id)
    form = PropertyPaymentForm()
    
    if form.validate_on_submit():
        try:
            payment = PropertyPayment(
                property_id=property_id,
                payment_date=form.payment_date.data,
                amount=form.amount.data,
                status=form.status.data,
                actual_payment_date=form.actual_payment_date.data,
                payment_method=form.payment_method.data,
                reference_number=form.reference_number.data,
                notes=form.notes.data
            )
            
            db.session.add(payment)
            db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='إضافة دفعة إيجار',
                entity_type='PropertyPayment',
                entity_id=payment.id,
                details=f'تم إضافة دفعة بقيمة {payment.amount} ريال للعقار: {property.contract_number}'
            )
            
            flash('تم إضافة الدفعة بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة الدفعة: {str(e)}', 'danger')
    
    return render_template('properties/add_payment.html', form=form, property=property)


@properties_bp.route('/payments/<int:payment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(payment_id):
    """تعديل دفعة"""
    payment = PropertyPayment.query.get_or_404(payment_id)
    property = payment.rental_property
    form = PropertyPaymentForm(obj=payment)
    
    if form.validate_on_submit():
        try:
            payment.payment_date = form.payment_date.data
            payment.amount = form.amount.data
            payment.status = form.status.data
            payment.actual_payment_date = form.actual_payment_date.data
            payment.payment_method = form.payment_method.data
            payment.reference_number = form.reference_number.data
            payment.notes = form.notes.data
            
            db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='تعديل دفعة إيجار',
                entity_type='PropertyPayment',
                entity_id=payment.id,
                details=f'تم تعديل دفعة بقيمة {payment.amount} ريال للعقار: {property.contract_number}'
            )
            
            flash('تم تحديث الدفعة بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث الدفعة: {str(e)}', 'danger')
    
    return render_template('properties/edit_payment.html', form=form, payment=payment, property=property)


@properties_bp.route('/payments/<int:payment_id>/delete', methods=['POST'])
@login_required
def delete_payment(payment_id):
    """حذف دفعة"""
    payment = PropertyPayment.query.get_or_404(payment_id)
    property_id = payment.property_id
    
    try:
        db.session.delete(payment)
        db.session.commit()
        
        flash('تم حذف الدفعة بنجاح!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الدفعة: {str(e)}', 'danger')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/<int:property_id>/furnishing', methods=['GET', 'POST'])
@login_required
def manage_furnishing(property_id):
    """إدارة تجهيزات العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    if not furnishing:
        furnishing = PropertyFurnishing(property_id=property_id)
    
    form = PropertyFurnishingForm(obj=furnishing)
    
    if form.validate_on_submit():
        try:
            furnishing.gas_cylinder = form.gas_cylinder.data or 0
            furnishing.stoves = form.stoves.data or 0
            furnishing.beds = form.beds.data or 0
            furnishing.blankets = form.blankets.data or 0
            furnishing.pillows = form.pillows.data or 0
            furnishing.other_items = form.other_items.data
            furnishing.notes = form.notes.data
            
            if not furnishing.id:
                db.session.add(furnishing)
            
            db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='تحديث تجهيزات عقار',
                entity_type='PropertyFurnishing',
                entity_id=furnishing.id,
                details=f'تم تحديث تجهيزات العقار: {property.contract_number}'
            )
            
            flash('تم تحديث التجهيزات بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث التجهيزات: {str(e)}', 'danger')
    
    return render_template('properties/furnishing.html', form=form, property=property, furnishing=furnishing)


@properties_bp.route('/<int:property_id>/export-excel')
@login_required
def export_excel(property_id):
    """تصدير بيانات العقار إلى Excel"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب البيانات المرتبطة
    payments = PropertyPayment.query.filter_by(property_id=property_id).order_by(PropertyPayment.payment_date).all()
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات العقار"
    
    # تعريف الألوان والتنسيقات
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان الرئيسي
    ws.merge_cells('A1:D1')
    ws['A1'] = f"تقرير العقار: {property.city}"
    ws['A1'].font = Font(bold=True, size=16, color="1F4788")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # معلومات العقار الأساسية
    ws['A3'] = "معلومات العقار"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:B3')
    
    property_data = [
        ('اسم العقار', property.city),
        ('نوع العقار', {'apartment': 'شقة', 'villa': 'فيلا', 'building': 'عمارة', 
                       'full_floor': 'دور كامل', 'office': 'مكتب', 'warehouse': 'مستودع'}.get(property.owner_id, '-')),
        ('العنوان', property.address),
        ('رقم العقد', property.contract_number or '-'),
        ('اسم المالك', property.owner_name),
        ('تاريخ البداية', property.contract_start_date.strftime('%Y-%m-%d')),
        ('تاريخ الانتهاء', property.contract_end_date.strftime('%Y-%m-%d')),
        ('الإيجار السنوي', f"{property.annual_rent_amount:,.0f} ريال"),
        ('طريقة الدفع', {'monthly': 'شهري', 'quarterly': 'ربع سنوي', 
                        'semi_annually': 'نصف سنوي', 'annually': 'سنوي'}.get(property.payment_method, '-')),
    ]
    
    row = 4
    for label, value in property_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # التجهيزات
    if furnishing:
        row += 1
        ws[f'A{row}'] = "التجهيزات"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        furnishing_data = [
            ('جرات الغاز', furnishing.gas_cylinder),
            ('الطباخات', furnishing.stoves),
            ('الأسرّة', furnishing.beds),
            ('البطانيات', furnishing.blankets),
            ('المخدات', furnishing.pillows),
        ]
        
        for label, value in furnishing_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
    
    # الدفعات
    if payments:
        row += 2
        ws[f'A{row}'] = "الدفعات"
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # عناوين جدول الدفعات
        headers = ['التاريخ المتوقع', 'المبلغ', 'الحالة', 'تاريخ الدفع الفعلي', 'ملاحظات']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # بيانات الدفعات
        for payment in payments:
            ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=f"{payment.amount:,.2f} ريال").border = border
            status_text = {'pending': 'معلق', 'paid': 'مدفوع', 'overdue': 'متأخر'}.get(payment.status, '-')
            ws.cell(row=row, column=3, value=status_text).border = border
            ws.cell(row=row, column=4, value=payment.actual_payment_date.strftime('%Y-%m-%d') if payment.actual_payment_date else '-').border = border
            ws.cell(row=row, column=5, value=payment.notes or '-').border = border
            row += 1
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    
    # حفظ في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # اسم الملف
    filename = f"عقار_{property.city}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
