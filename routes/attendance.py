from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required
from sqlalchemy import func, extract
from datetime import datetime, time, timedelta, date
from app import db
from models import Attendance, Employee, Department, SystemAudit, VehicleProject, Module, Permission, employee_departments
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from utils.excel import export_attendance_by_department
from utils.excel_dashboard import export_attendance_by_department_with_dashboard
from utils.user_helpers import check_module_access
from utils.audit_logger import log_attendance_activity, log_system_activity, log_activity
import calendar
import logging
import time as time_module  # Renamed to avoid conflict with datetime.time

# Setup logging
logger = logging.getLogger(__name__)

attendance_bp = Blueprint('attendance', __name__)

@attendance_bp.route('/')
def index():
    """List attendance records with filtering options"""
    # Get filter parameters
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # Parse date
    try:
        date = parse_date(date_str)
    except ValueError:
        date = datetime.now().date()
    
    # Build query
    query = Attendance.query.filter(Attendance.date == date)
    base_query = query  # Save the base query for statistics
    
    # Apply filters for the main query - استخدام علاقة many-to-many
    if department_id and department_id != '':
        # استخدام العلاقة many-to-many بين الموظفين والأقسام
        query = query.join(Employee).join(employee_departments).filter(employee_departments.c.department_id == int(department_id))
        base_query = base_query.join(Employee).join(employee_departments).filter(employee_departments.c.department_id == int(department_id))
    
    if status and status != '':
        query = query.filter(Attendance.status == status)
    
    # Execute query
    attendances = query.all()
    
    # Get departments for filter dropdown based on user permissions
    from flask_login import current_user
    
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
        
        # إذا كان لدى المستخدم قسم مخصص، فلتر البيانات تلقائياً لهذا القسم - استخدام many-to-many
        if current_user.assigned_department_id and not department_id:
            department_id = str(current_user.assigned_department_id)
            query = query.join(Employee).join(employee_departments).filter(employee_departments.c.department_id == current_user.assigned_department_id)
            base_query = base_query.join(Employee).join(employee_departments).filter(employee_departments.c.department_id == current_user.assigned_department_id)
            attendances = query.all()
    else:
        departments = Department.query.all()
    
    # Calculate attendance statistics for the current date and department filter
    present_count = base_query.filter(Attendance.status == 'present').count()
    absent_count = base_query.filter(Attendance.status == 'absent').count()
    leave_count = base_query.filter(Attendance.status == 'leave').count()
    sick_count = base_query.filter(Attendance.status == 'sick').count()
    
    # Format date for display in both calendars
    hijri_date = format_date_hijri(date)
    gregorian_date = format_date_gregorian(date)
    
    return render_template('attendance/index.html', 
                          attendances=attendances,
                          departments=departments,
                          date=date,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date,
                          selected_department=department_id,
                          selected_status=status,
                          present_count=present_count,
                          absent_count=absent_count,
                          leave_count=leave_count,
                          sick_count=sick_count)

@attendance_bp.route('/record', methods=['GET', 'POST'])
def record():
    """Record attendance for individual employees"""
    if request.method == 'POST':
        try:
            employee_id = request.form['employee_id']
            date_str = request.form['date']
            status = request.form['status']
            
            # Parse date
            date = parse_date(date_str)
            
            # Check if attendance record already exists
            existing = Attendance.query.filter_by(
                employee_id=employee_id,
                date=date
            ).first()
            
            if existing:
                # Update existing record
                existing.status = status
                existing.notes = request.form.get('notes', '')
                
                # Process check-in and check-out times if present
                if status == 'present':
                    check_in_str = request.form.get('check_in', '')
                    check_out_str = request.form.get('check_out', '')
                    
                    if check_in_str:
                        hours, minutes = map(int, check_in_str.split(':'))
                        existing.check_in = time(hours, minutes)
                    
                    if check_out_str:
                        hours, minutes = map(int, check_out_str.split(':'))
                        existing.check_out = time(hours, minutes)
                else:
                    existing.check_in = None
                    existing.check_out = None
                
                db.session.commit()
                
                # تسجيل العملية في سجل النشاط
                employee = Employee.query.get(employee_id)
                if employee:
                    log_attendance_activity(
                        action='update',
                        attendance_data={
                            'id': existing.id,
                            'employee_id': employee_id,
                            'date': date.isoformat(),
                            'status': status
                        },
                        employee_name=employee.name
                    )
                
                flash('تم تحديث سجل الحضور بنجاح', 'success')
            else:
                # Create new attendance record
                new_attendance = Attendance(
                    employee_id=employee_id,
                    date=date,
                    status=status,
                    notes=request.form.get('notes', '')
                )
                
                # Process check-in and check-out times if present and status is 'present'
                if status == 'present':
                    check_in_str = request.form.get('check_in', '')
                    check_out_str = request.form.get('check_out', '')
                    
                    if check_in_str:
                        hours, minutes = map(int, check_in_str.split(':'))
                        new_attendance.check_in = time(hours, minutes)
                    
                    if check_out_str:
                        hours, minutes = map(int, check_out_str.split(':'))
                        new_attendance.check_out = time(hours, minutes)
                
                db.session.add(new_attendance)
                db.session.commit()
                
                # تسجيل العملية في سجل النشاط
                employee = Employee.query.get(employee_id)
                if employee:
                    log_attendance_activity(
                        action='create',
                        attendance_data={
                            'id': new_attendance.id,
                            'employee_id': employee_id,
                            'date': date.isoformat(),
                            'status': status
                        },
                        employee_name=employee.name
                    )
                
                flash('تم تسجيل الحضور بنجاح', 'success')
            
            return redirect(url_for('attendance.index', date=date_str))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # الحصول على الموظفين النشطين حسب صلاحيات المستخدم
    from flask_login import current_user
    
    employees = []
    if current_user.is_authenticated:
        try:
            user_role = current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)
            print(f"المستخدم الحالي: {current_user.name}, الدور: {user_role}, القسم المخصص: {current_user.assigned_department_id}")
            
            if user_role in ['ADMIN', 'MANAGER', 'SUPERVISOR']:
                # المديرون والمشرفون يمكنهم رؤية جميع الموظفين
                employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
                print(f"تم العثور على {len(employees)} موظف نشط للمدير/المشرف")
            elif current_user.assigned_department_id:
                # المستخدمون مع قسم مخصص يرون موظفي قسمهم فقط
                employees = Employee.query.filter_by(
                    status='active',
                    department_id=current_user.assigned_department_id
                ).order_by(Employee.name).all()
                print(f"تم العثور على {len(employees)} موظف نشط للقسم {current_user.assigned_department_id}")
            else:
                # كحل بديل، عرض جميع الموظفين للمستخدمين المسجلين
                employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
                print(f"عرض جميع الموظفين كحل بديل: {len(employees)} موظف")
        except Exception as e:
            print(f"خطأ في تحديد صلاحيات المستخدم: {e}")
            # كحل بديل في حالة الخطأ، عرض جميع الموظفين
            employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
    else:
        print("المستخدم غير مسجل دخول")
    
    # Default to today's date
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    return render_template('attendance/record.html', 
                          employees=employees, 
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/department', methods=['GET', 'POST'])
def department_attendance():
    """Record attendance for an entire department at once"""
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            date_str = request.form['date']
            status = request.form['status']
            
            # التحقق من صلاحيات المستخدم للوصول إلى هذا القسم
            from flask_login import current_user
            
            if current_user.is_authenticated and not current_user.can_access_department(int(department_id)):
                flash('ليس لديك صلاحية لتسجيل حضور هذا القسم', 'error')
                return redirect(url_for('attendance.department_attendance'))
            
            # Parse date
            date = parse_date(date_str)
            
            # Get all employees in the department using many-to-many relationship
            department = Department.query.get_or_404(department_id)
            employees = [emp for emp in department.employees if emp.status == 'active']
            
            count = 0
            for employee in employees:
                # Check if attendance record already exists
                existing = Attendance.query.filter_by(
                    employee_id=employee.id,
                    date=date
                ).first()
                
                if existing:
                    # Update existing record
                    existing.status = status
                    if status != 'present':
                        existing.check_in = None
                        existing.check_out = None
                else:
                    # Create new attendance record
                    new_attendance = Attendance(
                        employee_id=employee.id,
                        date=date,
                        status=status
                    )
                    db.session.add(new_attendance)
                
                count += 1
            
            # تسجيل العملية في سجل النشاط
            department = Department.query.get(department_id)
            if department:
                log_attendance_activity(
                    action='bulk_create',
                    attendance_data={
                        'department_id': department_id,
                        'date': date.isoformat(),
                        'status': status,
                        'count': count
                    },
                    employee_name=f"جميع موظفي قسم {department.name}"
                )
            
            flash(f'تم تسجيل الحضور لـ {count} موظف بنجاح', 'success')
            return redirect(url_for('attendance.index', date=date_str))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # Get departments based on user permissions
    from flask_login import current_user
    
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
    else:
        departments = Department.query.all()
    
    # Default to today's date
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    return render_template('attendance/department.html', 
                          departments=departments,
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/bulk-record', methods=['GET', 'POST'])
def bulk_record():
    """تسجيل الحضور الجماعي للموظفين بفترات مختلفة"""
    from flask_login import current_user
    
    # التحقق من تسجيل الدخول
    if not current_user.is_authenticated:
        flash('يجب تسجيل الدخول أولاً', 'error')
        return redirect(url_for('auth.login'))
    
    # التحقق من وجود قسم مخصص للمستخدم
    if not current_user.assigned_department_id and current_user.role.value != 'admin':
        flash('يجب تخصيص قسم لك لتتمكن من تسجيل الحضور', 'error')
        return redirect(url_for('attendance.index'))
    
    if request.method == 'POST':
        try:
            period_type = request.form['period_type']
            default_status = request.form['default_status']
            employee_ids = request.form.getlist('employee_ids')
            skip_weekends = 'skip_weekends' in request.form
            overwrite_existing = 'overwrite_existing' in request.form
            
            if not employee_ids:
                flash('يجب اختيار موظف واحد على الأقل', 'error')
                return redirect(url_for('attendance.bulk_record'))
            
            # تحديد التواريخ حسب نوع الفترة
            dates = []
            
            if period_type == 'daily':
                single_date = parse_date(request.form['single_date'])
                dates = [single_date]
                
            elif period_type == 'weekly':
                week_start = parse_date(request.form['week_start'])
                dates = [week_start + timedelta(days=i) for i in range(7)]
                
            elif period_type == 'monthly':
                month_year = request.form['month_year']
                year, month = map(int, month_year.split('-'))
                import calendar
                days_in_month = calendar.monthrange(year, month)[1]
                dates = [date(year, month, day) for day in range(1, days_in_month + 1)]
                
            elif period_type == 'custom':
                start_date = parse_date(request.form['start_date'])
                end_date = parse_date(request.form['end_date'])
                current_date = start_date
                while current_date <= end_date:
                    dates.append(current_date)
                    current_date += timedelta(days=1)
            
            # تصفية عطلة نهاية الأسبوع إذا كان مطلوباً
            if skip_weekends:
                dates = [d for d in dates if d.weekday() not in [4, 5]]  # الجمعة والسبت
            
            # تسجيل الحضور
            count = 0
            for employee_id in employee_ids:
                for date_obj in dates:
                    # التحقق من وجود سجل سابق
                    existing = Attendance.query.filter_by(
                        employee_id=employee_id,
                        date=date_obj
                    ).first()
                    
                    if existing:
                        if overwrite_existing:
                            existing.status = default_status
                            count += 1
                    else:
                        attendance = Attendance(
                            employee_id=employee_id,
                            date=date_obj,
                            status=default_status
                        )
                        db.session.add(attendance)
                        count += 1
            
            db.session.commit()
            flash(f'تم تسجيل {count} سجل حضور بنجاح', 'success')
            return redirect(url_for('attendance.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # الحصول على موظفي القسم المخصص للمستخدم
    try:
        print(f"معلومات المستخدم: ID={current_user.id}, assigned_department_id={getattr(current_user, 'assigned_department_id', None)}")
        
        if hasattr(current_user, 'role') and current_user.role and current_user.role.value == 'admin':
            # المديرون العامون يمكنهم رؤية جميع الموظفين
            employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
            print(f"مدير عام - تم جلب {len(employees)} موظف")
        elif current_user.assigned_department_id:
            # المستخدمون مع قسم مخصص يرون موظفي قسمهم فقط
            employees = Employee.query.filter_by(
                status='active',
                department_id=current_user.assigned_department_id
            ).order_by(Employee.name).all()
            print(f"مستخدم قسم {current_user.assigned_department_id} - تم جلب {len(employees)} موظف")
        else:
            # جرب جلب جميع الموظفين النشطين للاختبار
            employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
            print(f"مستخدم بدون قسم - تم جلب {len(employees)} موظف للاختبار")
    except Exception as e:
        print(f"خطأ في جلب الموظفين: {str(e)}")
        employees = []
    
    today = datetime.now().date()
    
    return render_template('attendance/bulk_record.html', 
                         employees=employees,
                         today=today)

@attendance_bp.route('/all-departments-simple', methods=['GET', 'POST'])
def all_departments_attendance_simple():
    """تسجيل حضور لعدة أقسام لفترة زمنية محددة - نسخة مبسطة"""
    # التاريخ الافتراضي هو اليوم
    today = datetime.now().date()
    
    # CSRF Token - نستخدم المعالج الافتراضي
    from flask_wtf.csrf import generate_csrf
    form = {"csrf_token": generate_csrf()}
    
    if request.method == 'POST':
        try:
            # 1. استلام البيانات من النموذج
            department_ids = request.form.getlist('department_ids')
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            status = request.form.get('status', 'present')  # القيمة الافتراضية هي حاضر
            
            # 2. التحقق من البيانات المدخلة
            if not department_ids:
                flash('الرجاء اختيار قسم واحد على الأقل', 'warning')
                return redirect(url_for('attendance.all_departments_attendance_simple'))
                
            try:
                # التأكد من أن السلاسل ليست None
                if not start_date_str or not end_date_str:
                    flash('تنسيق التاريخ غير صحيح', 'danger')
                    return redirect(url_for('attendance.all_departments_attendance_simple'))
                
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('تنسيق التاريخ غير صحيح', 'danger')
                return redirect(url_for('attendance.all_departments_attendance_simple'))
                
            if end_date < start_date:
                flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية أو مساوياً له', 'warning')
                return redirect(url_for('attendance.all_departments_attendance_simple'))
                
            # 3. تهيئة المتغيرات للإحصائيات
            total_departments = 0
            total_employees = 0
            total_records = 0
            
            # 4. حساب عدد الأيام
            delta = end_date - start_date
            days_count = delta.days + 1  # لتضمين اليوم الأخير
            
            # 5. معالجة البيانات
            for dept_id in department_ids:
                try:
                    # التحقق من وجود القسم
                    department = Department.query.get(int(dept_id))
                    if not department:
                        continue
                        
                    total_departments += 1
                    
                    # الحصول على الموظفين النشطين في القسم
                    employees = Employee.query.filter_by(
                        department_id=int(dept_id),
                        status='active'
                    ).all()
                    
                    # عدد موظفي القسم
                    dept_employees_count = len(employees)
                    total_employees += dept_employees_count
                    dept_records = 0
                    
                    # معالجة كل يوم في نطاق التاريخ
                    curr_date = start_date
                    while curr_date <= end_date:
                        for employee in employees:
                            # التحقق من وجود سجل حضور سابق
                            existing = Attendance.query.filter_by(
                                employee_id=employee.id,
                                date=curr_date
                            ).first()
                            
                            if existing:
                                # تحديث السجل الموجود
                                existing.status = status
                                if status != 'present':
                                    existing.check_in = None
                                    existing.check_out = None
                            else:
                                # إنشاء سجل جديد
                                new_attendance = Attendance()
                                new_attendance.employee_id = employee.id
                                new_attendance.date = curr_date
                                new_attendance.status = status
                                if status == 'present':
                                    # يمكن إضافة وقت الدخول والخروج الافتراضي إذا كان حاضر
                                    pass
                                db.session.add(new_attendance)
                                
                            dept_records += 1
                            
                        # الانتقال لليوم التالي
                        curr_date += timedelta(days=1)
                    
                    total_records += dept_records
                    
                    try:
                        # تسجيل النشاط للقسم في سجل النظام
                        from flask_login import current_user
                        user_id = current_user.id if hasattr(current_user, 'id') else None
                        
                        SystemAudit.create_audit_record(
                            user_id=user_id,
                            action='mass_attendance',
                            entity_type='department',
                            entity_id=department.id,
                            entity_name=department.name,
                            details=f'تم تسجيل حضور لقسم {department.name} للفترة من {start_date} إلى {end_date} لعدد {dept_employees_count} موظف'
                        )
                    except Exception as audit_error:
                        # تخطي خطأ السجل
                        print(f"خطأ في تسجيل النشاط: {str(audit_error)}")
                        
                except Exception as dept_error:
                    print(f"خطأ في معالجة القسم رقم {dept_id}: {str(dept_error)}")
                    continue
            
            # حفظ التغييرات في قاعدة البيانات
            db.session.commit()
            
            # تسجيل العملية
            departments_names = [Department.query.get(dept_id).name for dept_id in department_ids if Department.query.get(dept_id)]
            log_activity('create', 'BulkAttendance', None, 
                        f"تم تسجيل حضور جماعي لـ {total_departments} قسم ({', '.join(departments_names[:3])}{'...' if len(departments_names) > 3 else ''}) - {total_employees} موظف عن {days_count} يوم")
            
            # عرض رسالة نجاح
            flash(f'تم تسجيل الحضور بنجاح لـ {total_departments} قسم و {total_employees} موظف عن {days_count} يوم (إجمالي {total_records} سجل)', 'success')
            return redirect(url_for('attendance.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء معالجة البيانات: {str(e)}', 'danger')
            print(f"خطأ: {str(e)}")
    
    # الحصول على الأقسام مع عدد الموظفين النشطين لكل قسم
    departments = []
    all_departments = Department.query.all()
    for dept in all_departments:
        active_count = Employee.query.filter_by(department_id=dept.id, status='active').count()
        # أضف جميع الأقسام حتى لو لم يكن لديها موظفين نشطين
        dept.active_employees_count = active_count
        departments.append(dept)
    
    return render_template('attendance/all_departments_simple.html',
                          departments=departments,
                          today=today,
                          form=form)

@attendance_bp.route('/all-departments', methods=['GET', 'POST'])
def all_departments_attendance():
    """تسجيل حضور لعدة أقسام لفترة زمنية محددة"""
    # التاريخ الافتراضي هو اليوم
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    if request.method == 'POST':
        try:
            department_ids = request.form.getlist('department_ids')
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            status = request.form.get('status')
            
            if not department_ids:
                flash('يجب اختيار قسم واحد على الأقل.', 'danger')
                return redirect(url_for('attendance.all_departments_attendance'))
                
            if not start_date_str or not end_date_str:
                flash('يجب تحديد تاريخ البداية والنهاية.', 'danger')
                return redirect(url_for('attendance.all_departments_attendance'))
                
            if not status:
                status = 'present'  # الحالة الافتراضية هي حاضر
                
            # تحليل التواريخ
            try:
                start_date = parse_date(start_date_str)
                end_date = parse_date(end_date_str)
                
                if not start_date or not end_date:
                    flash('تاريخ غير صالح، يرجى التحقق من التنسيق.', 'danger')
                    return redirect(url_for('attendance.all_departments_attendance'))
                
                # التحقق من صحة النطاق
                if end_date < start_date:
                    flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية أو مساوياً له.', 'danger')
                    return redirect(url_for('attendance.all_departments_attendance'))
            except (ValueError, TypeError) as e:
                flash(f'خطأ في تنسيق التاريخ: {str(e)}', 'danger')
                return redirect(url_for('attendance.all_departments_attendance'))
                
            # تهيئة المتغيرات للإحصائيات
            total_departments = len(department_ids)
            total_employees = 0
            total_records = 0
            
            # حساب عدد الأيام
            delta = end_date - start_date
            days_count = delta.days + 1  # لتضمين اليوم الأخير
            
            try:
                # العمل على كل قسم من الأقسام المحددة
                for department_id in department_ids:
                    try:
                        # التحويل إلى عدد صحيح
                        dept_id = int(department_id)
                        
                        # الحصول على القسم للتأكد من وجوده
                        department = Department.query.get(dept_id)
                        if not department:
                            continue
                            
                        # الحصول على جميع الموظفين في القسم
                        employees = Employee.query.filter_by(
                            department_id=dept_id,
                            status='active'
                        ).all()
                        
                        # عدد موظفي القسم
                        department_employee_count = len(employees)
                        total_employees += department_employee_count
                        
                        # التحضير لكل يوم في النطاق المحدد
                        current_date = start_date
                        department_records = 0
                        
                        while current_date <= end_date:
                            day_count = 0
                            
                            for employee in employees:
                                try:
                                    # التحقق من وجود سجل حضور مسبق
                                    existing = Attendance.query.filter_by(
                                        employee_id=employee.id,
                                        date=current_date
                                    ).first()
                                    
                                    if existing:
                                        # تحديث السجل الموجود
                                        existing.status = status
                                        if status != 'present':
                                            existing.check_in = None
                                            existing.check_out = None
                                    else:
                                        # إنشاء سجل حضور جديد
                                        new_attendance = Attendance(
                                            employee_id=employee.id,
                                            date=current_date,
                                            status=status
                                        )
                                        db.session.add(new_attendance)
                                    
                                    day_count += 1
                                except Exception as emp_error:
                                    # تسجيل الخطأ والاستمرار مع الموظف التالي
                                    print(f"خطأ مع الموظف {employee.id}: {str(emp_error)}")
                                    continue
                            
                            # الانتقال إلى اليوم التالي
                            current_date += timedelta(days=1)
                            department_records += day_count
                        
                        total_records += department_records
                        
                        # تسجيل العملية للقسم
                        log_activity('create', 'DepartmentAttendance', department.id, 
                                   f'تم تسجيل حضور لقسم {department.name} للفترة من {start_date} إلى {end_date} لعدد {department_employee_count} موظف')
                    
                    except Exception as dept_error:
                        # تسجيل الخطأ والاستمرار مع القسم التالي
                        print(f"خطأ مع القسم {department_id}: {str(dept_error)}")
                        continue
                
                # حفظ جميع التغييرات
                db.session.commit()
                
                # رسالة نجاح مفصلة
                flash(f'تم تسجيل الحضور لـ {total_departments} قسم و {total_employees} موظف عن {days_count} يوم بنجاح (إجمالي {total_records} سجل)', 'success')
                return redirect(url_for('attendance.index', date=start_date_str))
            
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء معالجة الأقسام: {str(e)}', 'danger')
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ عام: {str(e)}', 'danger')
    
    # الحصول على جميع الأقسام مع عدد الموظفين النشطين لكل قسم
    try:
        departments = []
        all_departments = Department.query.all()
        for dept in all_departments:
            active_count = Employee.query.filter_by(department_id=dept.id, status='active').count()
            # أضف جميع الأقسام حتى لو لم يكن لديها موظفين نشطين
            dept.active_employees_count = active_count
            departments.append(dept)
    except Exception as e:
        departments = []
        flash(f'خطأ في تحميل الأقسام: {str(e)}', 'warning')
    
    return render_template('attendance/all_departments.html', 
                          departments=departments,
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/multi-day-department', methods=['GET', 'POST'])
def multi_day_department_attendance():
    """تسجيل حضور لقسم بالكامل لفترة زمنية محددة"""
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            start_date_str = request.form['start_date']
            end_date_str = request.form['end_date']
            status = request.form['status']
            
            # تحليل التواريخ
            try:
                start_date = parse_date(start_date_str)
                end_date = parse_date(end_date_str)
                
                if not start_date or not end_date:
                    raise ValueError("تاريخ غير صالح")
                
                # التحقق من صحة النطاق
                if end_date < start_date:
                    flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية أو مساوياً له.', 'danger')
                    return redirect(url_for('attendance.multi_day_department_attendance'))
            except (ValueError, TypeError) as e:
                flash(f'خطأ في تنسيق التاريخ: {str(e)}', 'danger')
                return redirect(url_for('attendance.multi_day_department_attendance'))
                
            # الحصول على جميع الموظفين في القسم
            employees = Employee.query.filter_by(
                department_id=department_id,
                status='active'
            ).all()
            
            # حساب عدد الأيام
            delta = end_date - start_date
            days_count = delta.days + 1  # لتضمين اليوم الأخير
            
            # التحضير لكل يوم في النطاق المحدد
            total_count = 0
            current_date = start_date
            
            while current_date <= end_date:
                day_count = 0
                
                for employee in employees:
                    # التحقق من وجود سجل حضور مسبق
                    existing = Attendance.query.filter_by(
                        employee_id=employee.id,
                        date=current_date
                    ).first()
                    
                    if existing:
                        # تحديث السجل الموجود
                        existing.status = status
                        if status != 'present':
                            existing.check_in = None
                            existing.check_out = None
                    else:
                        # إنشاء سجل حضور جديد
                        new_attendance = Attendance(
                            employee_id=employee.id,
                            date=current_date,
                            status=status
                        )
                        db.session.add(new_attendance)
                    
                    day_count += 1
                
                # الانتقال إلى اليوم التالي
                current_date += timedelta(days=1)
                total_count += day_count
            
            # تسجيل العملية
            department = Department.query.get(department_id)
            if department:
                log_activity('create', 'MultiDayDepartmentAttendance', department.id,
                           f'تم تسجيل حضور لقسم {department.name} للفترة من {start_date} إلى {end_date} لعدد {len(employees)} موظف و {days_count} يوم ({total_count} سجل)')
            
            flash(f'تم تسجيل الحضور لـ {len(employees)} موظف عن {days_count} يوم بنجاح (إجمالي {total_count} سجل)', 'success')
            return redirect(url_for('attendance.index', date=start_date_str))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # الحصول على جميع الأقسام
    departments = Department.query.all()
    
    # التاريخ الافتراضي هو اليوم
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    return render_template('attendance/multi_day_department.html', 
                          departments=departments,
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/delete/<int:id>', methods=['POST'])
def delete(id):
    """Delete an attendance record"""
    attendance = Attendance.query.get_or_404(id)
    
    try:
        # Get associated employee
        employee = Employee.query.get(attendance.employee_id)
        
        # Delete attendance record
        db.session.delete(attendance)
        
        # Log the action
        entity_name = employee.name if employee else f'ID: {id}'
        SystemAudit.create_audit_record(
            user_id=None,  # يمكن تعديلها لاستخدام current_user.id
            action='delete',
            entity_type='attendance',
            entity_id=id,
            entity_name=entity_name,
            details=f'تم حذف سجل حضور للموظف: {employee.name if employee else "غير معروف"} بتاريخ {attendance.date}'
        )
        db.session.commit()
        
        flash('تم حذف سجل الحضور بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('attendance.index', date=attendance.date))

@attendance_bp.route('/stats')
def stats():
    """Get attendance statistics for a date range"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    department_id = request.args.get('department_id', '')
    
    try:
        start_date = parse_date(start_date_str) if start_date_str else datetime.now().date().replace(day=1)
        end_date = parse_date(end_date_str) if end_date_str else datetime.now().date()
    except ValueError:
        start_date = datetime.now().date().replace(day=1)
        end_date = datetime.now().date()
    
    query = db.session.query(
        Attendance.status,
        func.count(Attendance.id).label('count')
    ).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    )
    
    if department_id and department_id != '':
        query = query.join(Employee).filter(Employee.department_id == department_id)
    
    stats = query.group_by(Attendance.status).all()
    
    # Convert to a dict for easier consumption by charts
    result = {'present': 0, 'absent': 0, 'leave': 0, 'sick': 0}
    for status, count in stats:
        result[status] = count
    
    return jsonify(result)

@attendance_bp.route('/export/excel', methods=['POST', 'GET'])
def export_excel():
    """تصدير بيانات الحضور إلى ملف Excel"""
    try:
        # الحصول على البيانات من النموذج حسب طريقة الطلب
        if request.method == 'POST':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            department_id = request.form.get('department_id')
        else:  # GET
            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')
            department_id = request.args.get('department_id')
        
        # التحقق من المدخلات
        if not start_date_str:
            flash('تاريخ البداية مطلوب', 'danger')
            return redirect(url_for('attendance.export_page'))
        
        # تحليل التواريخ
        try:
            start_date = parse_date(start_date_str)
            if end_date_str:
                end_date = parse_date(end_date_str)
            else:
                end_date = datetime.now().date()
        except (ValueError, TypeError):
            flash('تاريخ غير صالح', 'danger')
            return redirect(url_for('attendance.export_page'))
        
        # التحقق من اختيار القسم
        if department_id and department_id != '':
            # تصدير قسم واحد فقط
            department = Department.query.get(department_id)
            if not department:
                flash('القسم غير موجود', 'danger')
                return redirect(url_for('attendance.export_page'))
            
            employees = department.employees
            
            if employees:
                attendances = Attendance.query.filter(
                    Attendance.date.between(start_date, end_date),
                    Attendance.employee_id.in_([emp.id for emp in employees])
                ).all()
            else:
                attendances = []
            
            excel_file = export_attendance_by_department(employees, attendances, start_date, end_date)
            
            if end_date_str:
                filename = f'سجل الحضور - {department.name} - {start_date_str} إلى {end_date_str}.xlsx'
            else:
                filename = f'سجل الحضور - {department.name} - {start_date_str}.xlsx'
        else:
            # تصدير جميع الأقسام
            departments = Department.query.all()
            all_employees = Employee.query.filter_by(status='active').all()
            
            if all_employees:
                attendances = Attendance.query.filter(
                    Attendance.date.between(start_date, end_date),
                    Attendance.employee_id.in_([emp.id for emp in all_employees])
                ).all()
            else:
                attendances = []
            
            excel_file = export_attendance_by_department_with_dashboard(all_employees, attendances, start_date, end_date)
            
            if end_date_str:
                filename = f'سجل الحضور - جميع الأقسام - {start_date_str} إلى {end_date_str}.xlsx'
            else:
                filename = f'سجل الحضور - جميع الأقسام - {start_date_str}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('attendance.export_page'))

@attendance_bp.route('/export')
def export_page():
    """صفحة تصدير بيانات الحضور إلى ملف Excel"""
    departments = Department.query.all()
    today = datetime.now().date()
    start_of_month = today.replace(day=1)
    
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    hijri_start = format_date_hijri(start_of_month)
    gregorian_start = format_date_gregorian(start_of_month)
    
    return render_template('attendance/export.html',
                          departments=departments,
                          today=today,
                          start_of_month=start_of_month,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today,
                          hijri_start=hijri_start,
                          gregorian_start=gregorian_start)

@attendance_bp.route('/api/departments/<int:department_id>/employees')
def get_department_employees(department_id):
    """API endpoint to get all employees in a department"""
    try:
        # الحصول على القسم أولاً
        department = Department.query.get_or_404(department_id)
        
        # استخدام العلاقة many-to-many للحصول على جميع الموظفين النشطين
        employees = [emp for emp in department.employees if emp.status == 'active']
        
        employee_data = []
        for employee in employees:
            employee_data.append({
                'id': employee.id,
                'name': employee.name,
                'employee_id': employee.employee_id,
                'job_title': employee.job_title or 'غير محدد',
                'status': employee.status
            })
        
        logger.info(f"تم جلب {len(employee_data)} موظف نشط من القسم {department_id} ({department.name})")
        return jsonify(employee_data)
    
    except Exception as e:
        logger.error(f"خطأ في جلب موظفي القسم {department_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
@attendance_bp.route('/dashboard')
def dashboard():
    """لوحة معلومات الحضور مع إحصائيات يومية وأسبوعية وشهرية"""
    
    # إضافة آلية إعادة المحاولة للتعامل مع أخطاء الاتصال المؤقتة
    max_retries = 3  # عدد محاولات إعادة الاتصال
    retry_count = 0
    retry_delay = 1  # ثانية واحدة للمحاولة الأولى
    
    while retry_count < max_retries:
        try:
            # إعادة تحميل جميع البيانات من قاعدة البيانات للتأكد من أن البيانات محدثة
            db.session.expire_all()
            
            # 1. الحصول على المشروع المحدد (إذا وجد)
            project_name = request.args.get('project', None)
            
            # 2. الحصول على التاريخ الحالي
            today = datetime.now().date()
            current_month = today.month
            current_year = today.year
            
            # 3. حساب تاريخ بداية ونهاية الأسبوع الحالي بناءً على تاريخ بداية الشهر
            start_of_month = today.replace(day=1)  # أول يوم في الشهر الحالي
            
            # نحسب عدد الأيام منذ بداية الشهر حتى اليوم الحالي
            days_since_month_start = (today - start_of_month).days
            
            # نحسب عدد الأسابيع الكاملة منذ بداية الشهر (كل أسبوع 7 أيام)
            weeks_since_month_start = days_since_month_start // 7
            
            # نحسب بداية الأسبوع الحالي (بناءً على أسابيع من بداية الشهر)
            start_of_week = start_of_month + timedelta(days=weeks_since_month_start * 7)
            
            # نهاية الأسبوع بعد 6 أيام من البداية
            end_of_week = start_of_week + timedelta(days=6)
            
            # إذا كانت نهاية الأسبوع بعد نهاية الشهر، نجعلها آخر يوم في الشهر
            last_day = calendar.monthrange(current_year, current_month)[1]
            end_of_month = today.replace(day=last_day)
            if end_of_week > end_of_month:
                end_of_week = end_of_month
            
            # 4. حساب تاريخ بداية ونهاية الشهر الحالي
            start_of_month = today.replace(day=1)
            last_day = calendar.monthrange(current_year, current_month)[1]
            end_of_month = today.replace(day=last_day)
            
            # 5. إنشاء قاعدة الاستعلام
            query_base = db.session.query(
                Attendance.status,
                func.count(Attendance.id).label('count')
            )
            
            # 6. إحصائيات الحضور حسب المشروع أو عامة
            # تعريف قائمة معرفات الموظفين (سيكون None للكل)
            employee_ids = None
            
            if project_name:
                # استعلام للموظفين في مشروع محدد
                # نحتاج للحصول على قائمة الموظفين المرتبطين بالمشروع
                project_employees = db.session.query(Employee.id).filter(
                    Employee.project == project_name,
                    Employee.status == 'active'
                ).all()
                
                # تحويل النتائج إلى قائمة بسيطة من المعرفات
                employee_ids = [emp[0] for emp in project_employees]
            
            # تعريف ونهيئة متغيرات إحصائية
            daily_stats = []
            weekly_stats = []
            monthly_stats = []
            
            # إذا كان هناك مشروع محدد ولا يوجد موظفين فيه، نترك الإحصائيات فارغة
            if project_name and not employee_ids:
                # لا يوجد موظفين في هذا المشروع، نترك الإحصائيات فارغة
                pass
            else:
                # بناء استعلامات الإحصائيات إما لجميع الموظفين أو لموظفي مشروع محدد
                if employee_ids:
                    # إحصائيات الموظفين في المشروع المحدد
                    
                    # إحصائيات اليوم
                    daily_stats = query_base.filter(
                        Attendance.date == today,
                        Attendance.employee_id.in_(employee_ids)
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الأسبوع
                    weekly_stats = query_base.filter(
                        Attendance.date >= start_of_week,
                        Attendance.date <= end_of_week,
                        Attendance.employee_id.in_(employee_ids)
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الشهر
                    monthly_stats = query_base.filter(
                        Attendance.date >= start_of_month,
                        Attendance.date <= end_of_month,
                        Attendance.employee_id.in_(employee_ids)
                    ).group_by(Attendance.status).all()
                else:
                    # إحصائيات عامة لجميع الموظفين
                    
                    # إحصائيات اليوم
                    daily_stats = query_base.filter(
                        Attendance.date == today
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الأسبوع
                    weekly_stats = query_base.filter(
                        Attendance.date >= start_of_week,
                        Attendance.date <= end_of_week
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الشهر
                    monthly_stats = query_base.filter(
                        Attendance.date >= start_of_month,
                        Attendance.date <= end_of_month
                    ).group_by(Attendance.status).all()
            
            # 7. إحصائيات الحضور اليومي خلال الشهر الحالي لعرضها في المخطط البياني
            daily_attendance_data = []
            
            for day in range(1, last_day + 1):
                current_date = date(current_year, current_month, day)
                
                # تخطي التواريخ المستقبلية
                if current_date > today:
                    break
                    
                # استخدام employee_ids مباشرة بعد التأكد من أنه تم تعريفه في خطوة سابقة
                if employee_ids:
                    present_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'present',
                        Attendance.employee_id.in_(employee_ids)
                    ).scalar() or 0
                    
                    absent_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'absent',
                        Attendance.employee_id.in_(employee_ids)
                    ).scalar() or 0
                else:
                    present_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'present'
                    ).scalar() or 0
                    
                    absent_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'absent'
                    ).scalar() or 0
                    
                daily_attendance_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': str(day),
                    'present': present_count,
                    'absent': absent_count
                })
                
            # 8. الحصول على قائمة المشاريع النشطة للفلتر
            active_projects = db.session.query(Employee.project).filter(
                Employee.status == 'active',
                Employee.project.isnot(None)
            ).distinct().all()
            
            active_projects = [project[0] for project in active_projects if project[0]]
            
            # 9. تحويل البيانات إلى قاموس
            def stats_to_dict(stats_data):
                result = {'present': 0, 'absent': 0, 'leave': 0, 'sick': 0}
                for item in stats_data:
                    result[item[0]] = item[1]
                return result
            
            daily_stats_dict = stats_to_dict(daily_stats)
            weekly_stats_dict = stats_to_dict(weekly_stats)
            monthly_stats_dict = stats_to_dict(monthly_stats)
            
            # 10. إعداد البيانات للمخططات البيانية
            # 10.أ. مخطط توزيع الحضور اليومي
            daily_chart_data = {
                'labels': ['حاضر', 'غائب', 'إجازة', 'مرضي'],
                'datasets': [{
                    'data': [
                        daily_stats_dict['present'],
                        daily_stats_dict['absent'],
                        daily_stats_dict['leave'],
                        daily_stats_dict['sick']
                    ],
                    'backgroundColor': ['#28a745', '#dc3545', '#ffc107', '#17a2b8']
                }]
            }
            
            # 10.ب. مخطط توزيع الحضور الأسبوعي
            weekly_chart_data = {
                'labels': ['حاضر', 'غائب', 'إجازة', 'مرضي'],
                'datasets': [{
                    'data': [
                        weekly_stats_dict['present'],
                        weekly_stats_dict['absent'],
                        weekly_stats_dict['leave'],
                        weekly_stats_dict['sick']
                    ],
                    'backgroundColor': ['#28a745', '#dc3545', '#ffc107', '#17a2b8']
                }]
            }
            
            # 10.ج. مخطط توزيع الحضور الشهري
            monthly_chart_data = {
                'labels': ['حاضر', 'غائب', 'إجازة', 'مرضي'],
                'datasets': [{
                    'data': [
                        monthly_stats_dict['present'],
                        monthly_stats_dict['absent'],
                        monthly_stats_dict['leave'],
                        monthly_stats_dict['sick']
                    ],
                    'backgroundColor': ['#28a745', '#dc3545', '#ffc107', '#17a2b8']
                }]
            }
            
            # 10.د. مخطط الحضور اليومي خلال الشهر
            daily_trend_chart_data = {
                'labels': [item['day'] for item in daily_attendance_data],
                'datasets': [
                    {
                        'label': 'الحضور',
                        'data': [item['present'] for item in daily_attendance_data],
                        'backgroundColor': 'rgba(40, 167, 69, 0.2)',
                        'borderColor': 'rgba(40, 167, 69, 1)',
                        'borderWidth': 1,
                        'tension': 0.4
                    },
                    {
                        'label': 'الغياب',
                        'data': [item['absent'] for item in daily_attendance_data],
                        'backgroundColor': 'rgba(220, 53, 69, 0.2)',
                        'borderColor': 'rgba(220, 53, 69, 1)',
                        'borderWidth': 1,
                        'tension': 0.4
                    }
                ]
            }
            
            # 11. حساب معدل الحضور
            # إجمالي سجلات الحضور اليومية
            total_days = (
                daily_stats_dict['present'] + 
                daily_stats_dict['absent'] + 
                daily_stats_dict['leave'] + 
                daily_stats_dict['sick']
            )
            
            # إجمالي سجلات الحضور المتوقعة لليوم (جميع الموظفين النشطين)
            # حساب إجمالي الموظفين النشطين يتم في سطور لاحقة من الكود
            
            daily_attendance_rate = 0
            if total_days > 0:
                daily_attendance_rate = round((daily_stats_dict['present'] / total_days) * 100)
            
            # حساب إجمالي الموظفين النشطين - استخدام علاقة many-to-many الصحيحة
            if employee_ids:
                active_employees_count = len(employee_ids)
            else:
                # عد الموظفين النشطين المرتبطين بأقسام عبر علاقة many-to-many
                active_employees_count = db.session.query(func.count(func.distinct(Employee.id))).join(
                    employee_departments, Employee.id == employee_departments.c.employee_id
                ).filter(
                    Employee.status == 'active'
                ).scalar() or 0
            
            # حساب كامل الأسبوع (7 أيام) × عدد الموظفين النشطين
            # حساب عدد الأيام في الأسبوع (من بداية الأسبوع إلى نهايته)
            days_in_week = (end_of_week - start_of_week).days + 1
            
            # إجمالي سجلات الحضور والغياب في الأسبوع
            total_days_week = (
                weekly_stats_dict['present'] + 
                weekly_stats_dict['absent'] + 
                weekly_stats_dict['leave'] + 
                weekly_stats_dict['sick']
            )
            
            # حساب إجمالي سجلات الحضور المفترضة للأسبوع
            expected_days_week = days_in_week * active_employees_count
            
            weekly_attendance_rate = 0
            if total_days_week > 0:
                weekly_attendance_rate = round((weekly_stats_dict['present'] / total_days_week) * 100)
            
            # حساب معدل الحضور الشهري
            # إجمالي سجلات الحضور والغياب في الشهر
            total_days_month = (
                monthly_stats_dict['present'] + 
                monthly_stats_dict['absent'] + 
                monthly_stats_dict['leave'] + 
                monthly_stats_dict['sick']
            )
            
            # حساب عدد الأيام في الشهر حتى اليوم الحالي
            days_in_month = (today - start_of_month).days + 1
            
            # حساب إجمالي سجلات الحضور المفترضة للشهر
            expected_days_month = days_in_month * active_employees_count
            
            monthly_attendance_rate = 0
            if total_days_month > 0:
                monthly_attendance_rate = round((monthly_stats_dict['present'] / total_days_month) * 100)
            
            # 12. تنسيق التواريخ للعرض
            formatted_today = {
                'gregorian': format_date_gregorian(today),
                'hijri': format_date_hijri(today)
            }
            
            formatted_start_of_week = {
                'gregorian': format_date_gregorian(start_of_week),
                'hijri': format_date_hijri(start_of_week)
            }
            
            formatted_end_of_week = {
                'gregorian': format_date_gregorian(end_of_week),
                'hijri': format_date_hijri(end_of_week)
            }
            
            formatted_start_of_month = {
                'gregorian': format_date_gregorian(start_of_month),
                'hijri': format_date_hijri(start_of_month)
            }
            
            formatted_end_of_month = {
                'gregorian': format_date_gregorian(end_of_month),
                'hijri': format_date_hijri(end_of_month)
            }
            
            # 13. إعداد اسم الشهر الحالي
            month_names = [
                'يناير', 'فبراير', 'مارس', 'إبريل', 'مايو', 'يونيو',
                'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
            ]
            current_month_name = month_names[current_month - 1]
            
            # 14. إضافة المتغيرات الأصلية للاحتياط مع المتغيرات المنسقة
            
            # 15. إعداد البيانات للعرض على الصفحة
            return render_template('attendance/dashboard_new.html',
                                today=today,
                                current_month=current_month,
                                current_year=current_year,
                                current_month_name=current_month_name,
                                formatted_today=formatted_today,
                                formatted_start_of_week=formatted_start_of_week,
                                formatted_end_of_week=formatted_end_of_week,
                                formatted_start_of_month=formatted_start_of_month,
                                formatted_end_of_month=formatted_end_of_month,
                                start_of_week=start_of_week,
                                end_of_week=end_of_week,
                                start_of_month=start_of_month,
                                end_of_month=end_of_month,
                                daily_stats=daily_stats_dict,
                                weekly_stats=weekly_stats_dict,
                                monthly_stats=monthly_stats_dict,
                                daily_chart_data=daily_chart_data,
                                weekly_chart_data=weekly_chart_data,
                                monthly_chart_data=monthly_chart_data,
                                daily_trend_chart_data=daily_trend_chart_data,
                                daily_attendance_rate=daily_attendance_rate,
                                weekly_attendance_rate=weekly_attendance_rate,
                                monthly_attendance_rate=monthly_attendance_rate,
                                active_employees_count=active_employees_count,
                                active_projects=active_projects,
                                selected_project=project_name)
                                
            # Si todo funciona bien, sal del bucle
            break
            
        except Exception as e:
            # Si hay un error, incrementa el contador y espera
            retry_count += 1
            logger.error(f"Error al cargar el dashboard (intento {retry_count}): {str(e)}")
            
            if retry_count < max_retries:
                # Espera un tiempo exponencial antes de reintentar
                time_module.sleep(retry_delay)
                retry_delay *= 2  # Duplica el tiempo de espera para el próximo intento
            else:
                # Si se han agotado los reintentos, muestra un mensaje de error
                logger.critical(f"Error al cargar el dashboard después de {max_retries} intentos: {str(e)}")
                return render_template('error.html', 
                                      error_title="خطأ في الاتصال",
                                      error_message="حدث خطأ أثناء الاتصال بقاعدة البيانات. الرجاء المحاولة مرة أخرى.",
                                      error_details=str(e))

@attendance_bp.route('/employee/<int:employee_id>')
def employee_attendance(employee_id):
    """عرض سجلات الحضور التفصيلية للموظف مرتبة حسب الشهر والسنة - Dashboard مميز"""
    # الحصول على الموظف
    employee = Employee.query.get_or_404(employee_id)
    
    # الحصول على التاريخ الحالي
    today = datetime.now().date()
    
    # الحصول على السنة والشهر من URL أو استخدام الحالي
    selected_year = request.args.get('year', today.year, type=int)
    selected_month = request.args.get('month', today.month, type=int)
    
    # تحديد فترة الاستعلام (الشهر المختار)
    year = selected_year
    month = selected_month
    start_of_month = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_of_month = date(year, month, last_day)
    
    # الحصول على سجلات الحضور للشهر المختار
    attendances = Attendance.query.filter(
        Attendance.employee_id == employee_id,
        Attendance.date >= start_of_month,
        Attendance.date <= end_of_month
    ).order_by(Attendance.date).all()
    
    # تنظيم السجلات حسب اليوم للتقويم
    attendance_by_day = {}
    for record in attendances:
        attendance_by_day[record.date.day] = record
    
    # حساب الإحصائيات الشاملة
    present_count = sum(1 for a in attendances if a.status == 'present')
    absent_count = sum(1 for a in attendances if a.status == 'absent')
    leave_count = sum(1 for a in attendances if a.status == 'leave')
    sick_count = sum(1 for a in attendances if a.status == 'sick')
    total_records = len(attendances)
    
    # حساب النسب المئوية
    present_percentage = (present_count / total_records * 100) if total_records > 0 else 0
    absent_percentage = (absent_count / total_records * 100) if total_records > 0 else 0
    leave_percentage = (leave_count / total_records * 100) if total_records > 0 else 0
    sick_percentage = (sick_count / total_records * 100) if total_records > 0 else 0
    
    # حساب معدل الحضور
    attendance_rate = round(present_percentage, 1) if total_records > 0 else 0
    
    # جمع كل الفترات المتاحة (السنوات والأشهر التي لديها سجلات)
    all_records = Attendance.query.filter(Attendance.employee_id == employee_id).all()
    attendance_periods = {}
    for record in all_records:
        if record.date.year not in attendance_periods:
            attendance_periods[record.date.year] = set()
        attendance_periods[record.date.year].add(record.date.month)
    
    # معلومات التقويم
    first_day_weekday = calendar.monthrange(year, month)[0]
    days_in_month = last_day
    
    # تنسيق التواريخ للعرض
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    return render_template('attendance/employee_attendance.html',
                          employee=employee,
                          attendances=attendances,
                          attendance_by_day=attendance_by_day,
                          year=year,
                          month=month,
                          selected_year=selected_year,
                          selected_month=selected_month,
                          first_day_weekday=first_day_weekday,
                          days_in_month=days_in_month,
                          attendance_periods=attendance_periods,
                          present_count=present_count,
                          absent_count=absent_count,
                          leave_count=leave_count,
                          sick_count=sick_count,
                          total_records=total_records,
                          present_percentage=present_percentage,
                          absent_percentage=absent_percentage,
                          leave_percentage=leave_percentage,
                          sick_percentage=sick_percentage,
                          attendance_rate=attendance_rate,
                          today=today,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today)

@attendance_bp.route('/department-stats')
def department_stats():
    """API لجلب إحصائيات الحضور حسب الأقسام"""
    period = request.args.get('period', 'monthly')  # weekly أو monthly
    project_name = request.args.get('project', None)
    
    today = datetime.now().date()
    
    # تحديد الفترة الزمنية - استخدام البيانات الشهرية الحقيقية
    start_date = today.replace(day=1)  # بداية الشهر الحالي
    end_date = today  # حتى اليوم الحالي
    
    # جلب الأقسام المسموح بالوصول إليها حسب صلاحيات المستخدم
    from flask_login import current_user
    
    if current_user.is_authenticated:
        # إذا كان المستخدم مسجل دخوله، عرض الأقسام المسموحة فقط
        departments = current_user.get_accessible_departments()
    else:
        # إذا لم يكن مسجل دخوله، عرض جميع الأقسام (للعرض العام)
        departments = Department.query.all()
    
    department_stats = []
    
    for dept in departments:
        # جلب الموظفين النشطين في القسم - استخدام علاقة many-to-many
        employees = [emp for emp in dept.employees if emp.status == 'active']
        
        # فلترة حسب المشروع إذا تم تحديده
        if project_name:
            employees = [emp for emp in employees if emp.project == project_name]
        
        total_employees = len(employees)
        
        # عرض جميع الأقسام حتى لو كانت فارغة لضمان الشمولية
        # if total_employees == 0:
        #     continue
        
        # حساب الإحصائيات
        employee_ids = [emp.id for emp in employees]
        
        # جلب سجلات الحضور للفترة المحددة
        attendance_records = []
        if employee_ids:
            attendance_records = Attendance.query.filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
        
        # حساب الإحصائيات
        present_count = sum(1 for record in attendance_records if record.status == 'present')
        absent_count = sum(1 for record in attendance_records if record.status == 'absent')
        leave_count = sum(1 for record in attendance_records if record.status == 'leave')
        sick_count = sum(1 for record in attendance_records if record.status == 'sick')
        total_records = len(attendance_records)
        
        # حساب الأيام والسجلات المتوقعة
        working_days = (end_date - start_date).days + 1
        expected_total_records = total_employees * working_days
        
        # للفترة الشهرية، نحسب أيام العمل الفعلية (عدا الجمع والسبوت)
        if period == 'monthly':
            working_days_actual = 0
            current = start_date
            while current <= end_date:
                # حساب أيام العمل (الأحد-الخميس في النظام السعودي)
                if current.weekday() < 5:  # 0-4 (الاثنين-الجمعة) نحسبها أيام عمل
                    working_days_actual += 1
                current += timedelta(days=1)
            working_days = working_days_actual
        
        # حساب معدل الحضور بناء على السجلات الفعلية الموجودة
        if total_records > 0:
            attendance_rate = (present_count / total_records) * 100
        else:
            attendance_rate = 0
        
        department_stats.append({
            'id': dept.id,
            'name': dept.name,
            'total_employees': total_employees,
            'present': present_count,
            'absent': absent_count,
            'leave': leave_count,
            'sick': sick_count,
            'attendance_rate': round(attendance_rate, 1),
            'total_records': total_records,
            'working_days': working_days,
            'expected_records': expected_total_records
        })
    
    # ترتيب الأقسام حسب معدل الحضور (تنازلي)
    department_stats.sort(key=lambda x: x['attendance_rate'], reverse=True)
    
    return jsonify({
        'departments': department_stats,
        'period': period,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'project': project_name
    })

@attendance_bp.route('/export-excel-dashboard')
def export_excel_dashboard():
    """تصدير لوحة المعلومات إلى Excel مع تصميم داش بورد خيالي ومبهر"""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
        
        # الحصول على المعاملات
        selected_department = request.args.get('department', None)
        selected_project = request.args.get('project', None)
        
        today = datetime.now().date()
        start_date = today.replace(day=1)
        end_date = today
        
        # جلب جميع الموظفين
        all_employees = Employee.query.filter_by(status='active').all()
        total_employees = len(all_employees)
        
        # جلب البيانات حسب الفلتر
        departments = Department.query.all()
        department_data = []
        
        total_present = 0
        total_absent = 0
        total_leave = 0
        total_sick = 0
        total_records = 0
        
        for dept in departments:
            if selected_department and dept.name != selected_department:
                continue
                
            employees = [emp for emp in all_employees if any(d.id == dept.id for d in emp.departments)]
            dept_employee_count = len(employees)
            
            if dept_employee_count == 0:
                continue
                
            employee_ids = [emp.id for emp in employees]
            attendance_records = Attendance.query.filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
            
            present_count = sum(1 for r in attendance_records if r.status == 'present')
            absent_count = sum(1 for r in attendance_records if r.status == 'absent')
            leave_count = sum(1 for r in attendance_records if r.status == 'leave')
            sick_count = sum(1 for r in attendance_records if r.status == 'sick')
            dept_total_records = len(attendance_records)
            
            attendance_rate = (present_count / dept_total_records * 100) if dept_total_records > 0 else 0
            
            total_present += present_count
            total_absent += absent_count
            total_leave += leave_count
            total_sick += sick_count
            total_records += dept_total_records
            
            department_data.append({
                'name': dept.name,
                'employees': dept_employee_count,
                'present': present_count,
                'absent': absent_count,
                'leave': leave_count,
                'sick': sick_count,
                'total': dept_total_records,
                'rate': round(attendance_rate, 1)
            })
        
        if not department_data:
            wb = Workbook()
            ws = wb.active
            ws.title = "لا توجد بيانات"
            ws['A1'] = "لا توجد بيانات للعرض"
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=f'تقرير_الحضور_{today.strftime("%Y%m%d")}.xlsx')
        
        # إنشاء ملف Excel خيالي
        wb = Workbook()
        ws = wb.active
        ws.title = "📊 لوحة المعلومات"
        
        # === العنوان الرئيسي الفاخر ===
        ws.merge_cells('A1:M3')
        title_cell = ws['A1']
        title_cell.value = f"📊 لوحة معلومات الحضور والغياب\n{start_date.strftime('%Y/%m/%d')} - {end_date.strftime('%Y/%m/%d')}"
        title_cell.font = Font(size=24, bold=True, color="FFFFFF", name="Arial")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        
        # === بطاقات KPI الملونة ===
        kpi_row = 5
        ws.row_dimensions[kpi_row].height = 35
        ws.row_dimensions[kpi_row + 1].height = 30
        
        # تعريف الحدود
        thick_border = Border(
            left=Side(style='thick', color='FFFFFF'),
            right=Side(style='thick', color='FFFFFF'),
            top=Side(style='thick', color='FFFFFF'),
            bottom=Side(style='thick', color='FFFFFF')
        )
        
        # KPI 1: إجمالي الموظفين
        ws.merge_cells(f'A{kpi_row}:C{kpi_row+1}')
        kpi1 = ws[f'A{kpi_row}']
        kpi1.value = f"👥 إجمالي الموظفين\n{total_employees}"
        kpi1.font = Font(size=16, bold=True, color="FFFFFF")
        kpi1.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        kpi1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi1.border = thick_border
        
        # KPI 2: الحضور
        ws.merge_cells(f'D{kpi_row}:F{kpi_row+1}')
        kpi2 = ws[f'D{kpi_row}']
        kpi2.value = f"✅ الحضور\n{total_present}"
        kpi2.font = Font(size=16, bold=True, color="FFFFFF")
        kpi2.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        kpi2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi2.border = thick_border
        
        # KPI 3: الغياب
        ws.merge_cells(f'G{kpi_row}:I{kpi_row+1}')
        kpi3 = ws[f'G{kpi_row}']
        kpi3.value = f"❌ الغياب\n{total_absent}"
        kpi3.font = Font(size=16, bold=True, color="FFFFFF")
        kpi3.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        kpi3.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi3.border = thick_border
        
        # KPI 4: الإجازات
        ws.merge_cells(f'J{kpi_row}:L{kpi_row+1}')
        kpi4 = ws[f'J{kpi_row}']
        kpi4.value = f"🏖️ الإجازات\n{total_leave}"
        kpi4.font = Font(size=16, bold=True, color="FFFFFF")
        kpi4.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        kpi4.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi4.border = thick_border
        
        # KPI 5: المرضي
        ws.merge_cells(f'M{kpi_row}:O{kpi_row+1}')
        kpi5 = ws[f'M{kpi_row}']
        kpi5.value = f"🏥 المرضي\n{total_sick}"
        kpi5.font = Font(size=16, bold=True, color="FFFFFF")
        kpi5.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        kpi5.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi5.border = thick_border
        
        # === جدول البيانات التفصيلي ===
        table_start_row = kpi_row + 3
        
        # عنوان الجدول
        ws.merge_cells(f'A{table_start_row}:H{table_start_row}')
        table_title = ws[f'A{table_start_row}']
        table_title.value = "📋 تفاصيل الأقسام"
        table_title.font = Font(size=14, bold=True, color="FFFFFF")
        table_title.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        table_title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[table_start_row].height = 25
        
        # رؤوس الأعمدة
        headers_row = table_start_row + 1
        headers = ['القسم', 'عدد الموظفين', 'حاضر', 'غائب', 'إجازة', 'مرضي', 'إجمالي السجلات', 'معدل الحضور %']
        header_colors = ['34495E', '5B9BD5', '70AD47', 'E74C3C', 'F39C12', '3498DB', '95A5A6', '16A085']
        
        for col_idx, (header, color) in enumerate(zip(headers, header_colors), 1):
            cell = ws.cell(row=headers_row, column=col_idx)
            cell.value = header
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        ws.row_dimensions[headers_row].height = 30
        
        # البيانات
        data_start_row = headers_row + 1
        for row_idx, dept in enumerate(department_data, data_start_row):
            values = [dept['name'], dept['employees'], dept['present'], dept['absent'], 
                     dept['leave'], dept['sick'], dept['total'], dept['rate']]
            
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # تلوين بديل للصفوف
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                
                # تلوين خاص للأعمدة
                if col_idx == 3:  # حاضر
                    cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                    cell.font = Font(bold=True, color="27AE60")
                elif col_idx == 4:  # غائب
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    cell.font = Font(bold=True, color="C0392B")
                elif col_idx == 5:  # إجازة
                    cell.fill = PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid")
                    cell.font = Font(bold=True, color="D68910")
                elif col_idx == 6:  # مرضي
                    cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                    cell.font = Font(bold=True, color="2874A6")
                elif col_idx == 8:  # معدل الحضور
                    cell.font = Font(bold=True, size=11)
                    if value >= 90:
                        cell.fill = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")
                    elif value >= 75:
                        cell.fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
        
        # === الرسوم البيانية ===
        chart_row = data_start_row + len(department_data) + 2
        
        # 1. مخطط عمودي - مقارنة الحضور بين الأقسام
        bar_chart = BarChart()
        bar_chart.title = "مقارنة حالات الحضور بين الأقسام"
        bar_chart.style = 13
        bar_chart.y_axis.title = 'العدد'
        bar_chart.x_axis.title = 'الأقسام'
        bar_chart.height = 12
        bar_chart.width = 20
        
        data_ref = Reference(ws, min_col=3, min_row=headers_row, max_row=data_start_row + len(department_data) - 1, max_col=6)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_start_row + len(department_data) - 1)
        
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        ws.add_chart(bar_chart, f"A{chart_row}")
        
        # 2. مخطط دائري - نسب الحضور الإجمالية
        pie_chart = PieChart()
        pie_chart.title = "نسب حالات الحضور الإجمالية"
        pie_chart.height = 12
        pie_chart.width = 15
        
        # إنشاء بيانات الإجمالي
        summary_row = data_start_row + len(department_data) + 15
        ws[f'K{summary_row}'] = 'حاضر'
        ws[f'L{summary_row}'] = total_present
        ws[f'K{summary_row+1}'] = 'غائب'
        ws[f'L{summary_row+1}'] = total_absent
        ws[f'K{summary_row+2}'] = 'إجازة'
        ws[f'L{summary_row+2}'] = total_leave
        ws[f'K{summary_row+3}'] = 'مرضي'
        ws[f'L{summary_row+3}'] = total_sick
        
        pie_data = Reference(ws, min_col=12, min_row=summary_row, max_row=summary_row+3)
        pie_labels = Reference(ws, min_col=11, min_row=summary_row, max_row=summary_row+3)
        pie_chart.add_data(pie_data, titles_from_data=False)
        pie_chart.set_categories(pie_labels)
        ws.add_chart(pie_chart, f"K{chart_row}")
        
        # تعيين عرض الأعمدة
        column_widths = [20, 15, 12, 12, 12, 12, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'📊_لوحة_معلومات_الحضور_{today.strftime("%Y%m%d")}.xlsx'
        if selected_department:
            filename = f'📊_{selected_department}_{today.strftime("%Y%m%d")}.xlsx'
            
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"خطأ في تصدير Excel: {str(e)}")
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.dashboard'))

@attendance_bp.route('/department-details')
def department_details():
    """صفحة تفاصيل الحضور لقسم معين"""
    department_name = request.args.get('department')
    period = request.args.get('period', 'weekly')
    project_name = request.args.get('project', None)
    
    if not department_name:
        flash('يجب تحديد القسم', 'error')
        return redirect(url_for('attendance.dashboard'))
    
    # جلب القسم
    department = Department.query.filter_by(name=department_name).first()
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('attendance.dashboard'))
    
    today = datetime.now().date()
    
    # تحديد الفترة الزمنية - دائماً عرض الشهر الكامل للتفاصيل
    start_date = today.replace(day=1)  # بداية الشهر الحالي
    end_date = today  # حتى اليوم الحالي
    
    # إنشاء قائمة بجميع أيام الشهر حتى اليوم
    date_range = []
    current = start_date
    while current <= end_date:
        date_range.append(current)
        current += timedelta(days=1)
    
    # جلب الموظفين النشطين في القسم
    employees_query = Employee.query.filter_by(
        department_id=department.id,
        status='active'
    )
    
    if project_name and project_name != 'None' and project_name.strip():
        employees_query = employees_query.filter_by(project=project_name)
    
    employees = employees_query.all()
    
    # تسجيل عدد الموظفين للتشخيص
    print(f"تفاصيل القسم - عدد الموظفين المجلوبين: {len(employees)} للقسم {department.name}")
    for emp in employees:
        print(f"  - {emp.name} (ID: {emp.id})")
    
    # جلب سجلات الحضور للموظفين في الفترة المحددة
    employee_attendance = {}
    for employee in employees:
        attendance_records = Attendance.query.filter(
            Attendance.employee_id == employee.id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).order_by(Attendance.date).all()
        
        employee_attendance[employee.id] = {
            'employee': employee,
            'records': attendance_records,
            'stats': {
                'present': sum(1 for r in attendance_records if r.status == 'present'),
                'absent': sum(1 for r in attendance_records if r.status == 'absent'),
                'leave': sum(1 for r in attendance_records if r.status == 'leave'),
                'sick': sum(1 for r in attendance_records if r.status == 'sick')
            }
        }
    
    # حساب الإحصائيات الإجمالية للقسم
    total_stats = {
        'total_employees': len(employees),
        'present': 0,
        'absent': 0,
        'leave': 0,
        'sick': 0,
        'total_records': 0,
        'working_days': len(date_range),
        'attendance_rate': 0
    }
    
    for emp_data in employee_attendance.values():
        total_stats['present'] += emp_data['stats']['present']
        total_stats['absent'] += emp_data['stats']['absent']
        total_stats['leave'] += emp_data['stats']['leave']
        total_stats['sick'] += emp_data['stats']['sick']
        total_stats['total_records'] += len(emp_data['records'])
    
    # حساب معدل الحضور
    if total_stats['total_records'] > 0:
        total_stats['attendance_rate'] = round((total_stats['present'] / total_stats['total_records']) * 100, 1)
    
    # حساب الإحصائيات اليومية للقسم
    daily_stats = {}
    for date in date_range:
        daily_count = {
            'present': 0,
            'absent': 0,
            'leave': 0,
            'sick': 0,
            'total': 0
        }
        
        for emp_data in employee_attendance.values():
            for record in emp_data['records']:
                if record.date == date:
                    daily_count[record.status] += 1
                    daily_count['total'] += 1
                    break
        
        daily_stats[date] = daily_count
    
    # إحصائيات أسبوعية
    weekly_stats = []
    week_start = start_date
    while week_start <= end_date:
        week_end = min(week_start + timedelta(days=6), end_date)
        
        week_data = {
            'start_date': week_start,
            'end_date': week_end,
            'present': 0,
            'absent': 0,
            'leave': 0,
            'sick': 0
        }
        
        current = week_start
        while current <= week_end:
            if current in daily_stats:
                week_data['present'] += daily_stats[current]['present']
                week_data['absent'] += daily_stats[current]['absent']
                week_data['leave'] += daily_stats[current]['leave']
                week_data['sick'] += daily_stats[current]['sick']
            current += timedelta(days=1)
        
        weekly_stats.append(week_data)
        week_start += timedelta(days=7)
    
    return render_template('attendance/department_details_enhanced.html',
                          department=department,
                          employee_attendance=employee_attendance,
                          date_range=date_range,
                          daily_stats=daily_stats,
                          weekly_stats=weekly_stats,
                          total_stats=total_stats,
                          period='monthly',  # دائماً عرض شهري
                          start_date=start_date,
                          end_date=end_date,
                          project_name=project_name)

@attendance_bp.route('/export-excel-department')
def export_excel_department():
    """تصدير تفاصيل القسم إلى Excel مع لوحة معلومات تفصيلية مميزة"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        department_name = request.args.get('department')
        selected_project = request.args.get('project', None)
        
        if not department_name:
            flash('يجب تحديد القسم', 'error')
            return redirect(url_for('attendance.dashboard'))
        
        # جلب القسم
        department = Department.query.filter_by(name=department_name).first()
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('attendance.dashboard'))
        
        today = datetime.now().date()
        start_date = today.replace(day=1)
        end_date = today
        
        # إنشاء قائمة بجميع أيام الشهر حتى اليوم
        date_range = []
        current = start_date
        while current <= end_date:
            date_range.append(current)
            current += timedelta(days=1)
        
        # جلب الموظفين والبيانات - استخدام علاقة many-to-many
        all_employees = [emp for emp in department.employees if emp.status == 'active']
        
        if selected_project and selected_project != 'None' and selected_project.strip():
            employees = [emp for emp in all_employees if emp.project == selected_project]
        else:
            employees = all_employees
        
        # إنشاء ملف Excel
        wb = Workbook()
        
        # صفحة لوحة المعلومات الرئيسية
        ws_dashboard = wb.active
        ws_dashboard.title = "لوحة المعلومات"
        
        # العنوان الرئيسي
        title = f"لوحة معلومات قسم {department.name}"
        if selected_project and selected_project != 'None':
            title += f" - مشروع {selected_project}"
        
        ws_dashboard['A1'] = title
        ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws_dashboard['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.merge_cells('A1:H3')
        
        # معلومات الفترة
        period_info = f"الفترة: من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
        ws_dashboard['A4'] = period_info
        ws_dashboard['A4'].font = Font(size=12, bold=True)
        ws_dashboard['A4'].alignment = Alignment(horizontal='center')
        ws_dashboard.merge_cells('A4:H4')
        
        # جمع بيانات الموظفين والحضور
        employee_data = []
        total_stats = {
            'total_employees': len(employees),
            'present': 0,
            'absent': 0,
            'leave': 0,
            'sick': 0,
            'total_records': 0
        }
        
        for employee in employees:
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
            
            stats = {
                'present': sum(1 for r in attendance_records if r.status == 'present'),
                'absent': sum(1 for r in attendance_records if r.status == 'absent'),
                'leave': sum(1 for r in attendance_records if r.status == 'leave'),
                'sick': sum(1 for r in attendance_records if r.status == 'sick')
            }
            
            total_records = sum(stats.values())
            attendance_rate = (stats['present'] / total_records * 100) if total_records > 0 else 0
            
            employee_data.append({
                'الاسم': employee.name,
                'الهوية': employee.employee_id or 'غير محدد',
                'حاضر': stats['present'],
                'غائب': stats['absent'],
                'إجازة': stats['leave'],
                'مرضي': stats['sick'],
                'الإجمالي': total_records,
                'المعدل %': round(attendance_rate, 1)
            })
            
            # إضافة للإحصائيات الإجمالية
            total_stats['present'] += stats['present']
            total_stats['absent'] += stats['absent']
            total_stats['leave'] += stats['leave']
            total_stats['sick'] += stats['sick']
            total_stats['total_records'] += total_records
        
        # حساب معدل الحضور الإجمالي
        overall_rate = (total_stats['present'] / total_stats['total_records'] * 100) if total_stats['total_records'] > 0 else 0
        
        # الإحصائيات الإجمالية في لوحة المعلومات
        stats_row = 6
        
        # عناوين الإحصائيات
        stats_headers = ['إجمالي الموظفين', 'إجمالي الحضور', 'إجمالي الغياب', 'إجمالي الإجازات', 'إجمالي المرضي', 'معدل الحضور %', 'أيام العمل']
        stats_values = [total_stats['total_employees'], total_stats['present'], total_stats['absent'], 
                       total_stats['leave'], total_stats['sick'], round(overall_rate, 1), len(date_range)]
        
        for col, (header, value) in enumerate(zip(stats_headers, stats_values), 1):
            # العنوان
            header_cell = ws_dashboard.cell(row=stats_row, column=col, value=header)
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # القيمة
            value_cell = ws_dashboard.cell(row=stats_row + 1, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # تلوين حسب النوع
            if 'حضور' in header:
                value_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif 'غياب' in header:
                value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif 'إجازة' in header:
                value_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif 'مرضي' in header:
                value_cell.fill = PatternFill(start_color="9CC3F7", end_color="9CC3F7", fill_type="solid")
            else:
                value_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # جدول تفاصيل الموظفين
        table_start_row = stats_row + 4
        
        # عنوان الجدول
        ws_dashboard[f'A{table_start_row}'] = "تفاصيل حضور الموظفين"
        ws_dashboard[f'A{table_start_row}'].font = Font(size=14, bold=True)
        ws_dashboard.merge_cells(f'A{table_start_row}:H{table_start_row}')
        
        # رؤوس الجدول
        headers = list(employee_data[0].keys()) if employee_data else []
        header_row = table_start_row + 1
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_dashboard.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # بيانات الموظفين
        for row_num, emp_data in enumerate(employee_data, header_row + 1):
            for col_num, value in enumerate(emp_data.values(), 1):
                cell = ws_dashboard.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # تلوين صفوف متناوبة
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # تعديل عرض الأعمدة
        ws_dashboard.column_dimensions['A'].width = 25
        ws_dashboard.column_dimensions['B'].width = 15
        ws_dashboard.column_dimensions['C'].width = 12
        ws_dashboard.column_dimensions['D'].width = 12
        ws_dashboard.column_dimensions['E'].width = 12
        ws_dashboard.column_dimensions['F'].width = 12
        ws_dashboard.column_dimensions['G'].width = 12
        ws_dashboard.column_dimensions['H'].width = 12
        
        # صفحة التفاصيل اليومية
        ws_daily = wb.create_sheet("التفاصيل اليومية")
        
        # عنوان صفحة التفاصيل اليومية
        ws_daily['A1'] = f"التفاصيل اليومية - قسم {department.name}"
        ws_daily['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_daily['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_daily['A1'].alignment = Alignment(horizontal='center')
        ws_daily.merge_cells('A1:AF3')
        
        # رؤوس الأعمدة للتفاصيل اليومية
        ws_daily['A5'] = 'اسم الموظف'
        ws_daily['B5'] = 'رقم الهوية'
        
        # تواريخ الشهر
        for col_num, date in enumerate(date_range, 3):
            cell = ws_daily.cell(row=5, column=col_num, value=date.strftime('%d-%m'))
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', textRotation=90)
        
        # بيانات الحضور اليومي
        for row_num, employee in enumerate(employees, 6):
            ws_daily.cell(row=row_num, column=1, value=employee.name)
            ws_daily.cell(row=row_num, column=2, value=employee.employee_id or 'غير محدد')
            
            # جلب سجلات الحضور للموظف
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
            
            attendance_dict = {record.date: record.status for record in attendance_records}
            
            for col_num, date in enumerate(date_range, 3):
                status = attendance_dict.get(date, '-')
                cell = ws_daily.cell(row=row_num, column=col_num, value=status)
                
                # تلوين الخلايا حسب الحالة
                if status == 'present':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.value = "✓"
                elif status == 'absent':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.value = "✗"
                elif status == 'leave':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.value = "إ"
                elif status == 'sick':
                    cell.fill = PatternFill(start_color="9CC3F7", end_color="9CC3F7", fill_type="solid")
                    cell.value = "م"
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنسيق أعمدة التفاصيل اليومية
        ws_daily.column_dimensions['A'].width = 30
        ws_daily.column_dimensions['B'].width = 15
        
        # تنسيق أعمدة التواريخ (تحديد عرض بسيط لتجنب مشاكل الأعمدة الكثيرة)
        try:
            for col_num in range(3, min(3 + len(date_range), 26)):  # حد أقصى 26 عمود
                if col_num <= 26:  # A-Z فقط
                    col_letter = chr(64 + col_num)
                    ws_daily.column_dimensions[col_letter].width = 4
        except Exception:
            pass  # تجاهل أخطاء الأعمدة
        
        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'تفاصيل_قسم_{department.name}_{today.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"خطأ في تصدير تفاصيل القسم: {str(e)}")
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.dashboard'))

@attendance_bp.route('/department/view', methods=['GET'])
def department_attendance_view():
    """عرض حضور الأقسام خلال فترة زمنية محددة"""
    from flask_login import current_user
    
    # الحصول على معاملات الفلترة
    department_id = request.args.get('department_id', '')
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    # تحديد التواريخ الافتراضية (آخر 30 يوم)
    if not start_date_str:
        start_date = datetime.now().date() - timedelta(days=30)
    else:
        start_date = parse_date(start_date_str)
    
    if not end_date_str:
        end_date = datetime.now().date()
    else:
        end_date = parse_date(end_date_str)
    
    # الحصول على الأقسام حسب صلاحيات المستخدم
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
    else:
        departments = Department.query.all()
    
    # بناء الاستعلام
    query = Attendance.query.join(Employee).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    )
    
    # تطبيق فلتر القسم إذا تم تحديده
    if department_id:
        query = query.join(employee_departments).filter(
            employee_departments.c.department_id == int(department_id)
        )
    
    # الحصول على السجلات مرتبة حسب التاريخ والموظف
    attendances = query.order_by(Attendance.date.desc(), Employee.name).all()
    
    # حساب الإحصائيات
    total_count = len(attendances)
    present_count = sum(1 for a in attendances if a.status == 'present')
    absent_count = sum(1 for a in attendances if a.status == 'absent')
    leave_count = sum(1 for a in attendances if a.status == 'leave')
    sick_count = sum(1 for a in attendances if a.status == 'sick')
    
    return render_template('attendance/department_period.html',
                          departments=departments,
                          department_id=department_id,
                          start_date=start_date,
                          end_date=end_date,
                          attendances=attendances,
                          total_count=total_count,
                          present_count=present_count,
                          absent_count=absent_count,
                          leave_count=leave_count,
                          sick_count=sick_count)

@attendance_bp.route('/department/export-period', methods=['GET'])
def export_department_period():
    """تصدير حضور قسم خلال فترة زمنية إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        department_id = request.args.get('department_id')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if not all([department_id, start_date_str, end_date_str]):
            flash('يجب تحديد القسم والفترة الزمنية', 'error')
            return redirect(url_for('attendance.department_attendance_view'))
        
        # تحليل التواريخ
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        
        # الحصول على القسم
        department = Department.query.get_or_404(department_id)
        
        # الحصول على بيانات الحضور
        attendances = Attendance.query.join(Employee).join(employee_departments).filter(
            employee_departments.c.department_id == int(department_id),
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).order_by(Employee.name, Attendance.date).all()
        
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "حضور القسم"
        
        # تنسيقات الألوان
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=16)
        
        present_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        absent_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        leave_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        sick_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        
        # العنوان
        ws.merge_cells('A1:H1')
        ws['A1'] = f'تقرير حضور قسم {department.name}'
        ws['A1'].fill = title_fill
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # معلومات الفترة
        ws.merge_cells('A2:H2')
        ws['A2'] = f'من {start_date.strftime("%Y-%m-%d")} إلى {end_date.strftime("%Y-%m-%d")}'
        ws['A2'].font = Font(bold=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # رؤوس الأعمدة
        headers = ['الموظف', 'الرقم الوظيفي', 'التاريخ', 'الحالة', 'وقت الدخول', 'وقت الخروج', 'ساعات العمل', 'ملاحظات']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # البيانات
        row_num = 5
        for attendance in attendances:
            ws.cell(row=row_num, column=1, value=attendance.employee.name)
            ws.cell(row=row_num, column=2, value=attendance.employee.employee_id or '-')
            ws.cell(row=row_num, column=3, value=attendance.date.strftime('%Y-%m-%d'))
            
            # حالة الحضور مع التلوين
            status_cell = ws.cell(row=row_num, column=4)
            if attendance.status == 'present':
                status_cell.value = 'حاضر'
                status_cell.fill = present_fill
            elif attendance.status == 'absent':
                status_cell.value = 'غائب'
                status_cell.fill = absent_fill
            elif attendance.status == 'leave':
                status_cell.value = 'إجازة'
                status_cell.fill = leave_fill
            elif attendance.status == 'sick':
                status_cell.value = 'مرضي'
                status_cell.fill = sick_fill
            
            status_cell.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_num, column=5, value=attendance.check_in.strftime('%H:%M') if attendance.check_in else '-')
            ws.cell(row=row_num, column=6, value=attendance.check_out.strftime('%H:%M') if attendance.check_out else '-')
            
            # حساب ساعات العمل
            if attendance.check_in and attendance.check_out:
                hours = attendance.hours_worked
                ws.cell(row=row_num, column=7, value=f"{hours:.1f} ساعة")
            else:
                ws.cell(row=row_num, column=7, value='-')
            
            ws.cell(row=row_num, column=8, value=attendance.notes or '-')
            
            row_num += 1
        
        # تنسيق عرض الأعمدة
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        
        # إضافة ورقة الإحصائيات
        ws_stats = wb.create_sheet("الإحصائيات")
        
        # عنوان الإحصائيات
        ws_stats.merge_cells('A1:B1')
        ws_stats['A1'] = 'إحصائيات الحضور'
        ws_stats['A1'].fill = title_fill
        ws_stats['A1'].font = title_font
        ws_stats['A1'].alignment = Alignment(horizontal='center')
        
        # حساب الإحصائيات
        total_count = len(attendances)
        present_count = sum(1 for a in attendances if a.status == 'present')
        absent_count = sum(1 for a in attendances if a.status == 'absent')
        leave_count = sum(1 for a in attendances if a.status == 'leave')
        sick_count = sum(1 for a in attendances if a.status == 'sick')
        
        stats_data = [
            ('إجمالي السجلات', total_count),
            ('حضور', present_count),
            ('غياب', absent_count),
            ('إجازات', leave_count),
            ('مرضي', sick_count),
            ('نسبة الحضور', f"{(present_count/total_count*100):.1f}%" if total_count > 0 else "0%")
        ]
        
        row_num = 3
        for label, value in stats_data:
            ws_stats.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            ws_stats.cell(row=row_num, column=2, value=value)
            row_num += 1
        
        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 15
        
        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'حضور_{department.name}_{start_date.strftime("%Y%m%d")}_إلى_{end_date.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"خطأ في تصدير حضور الفترة: {str(e)}")
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.department_attendance_view'))

@attendance_bp.route('/department/bulk', methods=['GET', 'POST'])
def department_bulk_attendance():
    """تسجيل حضور قسم كامل لفترة زمنية محددة"""
    from flask_login import current_user
    
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            start_date_str = request.form['start_date']
            end_date_str = request.form['end_date']
            status = request.form['status']
            skip_weekends = 'skip_weekends' in request.form
            overwrite_existing = 'overwrite_existing' in request.form
            
            # التحقق من صلاحيات المستخدم
            if current_user.is_authenticated and not current_user.can_access_department(int(department_id)):
                flash('ليس لديك صلاحية لتسجيل حضور هذا القسم', 'error')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # تحليل التواريخ
            start_date = parse_date(start_date_str)
            end_date = parse_date(end_date_str)
            
            # التحقق من صحة الفترة
            if start_date > end_date:
                flash('تاريخ البداية يجب أن يكون قبل تاريخ النهاية', 'error')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # الحد الأقصى للفترة (90 يوم)
            if (end_date - start_date).days > 90:
                flash('الفترة الزمنية لا يمكن أن تتجاوز 90 يوم', 'error')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # الحصول على القسم والموظفين
            department = Department.query.get_or_404(department_id)
            employees = [emp for emp in department.employees if emp.status == 'active']
            
            if not employees:
                flash('لا يوجد موظفين نشطين في هذا القسم', 'warning')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # إنشاء قائمة التواريخ
            date_list = []
            current_date = start_date
            while current_date <= end_date:
                # تخطي عطلة نهاية الأسبوع إذا تم تحديد الخيار
                if skip_weekends:
                    # 4 = الجمعة، 5 = السبت في Python (0=الإثنين)
                    if current_date.weekday() in [4, 5]:
                        current_date += timedelta(days=1)
                        continue
                
                date_list.append(current_date)
                current_date += timedelta(days=1)
            
            # حساب الإحصائيات
            created_count = 0
            updated_count = 0
            skipped_count = 0
            
            # تسجيل الحضور لكل موظف في كل يوم
            for employee in employees:
                for attendance_date in date_list:
                    # البحث عن سجل موجود
                    existing = Attendance.query.filter_by(
                        employee_id=employee.id,
                        date=attendance_date
                    ).first()
                    
                    if existing:
                        if overwrite_existing:
                            # تحديث السجل الموجود
                            existing.status = status
                            if status != 'present':
                                existing.check_in = None
                                existing.check_out = None
                            updated_count += 1
                        else:
                            # تخطي السجل الموجود
                            skipped_count += 1
                    else:
                        # إنشاء سجل جديد
                        new_attendance = Attendance(
                            employee_id=employee.id,
                            date=attendance_date,
                            status=status
                        )
                        db.session.add(new_attendance)
                        created_count += 1
            
            # حفظ التغييرات
            db.session.commit()
            
            # تسجيل العملية في سجل النشاط
            log_attendance_activity(
                action='bulk_create_period',
                attendance_data={
                    'department_id': department_id,
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat(),
                    'status': status,
                    'created': created_count,
                    'updated': updated_count,
                    'skipped': skipped_count
                },
                employee_name=f"جميع موظفي قسم {department.name}"
            )
            
            # رسالة النجاح
            message_parts = []
            if created_count > 0:
                message_parts.append(f'تم إنشاء {created_count} سجل جديد')
            if updated_count > 0:
                message_parts.append(f'تم تحديث {updated_count} سجل')
            if skipped_count > 0:
                message_parts.append(f'تم تخطي {skipped_count} سجل موجود')
            
            flash(' | '.join(message_parts), 'success')
            return redirect(url_for('attendance.index'))
            
        except Exception as e:
            db.session.rollback()
            print(f"خطأ في تسجيل الحضور الجماعي: {str(e)}")
            flash(f'حدث خطأ: {str(e)}', 'danger')
            return redirect(url_for('attendance.department_bulk_attendance'))
    
    # GET request
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
    else:
        departments = Department.query.all()
    
    return render_template('attendance/department_bulk.html', departments=departments)

@attendance_bp.route('/edit/<int:id>', methods=['GET'])
def edit_attendance_page(id):
    """عرض صفحة تعديل سجل الحضور"""
    from flask_login import current_user
    
    # الحصول على سجل الحضور
    attendance = Attendance.query.get_or_404(id)
    
    # التحقق من صلاحيات المستخدم
    if current_user.is_authenticated:
        employee_departments = [dept.id for dept in attendance.employee.departments]
        if employee_departments and not any(current_user.can_access_department(dept_id) for dept_id in employee_departments):
            flash('ليس لديك صلاحية لتعديل هذا السجل', 'error')
            return redirect(url_for('attendance.department_attendance_view'))
    
    return render_template('attendance/edit_attendance.html', attendance=attendance)

@attendance_bp.route('/edit/<int:id>', methods=['POST'])
def update_attendance_page(id):
    """تحديث سجل حضور موجود من صفحة التعديل"""
    from flask_login import current_user
    from datetime import time as dt_time
    
    try:
        # الحصول على سجل الحضور
        attendance = Attendance.query.get_or_404(id)
        
        # التحقق من صلاحيات المستخدم
        if current_user.is_authenticated:
            employee_departments = [dept.id for dept in attendance.employee.departments]
            if employee_departments and not any(current_user.can_access_department(dept_id) for dept_id in employee_departments):
                flash('ليس لديك صلاحية لتعديل هذا السجل', 'error')
                return redirect(url_for('attendance.department_attendance_view'))
        
        # الحصول على البيانات
        status = request.form.get('status')
        check_in_str = request.form.get('check_in', '')
        check_out_str = request.form.get('check_out', '')
        
        # تحديث الحالة
        old_status = attendance.status
        attendance.status = status
        
        # معالجة أوقات الدخول والخروج
        if status == 'present':
            if check_in_str:
                try:
                    hours, minutes = map(int, check_in_str.split(':'))
                    attendance.check_in = dt_time(hours, minutes)
                except:
                    attendance.check_in = None
            
            if check_out_str:
                try:
                    hours, minutes = map(int, check_out_str.split(':'))
                    attendance.check_out = dt_time(hours, minutes)
                except:
                    attendance.check_out = None
        else:
            # إذا لم تكن الحالة حاضر، نحذف أوقات الدخول والخروج
            attendance.check_in = None
            attendance.check_out = None
        
        # حفظ التغييرات
        db.session.commit()
        
        # تسجيل العملية في سجل النشاط
        log_attendance_activity(
            action='update',
            attendance_data={
                'id': attendance.id,
                'employee_id': attendance.employee_id,
                'date': attendance.date.isoformat(),
                'old_status': old_status,
                'new_status': status
            },
            employee_name=attendance.employee.name
        )
        
        flash('تم تحديث سجل الحضور بنجاح', 'success')
        
        # العودة لصفحة العرض مع المعاملات
        department_id = request.args.get('department_id', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        return redirect(url_for('attendance.department_attendance_view', 
                               department_id=department_id,
                               start_date=start_date,
                               end_date=end_date))
        
    except Exception as e:
        db.session.rollback()
        print(f"خطأ في تحديث سجل الحضور: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('attendance.edit_attendance_page', id=id))