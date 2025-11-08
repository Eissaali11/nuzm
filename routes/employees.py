import os
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
from flask_login import login_required
from app import db
from models import Employee, Department, SystemAudit, Document, Attendance, Salary, Module, Permission, Vehicle, VehicleHandover,User,Nationality, employee_departments, MobileDevice, DeviceAssignment, EmployeeLocation
from sqlalchemy import func, or_
from utils.excel import parse_employee_excel, generate_employee_excel, export_employee_attendance_to_excel
from utils.date_converter import parse_date
from utils.user_helpers import require_module_access
from utils.employee_comprehensive_report_updated import generate_employee_comprehensive_pdf, generate_employee_comprehensive_excel
from utils.employee_basic_report import generate_employee_basic_pdf
from utils.audit_logger import log_activity

employees_bp = Blueprint('employees', __name__)

# المجلد المخصص لحفظ صور الموظفين
UPLOAD_FOLDER = 'static/uploads/employees'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

def allowed_file(filename):
    """التحقق من أن الملف من الأنواع المسموحة"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_employee_image(file, employee_id, image_type):
    """حفظ صورة الموظف وإرجاع المسار"""
    if file and allowed_file(file.filename):
        # التأكد من وجود المجلد
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # إنشاء اسم ملف فريد
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        unique_filename = f"{employee_id}_{image_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)
        
        # إرجاع المسار النسبي للحفظ في قاعدة البيانات
        return f"static/uploads/employees/{unique_filename}"
    return None

@employees_bp.route('/')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def index():
    """List all employees with filtering options"""
    # الحصول على معاملات الفلترة من URL
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', '')
    multi_department_filter = request.args.get('multi_department', '')
    no_department_filter = request.args.get('no_department', '')
    duplicate_names_filter = request.args.get('duplicate_names', '')
    
    # بناء الاستعلام الأساسي
    query = Employee.query.options(
        db.joinedload(Employee.departments),
        db.joinedload(Employee.nationality_rel)
    )
    
    # فلترة الموظفين حسب القسم المحدد للمستخدم الحالي
    from flask_login import current_user
    if current_user.assigned_department_id:
        # إذا كان المستخدم مرتبط بقسم محدد، عرض موظفي ذلك القسم فقط
        query = query.join(employee_departments).join(Department).filter(Department.id == current_user.assigned_department_id)
    # إذا لم يكن المستخدم مرتبط بقسم، عرض جميع الموظفين (للمديرين العامين)
    
    # تطبيق فلتر القسم (إضافي للفلترة اليدوية)
    elif department_filter:
        query = query.join(employee_departments).join(Department).filter(Department.id == department_filter)
    
    # تطبيق فلتر الحالة
    if status_filter:
        query = query.filter(Employee.status == status_filter)
    
    # تطبيق فلتر الأسماء المكررة
    if duplicate_names_filter == 'yes':
        # البحث عن الأسماء المكررة
        duplicate_names_subquery = db.session.query(Employee.name, func.count(Employee.name).label('name_count'))\
                                           .group_by(Employee.name)\
                                           .having(func.count(Employee.name) > 1)\
                                           .subquery()
        query = query.join(duplicate_names_subquery, Employee.name == duplicate_names_subquery.c.name)
    
    # تطبيق فلتر الموظفين غير المربوطين بأي قسم
    if no_department_filter == 'yes':
        # الموظفين الذين لا يوجد لديهم أي أقسام
        query = query.outerjoin(employee_departments)\
                     .filter(employee_departments.c.employee_id.is_(None))
    elif multi_department_filter == 'yes':
        # الموظفين الذين لديهم أكثر من قسم
        subquery = db.session.query(employee_departments.c.employee_id, 
                                   func.count(employee_departments.c.department_id).label('dept_count'))\
                            .group_by(employee_departments.c.employee_id)\
                            .having(func.count(employee_departments.c.department_id) > 1)\
                            .subquery()
        query = query.join(subquery, Employee.id == subquery.c.employee_id)
    elif multi_department_filter == 'no':
        # الموظفين الذين لديهم قسم واحد فقط أو لا يوجد لديهم أقسام
        subquery = db.session.query(employee_departments.c.employee_id, 
                                   func.count(employee_departments.c.department_id).label('dept_count'))\
                            .group_by(employee_departments.c.employee_id)\
                            .having(func.count(employee_departments.c.department_id) <= 1)\
                            .subquery()
        query = query.outerjoin(subquery, Employee.id == subquery.c.employee_id)\
                     .filter(or_(subquery.c.employee_id.is_(None), 
                               subquery.c.dept_count <= 1))
    
    employees = query.all()
    
    # الحصول على الأقسام للفلتر - مفلترة حسب صلاحيات المستخدم
    if current_user.assigned_department_id:
        # إذا كان المستخدم مرتبط بقسم محدد، عرض ذلك القسم فقط
        departments = Department.query.filter(Department.id == current_user.assigned_department_id).all()
    else:
        # إذا لم يكن المستخدم مرتبط بقسم، عرض جميع الأقسام (للمديرين العامين)
        departments = Department.query.all()
    
    # حساب إحصائيات الموظفين متعددي الأقسام
    multi_dept_count = db.session.query(Employee.id)\
                                .join(employee_departments)\
                                .group_by(Employee.id)\
                                .having(func.count(employee_departments.c.department_id) > 1)\
                                .count()
    
    # حساب الموظفين بدون أقسام
    no_dept_count = db.session.query(Employee.id)\
                             .outerjoin(employee_departments)\
                             .filter(employee_departments.c.employee_id.is_(None))\
                             .count()
    
    # حساب الموظفين بأسماء مكررة - طريقة مبسطة
    duplicate_names_list = db.session.query(Employee.name)\
                                    .group_by(Employee.name)\
                                    .having(func.count(Employee.name) > 1)\
                                    .all()
    
    duplicate_names_count = 0
    duplicate_names_set = set()
    for name_tuple in duplicate_names_list:
        name = name_tuple[0]
        count = db.session.query(Employee).filter(Employee.name == name).count()
        duplicate_names_count += count
        duplicate_names_set.add(name)
    
    single_dept_count = db.session.query(Employee).count() - multi_dept_count - no_dept_count
    
    return render_template('employees/index.html', 
                         employees=employees, 
                         departments=departments,
                         current_department=department_filter,
                         current_status=status_filter,
                         current_multi_department=multi_department_filter,
                         current_no_department=no_department_filter,
                         current_duplicate_names=duplicate_names_filter,
                         multi_dept_count=multi_dept_count,
                         single_dept_count=single_dept_count,
                         no_dept_count=no_dept_count,
                         duplicate_names_count=duplicate_names_count,
                         duplicate_names_set=duplicate_names_set)

@employees_bp.route('/create', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.CREATE)
def create():
    """Create a new employee"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form['name']
            employee_id = request.form['employee_id']
            national_id = request.form['national_id']
            mobile = request.form['mobile']
            status = request.form['status']
            job_title = request.form['job_title']
            location = request.form['location']
            project = request.form['project']
            email = request.form.get('email', '')
            department_id = request.form.get('department_id', None)
            join_date = parse_date(request.form.get('join_date', ''))
            birth_date = parse_date(request.form.get('birth_date', ''))
            mobilePersonal = request.form.get('mobilePersonal')
            nationality_id = request.form.get('nationality_id')
            contract_status = request.form.get('contract_status')
            license_status = request.form.get('license_status')
            
            # الحقول الجديدة لنوع الموظف والعهدة
            employee_type = request.form.get('employee_type', 'regular')
            has_mobile_custody = 'has_mobile_custody' in request.form
            mobile_type = request.form.get('mobile_type', '') if has_mobile_custody else None
            mobile_imei = request.form.get('mobile_imei', '') if has_mobile_custody else None
            
            # حقول الكفالة الجديدة
            sponsorship_status = request.form.get('sponsorship_status', 'inside')
            current_sponsor_name = request.form.get('current_sponsor_name', '')
            
            # معلومات السكن
            residence_details = request.form.get('residence_details', '').strip() or None
            residence_location_url = request.form.get('residence_location_url', '').strip() or None
            
            # معالجة روابط Google Drive
            housing_drive_links = request.form.get('housing_drive_links', '').strip() or None
            
            # مقاسات الزي الموحد
            pants_size = request.form.get('pants_size', '').strip() or None
            shirt_size = request.form.get('shirt_size', '').strip() or None
            
            # الراتب الأساسي
            basic_salary_str = request.form.get('basic_salary', '').strip()
            basic_salary = float(basic_salary_str) if basic_salary_str else 0.0
            
            # حافز الدوام الكامل
            attendance_bonus_str = request.form.get('attendance_bonus', '').strip()
            attendance_bonus = float(attendance_bonus_str) if attendance_bonus_str else 0.0
            
            selected_dept_ids = {int(dept_id) for dept_id in request.form.getlist('department_ids')}
            
            # Convert empty department_id to None
            if department_id == '':
                department_id = None
                
            # Create new employee
            employee = Employee(
                name=name,
                employee_id=employee_id,
                national_id=national_id,
                mobile=mobile,
                status=status,
                job_title=job_title,
                location=location,
                project=project,
                email=email,
                department_id=department_id,
                join_date=join_date,
                birth_date=birth_date,
                mobilePersonal=mobilePersonal,
                nationality_id=int(nationality_id) if nationality_id else None,
                contract_status=contract_status,
                license_status=license_status,
                employee_type=employee_type,
                has_mobile_custody=has_mobile_custody,
                mobile_type=mobile_type,
                mobile_imei=mobile_imei,
                sponsorship_status=sponsorship_status,
                current_sponsor_name=current_sponsor_name,
                residence_details=residence_details,
                residence_location_url=residence_location_url,
                housing_drive_links=housing_drive_links,
                pants_size=pants_size,
                shirt_size=shirt_size,
                basic_salary=basic_salary,
                attendance_bonus=attendance_bonus
            )
            if selected_dept_ids:
                departments_to_assign = Department.query.filter(Department.id.in_(selected_dept_ids)).all()
                employee.departments.extend(departments_to_assign)
            
            db.session.add(employee)
            db.session.commit()
            
            # معالجة رفع صور السكن بعد حفظ الموظف (للحصول على ID)
            housing_images_files = request.files.getlist('housing_images')
            if housing_images_files and any(f.filename for f in housing_images_files):
                saved_images = []
                for img_file in housing_images_files:
                    if img_file and img_file.filename:
                        try:
                            saved_path = save_employee_image(img_file, employee.id, 'housing')
                            if saved_path:
                                saved_images.append(saved_path)
                        except Exception as img_error:
                            print(f"Error saving housing image: {str(img_error)}")
                
                if saved_images:
                    employee.housing_images = ','.join(saved_images)
                    db.session.commit()
            
            # Log the action
            log_activity('create', 'Employee', employee.id, f'تم إنشاء موظف جديد: {name}')
            
            flash('تم إنشاء الموظف بنجاح', 'success')
            return redirect(url_for('employees.index'))
        
        except IntegrityError as e:
            db.session.rollback()
            error_message = str(e)
            if "employee_id" in error_message.lower():
                flash(f"هذه المعلومات مسجلة مسبقاً: رقم الموظف موجود بالفعل في النظام", "danger")
            elif "national_id" in error_message.lower():
                flash(f"هذه المعلومات مسجلة مسبقاً: رقم الهوية موجود بالفعل في النظام", "danger")
            else:
                flash("هذه المعلومات مسجلة مسبقاً، لا يمكن تكرار بيانات الموظفين", "danger")
            
            # إرجاع المستخدم للنموذج مع البيانات المدخلة
            departments = Department.query.all()
            nationalities = Nationality.query.order_by(Nationality.name_ar).all()
            from models import ImportedPhoneNumber
            available_phone_numbers = ImportedPhoneNumber.query.filter(
                ImportedPhoneNumber.employee_id.is_(None)
            ).order_by(ImportedPhoneNumber.phone_number).all()
            from models import MobileDevice
            available_imei_numbers = MobileDevice.query.filter(
                MobileDevice.status == 'متاح',
                MobileDevice.employee_id.is_(None)
            ).order_by(MobileDevice.imei).all()
            
            return render_template('employees/create.html', 
                                 departments=departments,
                                 nationalities=nationalities,
                                 available_phone_numbers=available_phone_numbers,
                                 available_imei_numbers=available_imei_numbers,
                                 form_data=request.form)
                                 
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
            
            # إرجاع المستخدم للنموذج مع البيانات المدخلة
            departments = Department.query.all()
            nationalities = Nationality.query.order_by(Nationality.name_ar).all()
            from models import ImportedPhoneNumber
            available_phone_numbers = ImportedPhoneNumber.query.filter(
                ImportedPhoneNumber.employee_id.is_(None)
            ).order_by(ImportedPhoneNumber.phone_number).all()
            from models import MobileDevice
            available_imei_numbers = MobileDevice.query.filter(
                MobileDevice.status == 'متاح',
                MobileDevice.employee_id.is_(None)
            ).order_by(MobileDevice.imei).all()
            
            return render_template('employees/create.html', 
                                 departments=departments,
                                 nationalities=nationalities,
                                 available_phone_numbers=available_phone_numbers,
                                 available_imei_numbers=available_imei_numbers,
                                 form_data=request.form)
    
    # Get all departments for the dropdown
    departments = Department.query.all()
    nationalities = Nationality.query.order_by(Nationality.name_ar).all()
    
    # جلب الأرقام المتاحة فقط (غير المربوطة بأي موظف)
    from models import ImportedPhoneNumber
    available_phone_numbers = ImportedPhoneNumber.query.filter(
        ImportedPhoneNumber.employee_id.is_(None)  # الأرقام المتاحة فقط
    ).order_by(ImportedPhoneNumber.phone_number).all()
    
    # جلب أرقام IMEI المتاحة من إدارة الأجهزة
    from models import MobileDevice
    available_imei_numbers = MobileDevice.query.filter(
        MobileDevice.status == 'متاح',  # الأجهزة المتاحة فقط
        MobileDevice.employee_id.is_(None)  # غير مربوطة بموظف
    ).order_by(MobileDevice.imei).all()
    
    return render_template('employees/create.html', 
                         departments=departments,
                         nationalities=nationalities,
                         available_phone_numbers=available_phone_numbers,
                         available_imei_numbers=available_imei_numbers)



@employees_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def edit(id):
    """
    تعديل بيانات موظف موجود وأقسامه، مع التحقق من البيانات الفريدة،
    والتعامل الآمن مع تحديث العلاقات، ومزامنة المستخدم المرتبط.
    """
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # 1. استخراج البيانات الجديدة من النموذج
            new_name = request.form.get('name', '').strip()
            new_employee_id = request.form.get('employee_id', '').strip()
            new_national_id = request.form.get('national_id', '').strip()

            # 2. التحقق من صحة البيانات الفريدة قبل أي تعديل
            # التحقق من الرقم الوظيفي
            existing_employee = Employee.query.filter(Employee.employee_id == new_employee_id, Employee.id != id).first()
            if existing_employee:
                flash(f"رقم الموظف '{new_employee_id}' مستخدم بالفعل.", "danger")
                return redirect(url_for('employees.edit', id=id))

            # التحقق من الرقم الوطني
            existing_national = Employee.query.filter(Employee.national_id == new_national_id, Employee.id != id).first()
            if existing_national:
                flash(f"الرقم الوطني '{new_national_id}' مستخدم بالفعل.", "danger")
                return redirect(url_for('employees.edit', id=id))

            # 3. تحديث البيانات الأساسية للموظف
            employee.name = new_name
            employee.employee_id = new_employee_id
            employee.national_id = new_national_id
            # معالجة رقم الجوال مع دعم الإدخال المخصص
            mobile_value = request.form.get('mobile', '')
            print(f"DEBUG: Received mobile value from form: '{mobile_value}'")
            if mobile_value == 'custom':
                mobile_value = request.form.get('mobile_custom', '')
                print(f"DEBUG: Using custom mobile value: '{mobile_value}'")
            employee.mobile = mobile_value
            print(f"DEBUG: Final mobile value set to employee: '{employee.mobile}'")
            
            # تتبع حالة الموظف القديمة قبل التحديث
            old_status = employee.status
            new_status = request.form.get('status', 'active')
            employee.status = new_status
            
            employee.job_title = request.form.get('job_title', '')
            employee.location = request.form.get('location', '')
            employee.project = request.form.get('project', '')
            employee.email = request.form.get('email', '')
            employee.mobilePersonal = request.form.get('mobilePersonal', '')
            employee.contract_status = request.form.get('contract_status', '')
            employee.license_status = request.form.get('license_status', '')
            nationality_id = request.form.get('nationality_id')
            employee.nationality_id = int(nationality_id) if nationality_id else None
            
            # تحديث الحقول الجديدة لنوع الموظف والعهدة
            employee.employee_type = request.form.get('employee_type', 'regular')
            employee.has_mobile_custody = 'has_mobile_custody' in request.form
            employee.mobile_type = request.form.get('mobile_type', '') if employee.has_mobile_custody else None
            employee.mobile_imei = request.form.get('mobile_imei', '') if employee.has_mobile_custody else None
            
            # تحديث حقول الكفالة
            employee.sponsorship_status = request.form.get('sponsorship_status', 'inside')
            employee.current_sponsor_name = request.form.get('current_sponsor_name', '') if employee.sponsorship_status == 'inside' else None
            
            # تحديث حقول المعلومات البنكية
            employee.bank_iban = request.form.get('bank_iban', '').strip() or None
            
            # تحديث معلومات السكن
            employee.residence_details = request.form.get('residence_details', '').strip() or None
            employee.residence_location_url = request.form.get('residence_location_url', '').strip() or None
            
            # معالجة رفع صور السكن (multiple images)
            housing_images_files = request.files.getlist('housing_images')
            if housing_images_files and any(f.filename for f in housing_images_files):
                saved_images = []
                # الاحتفاظ بالصور القديمة إذا كانت موجودة
                if employee.housing_images:
                    saved_images = [img.strip() for img in employee.housing_images.split(',') if img.strip()]
                
                # حفظ الصور الجديدة
                for img_file in housing_images_files:
                    if img_file and img_file.filename:
                        try:
                            saved_path = save_employee_image(img_file, id, 'housing')
                            if saved_path:
                                saved_images.append(saved_path)
                        except Exception as img_error:
                            print(f"Error saving housing image: {str(img_error)}")
                
                # حفظ قائمة الصور كنص مفصول بفواصل
                employee.housing_images = ','.join(saved_images) if saved_images else None
            
            # معالجة روابط Google Drive
            employee.housing_drive_links = request.form.get('housing_drive_links', '').strip() or None
            
            # تحديث مقاسات الزي الموحد
            employee.pants_size = request.form.get('pants_size', '').strip() or None
            employee.shirt_size = request.form.get('shirt_size', '').strip() or None
            
            # تحديث الراتب الأساسي
            basic_salary_str = request.form.get('basic_salary', '').strip()
            employee.basic_salary = float(basic_salary_str) if basic_salary_str else 0.0
            
            # تحديث حافز الدوام الكامل
            attendance_bonus_str = request.form.get('attendance_bonus', '').strip()
            employee.attendance_bonus = float(attendance_bonus_str) if attendance_bonus_str else 0.0
            
            # معالجة رفع صورة شهادة الإيبان
            bank_iban_image_file = request.files.get('bank_iban_image')
            if bank_iban_image_file and bank_iban_image_file.filename:
                # حذف الصورة القديمة إذا كانت موجودة
                if employee.bank_iban_image:
                    old_image_path = os.path.join('static', employee.bank_iban_image)
                    if os.path.exists(old_image_path):
                        os.remove(old_image_path)
                
                # حفظ الصورة الجديدة
                employee.bank_iban_image = save_employee_image(bank_iban_image_file, id, 'iban')
            
            join_date_str = request.form.get('join_date')
            employee.join_date = parse_date(join_date_str) if join_date_str else None
            
            # إضافة معالجة تاريخ الميلاد
            birth_date_str = request.form.get('birth_date')
            employee.birth_date = parse_date(birth_date_str) if birth_date_str else None

            selected_dept_ids = {int(dept_id) for dept_id in request.form.getlist('department_ids')}
            current_dept_ids = {dept.id for dept in employee.departments}

            depts_to_add_ids = selected_dept_ids - current_dept_ids

            if depts_to_add_ids:
                    depts_to_add = Department.query.filter(Department.id.in_(depts_to_add_ids)).all()
                    for dept in depts_to_add:
                        employee.departments.append(dept)
                
            depts_to_remove_ids = current_dept_ids - selected_dept_ids


            if depts_to_remove_ids:
                    depts_to_remove = Department.query.filter(Department.id.in_(depts_to_remove_ids)).all()
                    for dept in depts_to_remove:
                        employee.departments.remove(dept)

            user_linked = User.query.filter_by(employee_id=employee.id).first()

            if user_linked:
                    # الطريقة الأسهل هنا هي فقط تعيين القائمة النهائية بعد تعديلها
                    # بما أننا داخل no_autoflush، يمكننا تعيينها مباشرة
                    # سيقوم SQLAlchemy بحساب الفرق بنفسه عند الـ commit
                    final_departments = Department.query.filter(Department.id.in_(selected_dept_ids)).all()
                    user_linked.departments = final_departments
            
            # 6. إذا تم تغيير الحالة إلى غير نشط، فك ربط جميع أرقام SIM والأجهزة
            if new_status == 'inactive' and old_status != 'inactive':
                try:
                    # استيراد النماذج المطلوبة
                    from models import SimCard, DeviceAssignment, MobileDevice
                    from flask import current_app
                    
                    current_app.logger.info(f"Employee {employee.id} ({employee.name}) became inactive - checking for SIM cards and devices")
                    
                    # البحث عن جميع أرقام SIM المرتبطة مباشرة بهذا الموظف
                    sim_cards = SimCard.query.filter_by(employee_id=employee.id).all()
                    
                    # البحث عن جميع تخصيصات الأجهزة النشطة للموظف
                    device_assignments = DeviceAssignment.query.filter_by(
                        employee_id=employee.id, 
                        is_active=True
                    ).all()
                    
                    total_unlinked = 0
                    
                    # فك ربط أرقام SIM المرتبطة مباشرة
                    current_app.logger.info(f"Found {len(sim_cards)} SIM cards directly linked to employee {employee.id}")
                    
                    if sim_cards:
                        for sim_card in sim_cards:
                            current_app.logger.info(f"Unlinking SIM card {sim_card.phone_number} (ID: {sim_card.id}) from employee {employee.id}")
                            
                            # فك الربط
                            sim_card.employee_id = None
                            sim_card.assigned_date = None
                            sim_card.status = 'متاح'
                            total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="SIM",
                                    entity_id=sim_card.id,
                                    details=f"فك ربط رقم SIM {sim_card.phone_number} تلقائياً بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log SIM audit: {str(audit_e)}")
                    
                    # فك ربط تخصيصات الأجهزة النشطة
                    current_app.logger.info(f"Found {len(device_assignments)} active device assignments for employee {employee.id}")
                    
                    if device_assignments:
                        for assignment in device_assignments:
                            current_app.logger.info(f"Deactivating device assignment {assignment.id} for employee {employee.id}")
                            
                            # إلغاء تنشيط التخصيص
                            assignment.is_active = False
                            assignment.end_date = datetime.now()
                            assignment.end_reason = f'فك ربط تلقائي - تغيير حالة الموظف إلى غير نشط'
                            
                            # فك ربط الجهاز إذا كان موجوداً
                            if assignment.device:
                                assignment.device.employee_id = None
                                assignment.device.status = 'متاح'
                            
                            # فك ربط SIM إذا كان موجوداً
                            if assignment.sim_card:
                                assignment.sim_card.employee_id = None
                                assignment.sim_card.assigned_date = None
                                assignment.sim_card.status = 'متاح'
                                total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                device_info = f"جهاز {assignment.device.brand} {assignment.device.model}" if assignment.device else "بدون جهاز"
                                sim_info = f"رقم {assignment.sim_card.phone_number}" if assignment.sim_card else "بدون رقم"
                                
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="DeviceAssignment",
                                    entity_id=assignment.id,
                                    details=f"فك ربط تخصيص الجهاز تلقائياً ({device_info} - {sim_info}) بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log device assignment audit: {str(audit_e)}")
                    
                    # رسالة نجاح شاملة
                    message_parts = []
                    if len(sim_cards) > 0:
                        message_parts.append(f'{len(sim_cards)} رقم SIM مرتبط مباشرة')
                    if len(device_assignments) > 0:
                        message_parts.append(f'{len(device_assignments)} تخصيص جهاز/رقم')
                    
                    if message_parts:
                        flash(f'تم فك ربط {" و ".join(message_parts)} بالموظف تلقائياً', 'info')
                    
                    current_app.logger.info(f"Successfully processed employee {employee.id} deactivation: {len(sim_cards)} SIM cards, {len(device_assignments)} device assignments")
                
                except Exception as e:
                    current_app.logger.error(f"Error unassigning SIM cards for inactive employee: {str(e)}")
                    flash('تحذير: حدث خطأ في فك ربط أرقام SIM. يرجى فحص الأرقام يدوياً', 'warning')
                    # لا نتوقف عن تحديث حالة الموظف حتى لو فشل فك ربط الأرقام

           
            # 7. حفظ كل التغييرات للموظف والمستخدم دفعة واحدة
            db.session.commit()
            
            # تسجيل عملية التحديث
            try:
                from utils.audit_logger import log_activity
                log_activity('update', 'Employee', employee.id, f'تم تحديث بيانات الموظف: {employee.name}')
            except Exception as audit_e:
                print(f"Failed to log employee update audit: {str(audit_e)}")
                
            flash('تم تحديث بيانات الموظف وأقسامه بنجاح.', 'success')
            
            # التحقق من مصدر الطلب للعودة إلى الصفحة المناسبة
            return_url = request.form.get('return_url')
            if not return_url:
                return_url = request.referrer
            
            if return_url and '/departments/' in return_url:
                # استخراج معرف القسم من الرابط المرجعي
                try:
                    department_id = return_url.split('/departments/')[1].split('/')[0]
                    return redirect(url_for('departments.view', id=department_id))
                except:
                    pass
            
            return redirect(url_for('employees.index'))
        
        except Exception as e:
            # تسجيل الخطأ للمطورين
            flash(f'حدث خطأ غير متوقع أثناء عملية التحديث. يرجى المحاولة مرة أخرى. Error updating employee (ID: {id}): {e}', 'danger')


    # في حالة GET request (عند فتح الصفحة لأول مرة)
    all_departments = Department.query.order_by(Department.name).all()
    all_nationalities = Nationality.query.order_by(Nationality.name_ar).all() # جلب كل الجنسيات
    
    # جلب الأرقام المتاحة فقط (غير المربوطة بأي موظف)
    from models import ImportedPhoneNumber
    available_phone_numbers = ImportedPhoneNumber.query.filter(
        ImportedPhoneNumber.employee_id.is_(None)  # الأرقام المتاحة فقط
    ).order_by(ImportedPhoneNumber.phone_number).all()
    
    # جلب أرقام IMEI المتاحة من إدارة الأجهزة
    from models import MobileDevice
    available_imei_numbers = MobileDevice.query.filter(
        MobileDevice.status == 'متاح',  # الأجهزة المتاحة فقط
        MobileDevice.employee_id.is_(None)  # غير مربوطة بموظف
    ).order_by(MobileDevice.imei).all()
    
    # جلب بيانات الجهاز و SIM المربوط بالموظف من DeviceAssignment
    from models import DeviceAssignment, SimCard
    active_assignment = DeviceAssignment.query.filter_by(
        employee_id=employee.id,
        is_active=True
    ).first()
    
    # بيانات الجهاز و SIM المربوط (سيتم عرضها في الصفحة)
    assigned_device = None
    assigned_sim = None
    
    if active_assignment:
        # جلب الجهاز مباشرة باستخدام device_id
        if active_assignment.device_id:
            assigned_device = MobileDevice.query.get(active_assignment.device_id)
        
        # جلب SIM مباشرة باستخدام sim_card_id
        if active_assignment.sim_card_id:
            assigned_sim = SimCard.query.get(active_assignment.sim_card_id)
    
    print(f"Passing {len(all_nationalities)} nationalities to the template.")
    return render_template('employees/edit.html', 
                         employee=employee, 
                         nationalities=all_nationalities, 
                         departments=all_departments,
                         available_phone_numbers=available_phone_numbers,
                         available_imei_numbers=available_imei_numbers,
                         assigned_device=assigned_device,
                         assigned_sim=assigned_sim)





# @employees_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
# @login_required
# @require_module_access(Module.EMPLOYEES, Permission.EDIT)
# def edit(id):
#     """
#     تعديل بيانات موظف موجود وأقسامه المرتبطة بها، مع مزامنة المستخدم المرتبط.
#     """
#     employee = Employee.query.get_or_404(id)
    
#     if request.method == 'POST':
#         try:
#             # 1. تحديث البيانات الأساسية للموظف
#             employee.name = request.form['name']
#             employee.employee_id = request.form['employee_id']
#             employee.national_id = request.form['national_id']
#             employee.mobile = request.form['mobile']
#             employee.status = request.form['status']
#             employee.job_title = request.form['job_title']
#             employee.location = request.form.get('location', '')
#             employee.project = request.form.get('project', '')
#             employee.email = request.form.get('email', '')
            
#             join_date_str = request.form.get('join_date', '')
#             if join_date_str:
#                 employee.join_date = parse_date(join_date_str) # افترض وجود دالة parse_date

#             # 2. *** تحديث الأقسام المرتبطة (منطق متعدد إلى متعدد) ***
#             # استلام قائمة معرفات الأقسام المحددة من مربعات الاختيار
#             selected_dept_ids = [int(dept_id) for dept_id in request.form.getlist('department_ids')]
            
#             # جلب كائنات الأقسام الفعلية من قاعدة البيانات
#             selected_departments = Department.query.filter(Department.id.in_(selected_dept_ids)).all()
            
#             # تعيين القائمة الجديدة للموظف، وSQLAlchemy سيتولى تحديث جدول الربط
#             employee.departments = selected_departments
            
#             # 3. *** المزامنة التلقائية للمستخدم المرتبط (مهم جداً) ***
#             # ابحث عن المستخدم المرتبط بهذا الموظف (إن وجد)
#             user_linked_to_employee = User.query.filter_by(employee_id=employee.id).first()
#             if user_linked_to_employee:
#                 # إذا وجد مستخدم، قم بمزامنة قائمة أقسامه لتكون مطابقة
#                 user_linked_to_employee.departments = selected_departments
#                 print(f"INFO: Synced departments for linked user: {user_linked_to_employee.name}")
            
#             # 4. حفظ كل التغييرات للموظف والمستخدم في قاعدة البيانات
#             db.session.commit()
            
#             # 5. تسجيل الإجراء والعودة
#             log_activity('update', 'Employee', employee.id, f'تم تحديث بيانات الموظف: {employee.name}')
#             flash('تم تحديث بيانات الموظف وأقسامه بنجاح.', 'success')
#             return redirect(url_for('employees.index'))
        
#         except  Exception as e:
#             db.session.rollback()
#             flash(f"خطأ في التكامل: رقم الموظف أو الرقم الوطني قد يكون مستخدماً بالفعل.{str(e)}", "danger")
#         except Exception as e:
#             db.session.rollback()
#             flash(f'حدث خطأ غير متوقع أثناء التحديث: {str(e)}', 'danger')
#             # من الجيد تسجيل الخطأ الكامل في السجلات للمطورين
#             # current_app.logger.error(f"Error editing employee {id}: {e}")
            
#     # في حالة GET request، جهز البيانات للعرض
#     all_departments = Department.query.order_by(Department.name).all()
#     return render_template('employees/edit.html', employee=employee, departments=all_departments)








@employees_bp.route('/<int:id>/view')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def view(id):
    """View detailed employee information"""
    employee = Employee.query.options(
        db.joinedload(Employee.departments),
        db.joinedload(Employee.nationality_rel)
    ).get_or_404(id)
    
    # Get employee documents
    documents = Document.query.filter_by(employee_id=id).all()
    
    # Get document types in Arabic
    document_types_map = {
        'national_id': 'الهوية الوطنية', 
        'passport': 'جواز السفر', 
        'health_certificate': 'الشهادة الصحية', 
        'work_permit': 'تصريح العمل', 
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'other': 'أخرى'
    }
    
    # Get documents by type for easier display
    documents_by_type = {}
    for doc_type in document_types_map.keys():
        documents_by_type[doc_type] = None
    
    today = datetime.now().date()
    
    for doc in documents:
        # Add expiry status
        days_to_expiry = (doc.expiry_date - today).days
        if days_to_expiry < 0:
            doc.status_class = "danger"
            doc.status_text = "منتهية"
        elif days_to_expiry < 30:
            doc.status_class = "warning"
            doc.status_text = f"تنتهي خلال {days_to_expiry} يوم"
        else:
            doc.status_class = "success"
            doc.status_text = "سارية"
        
        # Store document by type
        documents_by_type[doc.document_type] = doc
    
    # Get all attendance records for this employee
    attendances = Attendance.query.filter_by(employee_id=id).order_by(Attendance.date.desc()).all()
    
    # Get salary records
    salaries = Salary.query.filter_by(employee_id=id).order_by(Salary.year.desc(), Salary.month.desc()).all()
    
    # Get vehicle handover records
    vehicle_handovers = VehicleHandover.query.filter_by(employee_id=id).order_by(VehicleHandover.handover_date.desc()).all()
    
    # Get mobile devices assigned to this employee
    mobile_devices = MobileDevice.query.filter_by(employee_id=id).order_by(MobileDevice.assigned_date.desc()).all()
    
    # Get device assignments for this employee
    from models import DeviceAssignment
    device_assignments = DeviceAssignment.query.filter_by(
        employee_id=id, 
        is_active=True
    ).options(
        db.joinedload(DeviceAssignment.device),
        db.joinedload(DeviceAssignment.sim_card)
    ).all()
    
    all_departments = Department.query.order_by(Department.name).all()
    
    # جلب معلومات السكن (العقارات التي يقطن فيها الموظف)
    housing_properties = employee.housing_properties
    
    return render_template('employees/view.html', 
                          employee=employee, 
                          documents=documents,
                          documents_by_type=documents_by_type,
                          document_types_map=document_types_map,
                          attendances=attendances,
                          salaries=salaries,
                          vehicle_handovers=vehicle_handovers,
                          mobile_devices=mobile_devices,
                          device_assignments=device_assignments,
                          departments=all_departments,
                          housing_properties=housing_properties
                          )

@employees_bp.route('/<int:id>/upload_iban', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def upload_iban(id):
    """رفع صورة الإيبان البنكي للموظف"""
    employee = Employee.query.get_or_404(id)
    
    try:
        # الحصول على بيانات الإيبان والملف
        bank_iban = request.form.get('bank_iban', '').strip()
        iban_file = request.files.get('iban_image')
        
        # تحديث رقم الإيبان
        if bank_iban:
            employee.bank_iban = bank_iban
        
        # رفع صورة الإيبان إذا تم اختيارها
        if iban_file and iban_file.filename:
            # حذف الصورة القديمة إذا كانت موجودة
            if employee.bank_iban_image:
                old_image_path = os.path.join('static', employee.bank_iban_image)
                if os.path.exists(old_image_path):
                    os.remove(old_image_path)
            
            # حفظ الصورة الجديدة
            image_path = save_employee_image(iban_file, employee.id, 'iban')
            if image_path:
                employee.bank_iban_image = image_path
        
        db.session.commit()
        
        # تسجيل العملية
        log_activity('update', 'Employee', employee.id, f'تم تحديث بيانات الإيبان البنكي للموظف: {employee.name}')
        
        flash('تم حفظ بيانات الإيبان البنكي بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حفظ بيانات الإيبان: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/delete_iban_image', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def delete_iban_image(id):
    """حذف صورة الإيبان البنكي للموظف"""
    employee = Employee.query.get_or_404(id)
    
    try:
        if employee.bank_iban_image:
            # حذف الملف من الخادم
            image_path = os.path.join('static', employee.bank_iban_image)
            if os.path.exists(image_path):
                os.remove(image_path)
            
            # حذف المسار من قاعدة البيانات
            employee.bank_iban_image = None
            db.session.commit()
            
            # تسجيل العملية
            log_activity('delete', 'Employee', employee.id, f'تم حذف صورة الإيبان البنكي للموظف: {employee.name}')
            
            flash('تم حذف صورة الإيبان البنكي بنجاح', 'success')
        else:
            flash('لا توجد صورة إيبان لحذفها', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف صورة الإيبان: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/delete_housing_image', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def delete_housing_image(id):
    """حذف صورة من صور السكن التوضيحية"""
    employee = Employee.query.get_or_404(id)
    image_path = request.form.get('image_path', '').strip()
    
    try:
        if not image_path:
            flash('لم يتم تحديد الصورة المراد حذفها', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        if employee.housing_images:
            # تحويل القائمة إلى list
            image_list = [img.strip() for img in employee.housing_images.split(',')]
            
            # البحث عن الصورة في القائمة
            clean_image_path = image_path.replace('static/', '')
            image_to_remove = None
            
            for img in image_list:
                if img.replace('static/', '') == clean_image_path:
                    image_to_remove = img
                    break
            
            if image_to_remove:
                # حذف الصورة من القائمة
                image_list.remove(image_to_remove)
                
                # حذف الملف من الخادم
                full_path = os.path.join('static', clean_image_path)
                if os.path.exists(full_path):
                    os.remove(full_path)
                
                # تحديث قاعدة البيانات
                employee.housing_images = ','.join(image_list) if image_list else None
                db.session.commit()
                
                # تسجيل العملية
                log_activity('delete', 'Employee', employee.id, f'تم حذف صورة من صور السكن للموظف: {employee.name}')
                
                flash('تم حذف الصورة بنجاح', 'success')
            else:
                flash('لم يتم العثور على الصورة في القائمة', 'warning')
        else:
            flash('لا توجد صور سكن لحذفها', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الصورة: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/confirm_delete')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.DELETE)
def confirm_delete(id):
    """صفحة تأكيد حذف الموظف"""
    employee = Employee.query.get_or_404(id)
    
    # تحديد عنوان الصفحة التي تم تحويلنا منها للعودة إليها عند الإلغاء
    return_url = request.referrer
    if not return_url or '/employees/' in return_url:
        return_url = url_for('employees.index')
    
    return render_template('employees/confirm_delete.html', 
                          employee=employee, 
                          return_url=return_url)

@employees_bp.route('/<int:id>/delete', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.DELETE)
def delete(id):
    """Delete an employee"""
    employee = Employee.query.get_or_404(id)
    name = employee.name
    
    # إذا كان الطلب GET، نعرض صفحة التأكيد
    if request.method == 'GET':
        return redirect(url_for('employees.confirm_delete', id=id))
    
    # إذا كان الطلب POST، نتحقق من تأكيد الحذف
    confirmed = request.form.get('confirmed', 'no')
    
    if confirmed != 'yes':
        flash('لم يتم تأكيد عملية الحذف', 'warning')
        return redirect(url_for('employees.view', id=id))
    
    try:
        # حذف جميع تعيينات الأجهزة المرتبطة بالموظف
        DeviceAssignment.query.filter_by(employee_id=id).delete()
        
        db.session.delete(employee)
        db.session.commit()
        
        # Log the action
        log_activity('delete', 'Employee', id, f'تم حذف الموظف: {name}')
        
        flash('تم حذف الموظف بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الموظف: {str(e)}', 'danger')
    
    # التحقق من مصدر الطلب للعودة إلى الصفحة المناسبة
    referrer = request.form.get('return_url')
    if referrer and '/departments/' in referrer:
        try:
            department_id = referrer.split('/departments/')[1].split('/')[0]
            return redirect(url_for('departments.view', id=department_id))
        except:
            pass
    
    return redirect(url_for('employees.index'))

@employees_bp.route('/import', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.CREATE)
def import_excel():
    """Import employees from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                print(f"Received file: {file.filename}")
                
                # Parse Excel file
                employees_data = parse_employee_excel(file)
                print(f"Parsed {len(employees_data)} employee records from Excel")
                
                success_count = 0
                error_count = 0
                error_details = []
                
                for index, data in enumerate(employees_data):
                    try:
                        print(f"Processing employee {index+1}: {data.get('name', 'Unknown')}")
                        
                        # Check if employee with same employee_id already exists
                        existing = Employee.query.filter_by(employee_id=data['employee_id']).first()
                        if existing:
                            print(f"Employee with ID {data['employee_id']} already exists")
                            error_count += 1
                            error_details.append(f"الموظف برقم {data['employee_id']} موجود مسبقا")
                            continue
                            
                        # Check if employee with same national_id already exists
                        existing = Employee.query.filter_by(national_id=data['national_id']).first()
                        if existing:
                            print(f"Employee with national ID {data['national_id']} already exists")
                            error_count += 1
                            error_details.append(f"الموظف برقم هوية {data['national_id']} موجود مسبقا")
                            continue
                        
                        # Extract department data separately
                        department_name = data.pop('department', None)
                        
                        # Create employee without department field
                        employee = Employee(**data)
                        db.session.add(employee)
                        db.session.flush()  # Get the ID without committing
                        
                        # Handle department assignment if provided
                        if department_name:
                            department = Department.query.filter_by(name=department_name).first()
                            if department:
                                employee.departments.append(department)
                            else:
                                # Create new department if it doesn't exist
                                new_department = Department(name=department_name)
                                db.session.add(new_department)
                                db.session.flush()
                                employee.departments.append(new_department)
                        
                        db.session.commit()
                        success_count += 1
                        print(f"Successfully added employee: {data.get('name')}")
                    except Exception as e:
                        db.session.rollback()
                        error_count += 1
                        print(f"Error adding employee {index+1}: {str(e)}")
                        error_details.append(f"خطأ في السجل {index+1}: {str(e)}")
                
                # Log the import
                error_detail_str = ", ".join(error_details[:5])
                if len(error_details) > 5:
                    error_detail_str += f" وغيرها من الأخطاء..."
                
                details = f'تم استيراد {success_count} موظف بنجاح و {error_count} فشل'
                if error_details:
                    details += f". أخطاء: {error_detail_str}"
                    
                audit = SystemAudit(
                    action='import',
                    entity_type='employee',
                    entity_id=0,
                    details=details
                )
                db.session.add(audit)
                db.session.commit()
                
                if error_count > 0:
                    flash(f'تم استيراد {success_count} موظف بنجاح و {error_count} فشل. {error_detail_str}', 'warning')
                else:
                    flash(f'تم استيراد {success_count} موظف بنجاح', 'success')
                return redirect(url_for('employees.index'))
            except Exception as e:
                flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
        else:
            flash('الملف يجب أن يكون بصيغة Excel (.xlsx, .xls)', 'danger')
    
    return render_template('employees/import.html')

@employees_bp.route('/import/template')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def import_template():
    """Download Excel template for employee import with all comprehensive fields"""
    try:
        import pandas as pd
        
        # إنشاء قالب Excel مع جميع الحقول المطلوبة والاختيارية
        template_data = {
            'الاسم الكامل': ['محمد أحمد علي', 'فاطمة سالم محمد'],
            'رقم الموظف': ['EMP001', 'EMP002'],
            'رقم الهوية الوطنية': ['1234567890', '0987654321'],
            'رقم الجوال': ['0501234567', '0509876543'],
            'الجوال الشخصي': ['0551234567', ''],
            'المسمى الوظيفي': ['مطور برمجيات', 'محاسبة'],
            'الحالة الوظيفية': ['active', 'active'],
            'الموقع': ['الرياض', 'جدة'],
            'المشروع': ['مشروع الرياض', 'مشروع جدة'],
            'البريد الإلكتروني': ['mohamed@company.com', 'fatima@company.com'],
            'الأقسام': ['تقنية المعلومات', 'المحاسبة'],
            'تاريخ الانضمام': ['2024-01-15', '2024-02-01'],
            'تاريخ انتهاء الإقامة': ['2025-12-31', '2025-11-30'],
            'حالة العقد': ['محدد المدة', 'دائم'],
            'حالة الرخصة': ['سارية', 'سارية'],
            'الجنسية': ['سعودي', 'مصري'],
            'ملاحظات': ['موظف متميز', '']
        }
        
        # إنشاء DataFrame
        df = pd.DataFrame(template_data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # كتابة البيانات النموذجية
            df.to_excel(writer, sheet_name='البيانات النموذجية', index=False)
            
            # إنشاء ورقة فارغة للاستخدام
            empty_df = pd.DataFrame(columns=template_data.keys())
            empty_df.to_excel(writer, sheet_name='استيراد الموظفين', index=False)
            
            # إنشاء ورقة التعليمات
            instructions_data = {
                'العمود': list(template_data.keys()),
                'مطلوب/اختياري': ['مطلوب', 'مطلوب', 'مطلوب', 'مطلوب', 'اختياري', 'مطلوب', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري'],
                'التنسيق المطلوب': [
                    'نص',
                    'نص فريد',
                    'رقم من 10 أرقام',
                    'رقم جوال سعودي',
                    'رقم جوال (اختياري)',
                    'نص',
                    'active/inactive/on_leave',
                    'نص',
                    'نص',
                    'بريد إلكتروني صحيح',
                    'اسم القسم',
                    'YYYY-MM-DD',
                    'YYYY-MM-DD',
                    'نص',
                    'نص',
                    'اسم الجنسية',
                    'نص (اختياري)'
                ]
            }
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='التعليمات', index=False)
        
        output.seek(0)
        
        # تسجيل العملية
        audit = SystemAudit(
            action='download_template',
            entity_type='employee_import',
            entity_id=0,
            details='تم تحميل قالب استيراد الموظفين المحسن'
        )
        db.session.add(audit)
        db.session.commit()
        
        return send_file(
            output,
            download_name='قالب_استيراد_الموظفين_شامل.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ في إنشاء القالب: {str(e)}', 'danger')
        return redirect(url_for('employees.import_excel'))

@employees_bp.route('/import/empty_template')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def empty_import_template():
    """Download empty Excel template for employee import"""
    try:
        import pandas as pd
        
        # إنشاء قالب فارغ مع جميع الحقول المطلوبة
        empty_template_data = {
            'الاسم الكامل': [],
            'رقم الموظف': [],
            'رقم الهوية الوطنية': [],
            'رقم الجوال': [],
            'الجوال الشخصي': [],
            'المسمى الوظيفي': [],
            'الحالة الوظيفية': [],
            'الموقع': [],
            'المشروع': [],
            'البريد الإلكتروني': [],
            'الأقسام': [],
            'تاريخ الانضمام': [],
            'تاريخ انتهاء الإقامة': [],
            'حالة العقد': [],
            'حالة الرخصة': [],
            'الجنسية': [],
            'ملاحظات': []
        }
        
        # إنشاء DataFrame فارغ
        df = pd.DataFrame(empty_template_data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # كتابة القالب الفارغ
            df.to_excel(writer, sheet_name='استيراد الموظفين', index=False)
            
            # إنشاء ورقة التعليمات
            instructions_data = {
                'العمود': [
                    'الاسم الكامل', 'رقم الموظف', 'رقم الهوية الوطنية', 'رقم الجوال', 
                    'الجوال الشخصي', 'المسمى الوظيفي', 'الحالة الوظيفية', 'الموقع', 
                    'المشروع', 'البريد الإلكتروني', 'الأقسام', 'تاريخ الانضمام', 
                    'تاريخ انتهاء الإقامة', 'حالة العقد', 'حالة الرخصة', 'الجنسية', 'ملاحظات'
                ],
                'مطلوب/اختياري': [
                    'مطلوب', 'مطلوب', 'مطلوب', 'مطلوب', 'اختياري', 'مطلوب', 
                    'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 
                    'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري'
                ],
                'التنسيق المطلوب': [
                    'نص', 'نص فريد', 'رقم من 10 أرقام', 'رقم جوال سعودي', 
                    'رقم جوال (اختياري)', 'نص', 'active/inactive/on_leave', 'نص', 
                    'نص', 'بريد إلكتروني صحيح', 'اسم القسم', 'YYYY-MM-DD', 
                    'YYYY-MM-DD', 'نص', 'نص', 'اسم الجنسية', 'نص (اختياري)'
                ],
                'مثال': [
                    'محمد أحمد علي', 'EMP001', '1234567890', '0501234567',
                    '0551234567', 'مطور برمجيات', 'active', 'الرياض',
                    'مشروع الرياض', 'mohamed@company.com', 'تقنية المعلومات', '2024-01-15',
                    '2025-12-31', 'محدد المدة', 'سارية', 'سعودي', 'موظف متميز'
                ]
            }
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='التعليمات والأمثلة', index=False)
        
        output.seek(0)
        
        # تسجيل العملية
        audit = SystemAudit(
            action='download_empty_template',
            entity_type='employee_import',
            entity_id=0,
            details='تم تحميل نموذج فارغ لاستيراد الموظفين'
        )
        db.session.add(audit)
        db.session.commit()
        
        return send_file(
            output,
            download_name='نموذج_استيراد_الموظفين_فارغ.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ في إنشاء النموذج الفارغ: {str(e)}', 'danger')
        return redirect(url_for('employees.import_excel'))

@employees_bp.route('/<int:id>/update_status', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def update_status(id):
    """تحديث حالة الموظف"""
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            new_status = request.form.get('status')
            if new_status not in ['active', 'inactive', 'on_leave']:
                flash('حالة غير صالحة', 'danger')
                return redirect(url_for('employees.view', id=id))
            
            old_status = employee.status
            employee.status = new_status
            
            note = request.form.get('note', '')
            
            # إذا تم تغيير الحالة إلى غير نشط، فك ربط جميع أرقام SIM والأجهزة
            if new_status == 'inactive' and old_status != 'inactive':
                try:
                    # استيراد النماذج المطلوبة
                    from models import SimCard, DeviceAssignment, MobileDevice
                    from flask import current_app
                    
                    current_app.logger.info(f"Checking SIM cards and devices for employee {employee.id} ({employee.name}) who became inactive")
                    
                    # 1. البحث عن جميع أرقام SIM المرتبطة مباشرة بهذا الموظف
                    sim_cards = SimCard.query.filter_by(employee_id=employee.id).all()
                    
                    # 2. البحث عن جميع تخصيصات الأجهزة النشطة للموظف
                    device_assignments = DeviceAssignment.query.filter_by(
                        employee_id=employee.id, 
                        is_active=True
                    ).all()
                    
                    total_unlinked = 0
                    
                    # فك ربط أرقام SIM المرتبطة مباشرة
                    current_app.logger.info(f"Found {len(sim_cards)} SIM cards directly linked to employee {employee.id}")
                    
                    if sim_cards:
                        for sim_card in sim_cards:
                            current_app.logger.info(f"Unlinking SIM card {sim_card.phone_number} (ID: {sim_card.id}) from employee {employee.id}")
                            
                            # فك الربط
                            sim_card.employee_id = None
                            sim_card.assigned_date = None
                            sim_card.status = 'متاح'
                            total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="SIM",
                                    entity_id=sim_card.id,
                                    details=f"فك ربط رقم SIM {sim_card.phone_number} تلقائياً بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log SIM audit: {str(audit_e)}")
                    
                    # فك ربط تخصيصات الأجهزة النشطة
                    current_app.logger.info(f"Found {len(device_assignments)} active device assignments for employee {employee.id}")
                    
                    if device_assignments:
                        for assignment in device_assignments:
                            current_app.logger.info(f"Deactivating device assignment {assignment.id} for employee {employee.id}")
                            
                            # إلغاء تنشيط التخصيص
                            assignment.is_active = False
                            assignment.end_date = datetime.now()
                            assignment.end_reason = f'فك ربط تلقائي - تغيير حالة الموظف إلى غير نشط'
                            
                            # فك ربط الجهاز إذا كان موجوداً
                            if assignment.device:
                                assignment.device.employee_id = None
                                assignment.device.status = 'متاح'
                            
                            # فك ربط SIM إذا كان موجوداً
                            if assignment.sim_card:
                                assignment.sim_card.employee_id = None
                                assignment.sim_card.assigned_date = None
                                assignment.sim_card.status = 'متاح'
                                total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                device_info = f"جهاز {assignment.device.brand} {assignment.device.model}" if assignment.device else "بدون جهاز"
                                sim_info = f"رقم {assignment.sim_card.phone_number}" if assignment.sim_card else "بدون رقم"
                                
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="DeviceAssignment",
                                    entity_id=assignment.id,
                                    details=f"فك ربط تخصيص الجهاز تلقائياً ({device_info} - {sim_info}) بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log device assignment audit: {str(audit_e)}")
                    
                    # حفظ التغييرات في قاعدة البيانات
                    db.session.commit()
                    
                    # رسالة نجاح شاملة
                    message_parts = []
                    if len(sim_cards) > 0:
                        message_parts.append(f'{len(sim_cards)} رقم SIM مرتبط مباشرة')
                    if len(device_assignments) > 0:
                        message_parts.append(f'{len(device_assignments)} تخصيص جهاز/رقم')
                    
                    if message_parts:
                        flash(f'تم فك ربط {" و ".join(message_parts)} بالموظف تلقائياً', 'info')
                    
                    current_app.logger.info(f"Successfully processed employee {employee.id} deactivation: {len(sim_cards)} SIM cards, {len(device_assignments)} device assignments")
                
                except Exception as e:
                    current_app.logger.error(f"Error unassigning SIM cards for inactive employee: {str(e)}")
                    db.session.rollback()
                    flash('تحذير: حدث خطأ في فك ربط أرقام SIM. يرجى فحص الأرقام يدوياً', 'warning')
                    # لا نتوقف عن تحديث حالة الموظف حتى لو فشل فك ربط الأرقام
            
            # توثيق التغيير في السجل
            status_names = {
                'active': 'نشط',
                'inactive': 'غير نشط',
                'on_leave': 'في إجازة'
            }
            
            details = f'تم تغيير حالة الموظف {employee.name} من "{status_names.get(old_status, old_status)}" إلى "{status_names.get(new_status, new_status)}"'
            if note:
                details += f" - ملاحظات: {note}"
                
            # تسجيل العملية
            audit = SystemAudit(
                action='update_status',
                entity_type='employee',
                entity_id=employee.id,
                details=details
            )
            db.session.add(audit)
            db.session.commit()
            
            flash(f'تم تحديث حالة الموظف إلى {status_names.get(new_status, new_status)} بنجاح', 'success')
            
            # العودة إلى الصفحة السابقة
            referrer = request.referrer
            if referrer and '/departments/' in referrer:
                department_id = referrer.split('/departments/')[1].split('/')[0]
                return redirect(url_for('departments.view', id=department_id))
            
            return redirect(url_for('employees.view', id=id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث حالة الموظف: {str(e)}', 'danger')
            return redirect(url_for('employees.view', id=id))

@employees_bp.route('/export')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def export_excel():
    """Export employees to Excel file"""
    try:
        employees = Employee.query.options(
            db.joinedload(Employee.departments),
            db.joinedload(Employee.nationality_rel)
        ).all()
        output = generate_employee_excel(employees)
        
        # Log the export
        audit = SystemAudit(
            action='export',
            entity_type='employee',
            entity_id=0,
            details=f'تم تصدير {len(employees)} موظف إلى ملف Excel'
        )
        db.session.add(audit)
        db.session.commit()
        
        return send_file(
            BytesIO(output.getvalue()),
            download_name='employees.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('employees.index'))

@employees_bp.route('/export_comprehensive')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def export_comprehensive():
    """تصدير شامل لبيانات الموظفين مع جميع التفاصيل والعُهد والمعلومات البنكية"""
    try:
        from utils.basic_comprehensive_export import generate_comprehensive_employee_excel
        
        employees = Employee.query.options(
            db.joinedload(Employee.departments),
            db.joinedload(Employee.nationality_rel),
            db.joinedload(Employee.salaries),
            db.joinedload(Employee.attendances),
            db.joinedload(Employee.documents)
        ).all()
        
        output = generate_comprehensive_employee_excel(employees)
        
        # تسجيل العملية
        audit = SystemAudit(
            action='export_comprehensive',
            entity_type='employee',
            entity_id=0,
            details=f'تم التصدير الشامل لبيانات {len(employees)} موظف مع جميع التفاصيل'
        )
        db.session.add(audit)
        db.session.commit()
        
        # إنشاء اسم الملف مع التاريخ
        current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'تصدير_شامل_الموظفين_{current_date}.xlsx'
        
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        print(f"Error in comprehensive export: {str(e)}")
        print(traceback.format_exc())
        flash(f'حدث خطأ أثناء التصدير الشامل: {str(e)}', 'danger')
        return redirect(url_for('employees.index'))
        
@employees_bp.route('/<int:id>/export_attendance_excel')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def export_attendance_excel(id):
    """تصدير بيانات الحضور كملف إكسل"""
    try:
        # الحصول على بيانات الموظف
        employee = Employee.query.get_or_404(id)
        
        # الحصول على الشهر والسنة من معاملات الطلب
        month = request.args.get('month')
        year = request.args.get('year')
        
        # تحويل البيانات إلى أرقام صحيحة إذا كانت موجودة
        if month:
            try:
                month = int(month)
            except (ValueError, TypeError):
                flash('قيمة الشهر غير صالحة، تم استخدام الشهر الحالي', 'warning')
                month = None
                
        if year:
            try:
                year = int(year)
            except (ValueError, TypeError):
                flash('قيمة السنة غير صالحة، تم استخدام السنة الحالية', 'warning')
                year = None
        
        # توليد ملف الإكسل
        output = export_employee_attendance_to_excel(employee, month, year)
        
        # تعيين اسم الملف مع التاريخ الحالي
        current_date = datetime.now().strftime('%Y%m%d')
        
        # إضافة الشهر والسنة إلى اسم الملف إذا كانا موجودين
        if month and year:
            filename = f"attendance_{employee.name}_{year}_{month}_{current_date}.xlsx"
        else:
            # استخدام الشهر والسنة الحالية إذا لم يتم توفيرهما
            current_month = datetime.now().month
            current_year = datetime.now().year
            filename = f"attendance_{employee.name}_{current_year}_{current_month}_{current_date}.xlsx"
        
        # تسجيل الإجراء
        audit = SystemAudit(
            action='export',
            entity_type='attendance',
            entity_id=employee.id,
            details=f'تم تصدير سجل الحضور للموظف: {employee.name}'
        )
        db.session.add(audit)
        db.session.commit()
        
        # إرسال ملف الإكسل
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # طباعة تتبع الخطأ في سجل الخادم للمساعدة في التشخيص
        import traceback
        print(f"Error exporting attendance: {str(e)}")
        print(traceback.format_exc())
        
        flash(f'حدث خطأ أثناء تصدير ملف الحضور: {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/upload_image', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def upload_image(id):
    """رفع صورة للموظف (الصورة الشخصية، صورة الهوية، أو صورة الرخصة)"""
    employee = Employee.query.get_or_404(id)
    
    image_type = request.form.get('image_type')
    if not image_type or image_type not in ['profile', 'national_id', 'license']:
        flash('نوع الصورة غير صحيح', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    if 'image' not in request.files:
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    file = request.files['image']
    if file.filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    # حفظ الصورة
    image_path = save_employee_image(file, employee.employee_id, image_type)
    if image_path:
        try:
            # حذف الصورة القديمة إذا كانت موجودة
            old_path = None
            if image_type == 'profile':
                old_path = employee.profile_image
                employee.profile_image = image_path
            elif image_type == 'national_id':
                old_path = employee.national_id_image
                employee.national_id_image = image_path
            elif image_type == 'license':
                old_path = employee.license_image
                employee.license_image = image_path
            
            # حذف الصورة القديمة من الملفات
            if old_path:
                old_file_path = os.path.join('static', old_path)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            
            db.session.commit()
            
            # رسائل النجاح حسب نوع الصورة
            success_messages = {
                'profile': 'تم رفع الصورة الشخصية بنجاح',
                'national_id': 'تم رفع صورة الهوية بنجاح',
                'license': 'تم رفع صورة الرخصة بنجاح'
            }
            flash(success_messages[image_type], 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في حفظ الصورة: {str(e)}', 'danger')
    else:
        flash('فشل في رفع الصورة. تأكد من أن الملف من النوع المسموح (PNG, JPG, JPEG, GIF)', 'danger')
    
    return redirect(url_for('employees.view', id=id))


@employees_bp.route('/<int:id>/basic_report')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def basic_report(id):
    """تقرير المعلومات الأساسية للموظف"""
    try:
                # طباعة رسالة تشخيصية
        print("بدء إنشاء التقرير الشامل للموظف")
        
        # التحقق من وجود الموظف
        employee = Employee.query.get_or_404(id)
        print(f"تم العثور على الموظف: {employee.name}")
        
        # إنشاء ملف PDF
        print("استدعاء دالة إنشاء PDF")
        pdf_buffer = generate_employee_basic_pdf(id)
        print("تم استلام ناتج ملف PDF")
        
        if not pdf_buffer:
            flash('لم يتم العثور على بيانات كافية لإنشاء التقرير', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        if pdf_buffer:
            employee = Employee.query.get_or_404(id)
            current_date = datetime.now().strftime('%Y%m%d')
            filename = f'تقرير_أساسي_{employee.name}_{current_date}.pdf'
            
            # تسجيل الإجراء
            audit = SystemAudit(
                action='export',
                entity_type='employee_basic_report',
                entity_id=employee.id,
                details=f'تم تصدير التقرير الأساسي للموظف: {employee.name}'
            )
            db.session.add(audit)
            db.session.commit()
            
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        else:
            flash('خطأ في إنشاء ملف PDF', 'danger')
            return redirect(url_for('employees.view', id=id))
    except Exception as e:
        flash(f'خطأ في تصدير PDF: {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))


@employees_bp.route('/<int:id>/comprehensive_report')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def comprehensive_report(id):
    """تقرير شامل عن الموظف بصيغة PDF"""
    try:
        # طباعة رسالة تشخيصية
        print("بدء إنشاء التقرير الشامل للموظف")
        
        # التحقق من وجود الموظف
        employee = Employee.query.get_or_404(id)
        print(f"تم العثور على الموظف: {employee.name}")
        
        # إنشاء ملف PDF
        print("استدعاء دالة إنشاء PDF")
        output = generate_employee_comprehensive_pdf(id)
        print("تم استلام ناتج ملف PDF")
        
        if not output:
            flash('لم يتم العثور على بيانات كافية لإنشاء التقرير', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        # اسم الملف المُصدَّر
        filename = f"تقرير_شامل_{employee.name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        print(f"اسم الملف: {filename}")
        
        # تسجيل عملية التصدير
        audit = SystemAudit(
            action='export',
            entity_type='employee_report',
            entity_id=employee.id,
            details=f'تم إنشاء تقرير شامل للموظف: {employee.name}'
        )
        db.session.add(audit)
        db.session.commit()
        print("تم تسجيل العملية في سجل النظام")
        
        # طباعة نوع ناتج الملف للتشخيص
        print(f"نوع ناتج الملف: {type(output)}")
        print(f"حجم البيانات: {output.getbuffer().nbytes} بايت")
        
        # إرسال ملف PDF
        print("إرسال الملف للمتصفح")
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        # طباعة تتبع الخطأ في سجل الخادم للمساعدة في التشخيص
        import traceback
        print(f"Error generating comprehensive report: {str(e)}")
        print(traceback.format_exc())
        
        flash(f'حدث خطأ أثناء إنشاء التقرير الشامل: {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))


@employees_bp.route('/<int:id>/comprehensive_report_excel')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def comprehensive_report_excel(id):
    """تقرير شامل عن الموظف بصيغة Excel"""
    try:
        # التحقق من وجود الموظف
        employee = Employee.query.get_or_404(id)
        
        # إنشاء ملف Excel
        output = generate_employee_comprehensive_excel(id)
        
        if not output:
            flash('لم يتم العثور على بيانات كافية لإنشاء التقرير', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        # اسم الملف المُصدَّر
        filename = f"تقرير_شامل_{employee.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # تسجيل عملية التصدير
        audit = SystemAudit(
            action='export',
            entity_type='employee_report_excel',
            entity_id=employee.id,
            details=f'تم تصدير تقرير شامل (إكسل) للموظف: {employee.name}'
        )
        db.session.add(audit)
        db.session.commit()
        
        # إرسال ملف الإكسل
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # طباعة تتبع الخطأ في سجل الخادم للمساعدة في التشخيص
        import traceback
        print(f"Error generating comprehensive Excel report: {str(e)}")
        print(traceback.format_exc())
        
        flash(f'حدث خطأ أثناء إنشاء التقرير الشامل (إكسل): {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))


@employees_bp.route('/tracking')
@login_required
def tracking():
    """صفحة تتبع مواقع الموظفين عبر GPS"""
    from flask_login import current_user
    
    # TODO: لاحقاً فعّل هذا القيد للمديرين فقط
    # if current_user.role != 'admin':
    #     flash('هذه الصفحة متاحة للمديرين فقط', 'danger')
    #     return redirect(url_for('dashboard.index'))
    
    # الحصول على معاملات الفلترة
    department_filter = request.args.get('department', '')
    search_query = request.args.get('search', '')
    
    # بناء استعلام الموظفين النشطين فقط
    query = Employee.query.filter(Employee.status == 'active').options(
        db.joinedload(Employee.departments)
    )
    
    # تطبيق فلتر القسم
    if department_filter:
        query = query.join(employee_departments).join(Department).filter(Department.id == department_filter)
    
    # تطبيق فلتر البحث (اسم أو رقم وظيفي)
    if search_query:
        query = query.filter(
            or_(
                Employee.name.contains(search_query),
                Employee.employee_id.contains(search_query)
            )
        )
    
    # جلب جميع الموظفين
    all_employees = query.all()
    
    # جلب آخر موقع لكل موظف وترتيبهم
    employee_locations = {}
    employees_with_location = []
    employees_without_location = []
    
    for emp in all_employees:
        # جلب أحدث موقع للموظف باستخدام employee.id
        latest_location = EmployeeLocation.query.filter_by(
            employee_id=emp.id
        ).order_by(EmployeeLocation.recorded_at.desc()).first()
        
        if latest_location:
            # حساب عمر الموقع بالساعات
            age_hours = (datetime.utcnow() - latest_location.recorded_at).total_seconds() / 3600
            
            # تحديد اللون حسب عمر الموقع
            if age_hours < 1:
                color = 'green'
                status_text = 'نشط'
            elif age_hours < 6:
                color = 'orange'
                status_text = 'متوسط'
            else:
                color = 'red'
                status_text = 'قديم'
            
            employee_locations[emp.id] = {
                'latitude': latest_location.latitude,
                'longitude': latest_location.longitude,
                'accuracy': getattr(latest_location, 'accuracy_m', None), 
                'recorded_at': latest_location.recorded_at,
                'age_hours': age_hours,
                'color': color,
                'status_text': status_text
            }
            employees_with_location.append(emp)
        else:
            employees_without_location.append(emp)
    
    # ترتيب: الموظفون الذين لديهم موقع أولاً
    employees = employees_with_location + employees_without_location
    
    # تحويل الموظفين إلى قواميس لكي يمكن تحويلها إلى JSON
    employees_data = []
    for emp in employees:
        emp_dict = {
            'id': emp.id,
            'name': emp.name,
            'employee_id': emp.employee_id,
            'photo_url': emp.profile_image,
            'departments': [{'id': d.id, 'name': d.name} for d in emp.departments] if emp.departments else []
        }
        employees_data.append(emp_dict)
    
    # جلب جميع الأقسام للفلترة
    departments = Department.query.all()
    
    return render_template(
        'employees/tracking.html',
        employees=employees_data,
        employee_locations=employee_locations,
        departments=departments,
        department_filter=department_filter,
        search_query=search_query
    )




@employees_bp.route('/<int:id>/track-history')
@login_required
def track_history(id):
    """صفحة تتبع تحركات موظف واحد خلال 24 ساعة"""
    from flask import jsonify
    
    employee = Employee.query.get_or_404(id)
    
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    
    locations = EmployeeLocation.query.filter(
        EmployeeLocation.employee_id == id,
        EmployeeLocation.recorded_at >= cutoff_time
    ).order_by(EmployeeLocation.recorded_at.asc()).all()
    
    locations_data = []
    for loc in locations:
        loc_dict = {
            'latitude': float(loc.latitude),
            'longitude': float(loc.longitude),
            'speed': float(loc.speed_kmh) if loc.speed_kmh else 0,
            'vehicle_id': loc.vehicle_id,
            'recorded_at': loc.recorded_at.strftime('%Y-%m-%d %H:%M:%S'),
            'accuracy': float(loc.accuracy_m) if loc.accuracy_m else None
        }
        
        if loc.vehicle_id and loc.vehicle:
            loc_dict['vehicle'] = {
                'id': loc.vehicle.id,
                'plate_number': loc.vehicle.plate_number,
                'make': loc.vehicle.make,
                'model': loc.vehicle.model
            }
        
        locations_data.append(loc_dict)
    
    departments = Department.query.all()
    
    return render_template(
        'employees/track_history.html',
        employee=employee,
        locations=locations_data,
        departments=departments
    )


@employees_bp.route('/<int:employee_id>/track-history/export-pdf')
@login_required
def export_track_history_pdf(employee_id):
    """تصدير سجل التحركات إلى PDF مع روابط قابلة للنقر"""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    import requests
    
    employee = Employee.query.get_or_404(employee_id)
    
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    locations = EmployeeLocation.query.filter(
        EmployeeLocation.employee_id == employee_id,
        EmployeeLocation.recorded_at >= cutoff_time
    ).order_by(EmployeeLocation.recorded_at.asc()).all()
    
    pdfmetrics.registerFont(TTFont('Cairo', 'static/fonts/Cairo.ttf'))
    pdfmetrics.registerFont(TTFont('CairoReg', 'static/fonts/Cairo.ttf'))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    story = []
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=getSampleStyleSheet()['Heading1'],
        fontName='Cairo',
        fontSize=20,
        textColor=colors.HexColor('#1e1b4b'),
        alignment=TA_CENTER,
        spaceAfter=20,
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=getSampleStyleSheet()['Normal'],
        fontName='CairoReg',
        fontSize=11,
        alignment=TA_RIGHT,
        rightIndent=0,
        leftIndent=0,
        textColor=colors.HexColor('#374151'),
    )
    
    title_text = get_display(reshape(f"سجل تحركات الموظف - {employee.name}"))
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 0.3*cm))
    
    info_data = [
        [get_display(reshape('رقم الموظف:')), get_display(reshape(str(employee.employee_id)))],
        [get_display(reshape('الاسم:')), get_display(reshape(employee.name))],
        [get_display(reshape('عدد النقاط:')), str(len(locations))],
        [get_display(reshape('التاريخ:')), datetime.now().strftime('%Y-%m-%d %H:%M')],
    ]
    
    if employee.departments:
        info_data.insert(2, [get_display(reshape('القسم:')), get_display(reshape(employee.departments[0].name))])
    
    info_table = Table(info_data, colWidths=[5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'CairoReg', 11),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#312e81')),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c7d2fe')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f5f7ff')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 0.8*cm))
    
    if locations and len(locations) > 0:
        max_speed = max([float(loc.speed_kmh) if loc.speed_kmh else 0 for loc in locations])
        total_distance = 0
        vehicle_count = sum([1 for loc in locations if loc.vehicle_id])
        
        for i in range(1, len(locations)):
            prev = locations[i-1]
            curr = locations[i]
            lat1, lon1 = float(prev.latitude), float(prev.longitude)
            lat2, lon2 = float(curr.latitude), float(curr.longitude)
            
            from math import radians, sin, cos, sqrt, atan2
            R = 6371
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
            c = 2 * atan2(sqrt(a), sqrt(1-a))
            total_distance += R * c
        
        stats_data = [
            [get_display(reshape('إجمالي المسافة:')), f"{total_distance:.2f} " + get_display(reshape('كم'))],
            [get_display(reshape('أقصى سرعة:')), f"{max_speed:.1f} " + get_display(reshape('كم/س'))],
            [get_display(reshape('عدد النقاط على سيارة:')), str(vehicle_count)],
        ]
        
        stats_table = Table(stats_data, colWidths=[5*cm, 10*cm])
        stats_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'CairoReg', 11),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dbeafe')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e3a8a')),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#93c5fd')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 0.8*cm))
        
        subtitle = get_display(reshape('سجل التحركات التفصيلي'))
        story.append(Paragraph(subtitle, title_style))
        story.append(Spacer(1, 0.3*cm))
        
        data = [[
            get_display(reshape('الإحداثيات')),
            get_display(reshape('السرعة')),
            get_display(reshape('السيارة')),
            get_display(reshape('الوقت')),
            get_display(reshape('#'))
        ]]
        
        for idx, loc in enumerate(locations, 1):
            coords = f"{float(loc.latitude):.6f}, {float(loc.longitude):.6f}"
            coords_link = f'<link href="https://www.google.com/maps?q={float(loc.latitude)},{float(loc.longitude)}" color="blue">{coords}</link>'
            
            speed = f"{float(loc.speed_kmh):.1f} " + get_display(reshape('كم/س')) if loc.speed_kmh and float(loc.speed_kmh) > 0 else "-"
            
            vehicle_info = "-"
            if loc.vehicle_id and loc.vehicle:
                vehicle_info = get_display(reshape(f"{loc.vehicle.plate_number} - {loc.vehicle.make}"))
            
            time_str = loc.recorded_at.strftime('%Y-%m-%d %H:%M:%S')
            
            data.append([
                Paragraph(coords_link, normal_style),
                get_display(reshape(speed)),
                vehicle_info,
                time_str,
                str(idx)
            ])
        
        table = Table(data, colWidths=[6*cm, 2.5*cm, 4*cm, 4*cm, 1*cm])
        table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Cairo', 12),
            ('FONT', (0, 1), (-1, -1), 'CairoReg', 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c7d2fe')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7ff')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(table)
    else:
        no_data_text = get_display(reshape('لا توجد بيانات تتبع خلال آخر 24 ساعة'))
        story.append(Paragraph(no_data_text, normal_style))
    
    doc.build(story)
    buffer.seek(0)
    
    filename = f"track_history_{employee.employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


@employees_bp.route('/<int:employee_id>/track-history/export-excel')
@login_required
def export_track_history_excel(employee_id):
    """تصدير سجل التحركات إلى Excel مع صورة الخريطة"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import requests
    from io import BytesIO
    
    employee = Employee.query.get_or_404(employee_id)
    
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    locations = EmployeeLocation.query.filter(
        EmployeeLocation.employee_id == employee_id,
        EmployeeLocation.recorded_at >= cutoff_time
    ).order_by(EmployeeLocation.recorded_at.asc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "سجل التحركات"
    
    ws.right_to_left = True
    
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    title_font = Font(name='Arial', size=18, bold=True, color='1E1B4B')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"سجل تحركات الموظف - {employee.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:B2')
    ws['A2'] = "معلومات الموظف"
    ws['A2'].font = Font(name='Arial', size=12, bold=True, color='312E81')
    ws['A2'].fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws['A3'] = "رقم الموظف:"
    ws['B3'] = employee.employee_id
    ws['A4'] = "الاسم:"
    ws['B4'] = employee.name
    
    if employee.departments:
        ws['A5'] = "القسم:"
        ws['B5'] = employee.departments[0].name
        start_row = 6
    else:
        start_row = 5
    
    ws[f'A{start_row}'] = "عدد النقاط:"
    ws[f'B{start_row}'] = len(locations)
    ws[f'A{start_row+1}'] = "تاريخ التقرير:"
    ws[f'B{start_row+1}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    if locations and len(locations) > 0:
        max_speed = max([float(loc.speed_kmh) if loc.speed_kmh else 0 for loc in locations])
        total_distance = 0
        vehicle_count = sum([1 for loc in locations if loc.vehicle_id])
        
        for i in range(1, len(locations)):
            prev = locations[i-1]
            curr = locations[i]
            lat1, lon1 = float(prev.latitude), float(prev.longitude)
            lat2, lon2 = float(curr.latitude), float(curr.longitude)
            
            from math import radians, sin, cos, sqrt, atan2
            R = 6371
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
            c = 2 * atan2(sqrt(a), sqrt(1-a))
            total_distance += R * c
        
        stats_row = start_row + 3
        ws.merge_cells(f'A{stats_row}:B{stats_row}')
        ws[f'A{stats_row}'] = "إحصائيات التحركات"
        ws[f'A{stats_row}'].font = Font(name='Arial', size=12, bold=True, color='1E3A8A')
        ws[f'A{stats_row}'].fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        ws[f'A{stats_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'A{stats_row+1}'] = "إجمالي المسافة:"
        ws[f'B{stats_row+1}'] = f"{total_distance:.2f} كم"
        ws[f'A{stats_row+2}'] = "أقصى سرعة:"
        ws[f'B{stats_row+2}'] = f"{max_speed:.1f} كم/س"
        ws[f'A{stats_row+3}'] = "عدد النقاط على سيارة:"
        ws[f'B{stats_row+3}'] = vehicle_count
        
        table_start = stats_row + 5
        
        if len(locations) > 0:
            center_lat = sum([float(loc.latitude) for loc in locations]) / len(locations)
            center_lon = sum([float(loc.longitude) for loc in locations]) / len(locations)
            
            try:
                map_url = f"https://maps.geoapify.com/v1/staticmap?style=osm-bright&width=600&height=400&center=lonlat:{center_lon},{center_lat}&zoom=12&marker=lonlat:{center_lon},{center_lat};color:%234f46e5;size:medium&apiKey=YOUR_API_KEY"
                
                lats = [float(loc.latitude) for loc in locations]
                lons = [float(loc.longitude) for loc in locations]
                min_lat, max_lat = min(lats), max(lats)
                min_lon, max_lon = min(lons), max(lons)
                
                map_url = f"https://api.mapbox.com/styles/v1/mapbox/satellite-streets-v11/static/[{min_lon},{min_lat},{max_lon},{max_lat}]/600x400?access_token=pk.YOUR_TOKEN"
                
            except:
                pass
        
        headers = ['#', 'الوقت', 'خط العرض', 'خط الطول', 'السرعة (كم/س)', 'السيارة', 'رابط الموقع']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for idx, loc in enumerate(locations, 1):
            row = table_start + idx
            
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = loc.recorded_at.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row, column=3).value = float(loc.latitude)
            ws.cell(row=row, column=4).value = float(loc.longitude)
            ws.cell(row=row, column=5).value = f"{float(loc.speed_kmh):.1f}" if loc.speed_kmh and float(loc.speed_kmh) > 0 else "-"
            
            if loc.vehicle_id and loc.vehicle:
                ws.cell(row=row, column=6).value = f"{loc.vehicle.plate_number} - {loc.vehicle.make} {loc.vehicle.model}"
            else:
                ws.cell(row=row, column=6).value = "-"
            
            maps_link = f"https://www.google.com/maps?q={float(loc.latitude)},{float(loc.longitude)}"
            ws.cell(row=row, column=7).value = maps_link
            ws.cell(row=row, column=7).hyperlink = maps_link
            ws.cell(row=row, column=7).font = Font(color='0000FF', underline='single')
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='C7D2FE'),
                    right=Side(style='thin', color='C7D2FE'),
                    top=Side(style='thin', color='C7D2FE'),
                    bottom=Side(style='thin', color='C7D2FE')
                )
                
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F5F7FF', end_color='F5F7FF', fill_type='solid')
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"track_history_{employee.employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
